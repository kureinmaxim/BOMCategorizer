# -*- coding: utf-8 -*-
"""
Тестовый скрипт для проверки повторения заголовков категорий в PDF
"""

import os
from bom_categorizer.pdf_exporter import PDFExporter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_test_excel():
    """Создает тестовый Excel файл с большой таблицей"""
    wb = Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Создаем лист SUMMARY
    summary_sheet = wb.create_sheet("SUMMARY")
    summary_sheet.append(["Категория", "Кол-во позиций", "Кол-во компонентов"])
    summary_sheet.append(["КОНДЕНСАТОРЫ", 50, 150])
    summary_sheet.append(["РЕЗИСТОРЫ", 30, 100])
    
    # Создаем лист SOURCES
    sources_sheet = wb.create_sheet("SOURCES")
    sources_sheet.append(["source_file", "source_sheet"])
    sources_sheet.append(["test_file.xlsx", "Sheet1"])
    
    # Создаем лист КОНДЕНСАТОРЫ с БОЛЬШИМ количеством строк (для проверки переноса на новую страницу)
    cond_sheet = wb.create_sheet("КОНДЕНСАТОРЫ")
    
    # Заголовок
    headers = ["№ п/п", "Наименование ИВП", "ТУ", "Источник", "Кол-во", "Примечание", "№ ТРУ", "Стоимость"]
    cond_sheet.append(headers)
    
    # Форматируем заголовок
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in cond_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Добавляем МНОГО строк данных (чтобы таблица перешла на следующую страницу)
    for i in range(1, 51):  # 50 строк данных
        cond_sheet.append([
            i,
            f"Конденсатор К10-17б-М1500-{i*10}В-{i*100}пФ±5%",
            "ТУ 3.456.789-12",
            "test_file.xlsx",
            i,
            f"Примечание {i}",
            f"ТРУ-{i:04d}",
            f"{i*10.5:.2f}"
        ])
    
    # Создаем лист РЕЗИСТОРЫ
    res_sheet = wb.create_sheet("РЕЗИСТОРЫ")
    res_sheet.append(headers)
    
    # Форматируем заголовок
    for cell in res_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Добавляем данные
    for i in range(1, 31):  # 30 строк данных
        res_sheet.append([
            i,
            f"Резистор С2-33Н-0,125Вт-{i*100}Ом±5%",
            "ТУ 1.234.567-89",
            "test_file.xlsx",
            i,
            f"Примечание {i}",
            f"ТРУ-{i:04d}",
            f"{i*5.5:.2f}"
        ])
    
    # Сохраняем
    test_file = "test_pdf_output.xlsx"
    wb.save(test_file)
    print(f"✓ Создан тестовый Excel: {test_file}")
    return test_file


def test_pdf_export():
    """Тестирует экспорт в PDF"""
    print("=" * 70)
    print("Тест экспорта в PDF с повторением заголовков категорий")
    print("=" * 70)
    print()
    
    # Создаем тестовый Excel
    excel_file = create_test_excel()
    
    # Экспортируем в PDF
    print("\nЭкспорт в PDF...")
    exporter = PDFExporter()
    
    pdf_file = excel_file.replace('.xlsx', '.pdf')
    result = exporter.export_excel_to_pdf(excel_file, pdf_file)
    
    print(f"\n✓ PDF создан: {result}")
    print()
    print("=" * 70)
    print("ПРОВЕРКА:")
    print("=" * 70)
    print("1. Откройте файл:", result)
    print("2. Найдите страницу, где таблица КОНДЕНСАТОРЫ переходит на новый лист")
    print("3. Проверьте, что на новой странице есть:")
    print("   ✓ Заголовок категории 'КОНДЕНСАТОРЫ' (синий, по центру)")
    print("   ✓ Заголовки столбцов (№ п/п, Наименование ИВП, и т.д.)")
    print("4. Заголовок категории должен повторяться на КАЖДОЙ странице таблицы")
    print("=" * 70)
    
    # Проверяем наличие файла
    if os.path.exists(result):
        file_size = os.path.getsize(result) / 1024  # В KB
        print(f"\n✓ Файл существует, размер: {file_size:.2f} KB")
        
        # Пробуем открыть (только Windows)
        import platform
        if platform.system() == 'Windows':
            import subprocess
            try:
                subprocess.Popen([result], shell=True)
                print("✓ PDF открыт в программе просмотра")
            except Exception as e:
                print(f"⚠ Не удалось открыть PDF автоматически: {e}")
                print(f"  Откройте вручную: {os.path.abspath(result)}")
    
    print()
    return result


if __name__ == '__main__':
    test_pdf_export()

