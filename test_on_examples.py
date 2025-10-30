"""
Скрипт для тестирования проекта на реальных файлах из example/

Использование:
    python test_on_examples.py                      # Протестировать все файлы
    python test_on_examples.py plata_MKVH.doc      # Протестировать конкретный файл
    python test_on_examples.py --output test_results/  # Указать папку для результатов
"""

import sys
import os
from pathlib import Path
import argparse
from datetime import datetime
import subprocess

# Добавляем путь к модулю
sys.path.insert(0, str(Path(__file__).parent))


def test_file(input_path: Path, output_dir: Path, verbose: bool = True):
    """
    Тестирует один файл
    
    Args:
        input_path: Путь к входному файлу
        output_dir: Директория для выходных файлов
        verbose: Выводить подробную информацию
    
    Returns:
        True если успешно, False если ошибка
    """
    if not input_path.exists():
        print(f"❌ Файл не найден: {input_path}")
        return False
    
    # Создаем имя выходного файла
    output_name = f"{input_path.stem}_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = output_dir / output_name
    
    print(f"\n{'='*80}")
    print(f"📄 Тестирование: {input_path.name}")
    print(f"{'='*80}")
    
    try:
        # Запускаем обработку через split_bom.py
        print(f"⏳ Обработка файла...")
        
        # Формируем команду
        cmd = [
            sys.executable,
            "split_bom.py",
            "--inputs", str(input_path),
            "--xlsx", str(output_path),
            "--combine",
            "--no-interactive"
        ]
        
        # Запускаем команду
        process = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent
        )
        
        if process.returncode != 0:
            print(f"❌ Ошибка выполнения команды")
            if verbose:
                print(f"STDOUT:\n{process.stdout}")
                print(f"STDERR:\n{process.stderr}")
            return False
        
        if output_path.exists():
            # Проверяем результат
            from openpyxl import load_workbook
            wb = load_workbook(str(output_path))
            
            print(f"✅ Файл успешно обработан!")
            print(f"   Выходной файл: {output_path}")
            print(f"   Размер: {output_path.stat().st_size / 1024:.1f} KB")
            print(f"   Листов создано: {len(wb.sheetnames)}")
            
            if verbose:
                print(f"\n📊 Листы в выходном файле:")
                for i, sheet_name in enumerate(wb.sheetnames, 1):
                    ws = wb[sheet_name]
                    row_count = ws.max_row - 1  # -1 для заголовка
                    print(f"   {i}. {sheet_name:<30} - {row_count} строк")
            
            wb.close()
            return True
        else:
            print(f"❌ Выходной файл не создан")
            return False
            
    except Exception as e:
        print(f"❌ Ошибка при обработке: {e}")
        if verbose:
            import traceback
            traceback.print_exc()
        return False


def main():
    parser = argparse.ArgumentParser(
        description='Тестирование BOM Categorizer на примерах',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        'files',
        nargs='*',
        help='Файлы для тестирования (если не указано - все из example/)'
    )
    parser.add_argument(
        '--output',
        default='test_output',
        help='Папка для выходных файлов (по умолчанию: test_output/)'
    )
    parser.add_argument(
        '--verbose',
        '-v',
        action='store_true',
        help='Подробный вывод'
    )
    parser.add_argument(
        '--example-dir',
        default='example',
        help='Папка с примерами (по умолчанию: example/)'
    )
    
    args = parser.parse_args()
    
    # Определяем директории
    project_root = Path(__file__).parent
    example_dir = project_root / args.example_dir
    output_dir = project_root / args.output
    
    # Создаем выходную директорию
    output_dir.mkdir(exist_ok=True)
    
    print("=" * 80)
    print("🧪 BOM Categorizer - Тестирование на примерах")
    print("=" * 80)
    print(f"Папка с примерами: {example_dir}")
    print(f"Папка для результатов: {output_dir}")
    
    # Определяем файлы для тестирования
    if args.files:
        # Конкретные файлы
        test_files = []
        for file_arg in args.files:
            file_path = Path(file_arg)
            if not file_path.is_absolute():
                # Пробуем относительно example/
                file_path = example_dir / file_arg
            
            if file_path.exists():
                test_files.append(file_path)
            else:
                print(f"⚠️ Файл не найден: {file_arg}")
    else:
        # Все поддерживаемые файлы из example/
        if not example_dir.exists():
            print(f"❌ Папка с примерами не найдена: {example_dir}")
            return 1
        
        test_files = []
        for pattern in ['*.doc', '*.docx', '*.xlsx', '*.txt']:
            test_files.extend(example_dir.glob(pattern))
        
        # Исключаем файлы из подпапок debug, final и т.д.
        test_files = [f for f in test_files if f.parent == example_dir]
    
    if not test_files:
        print("❌ Не найдено файлов для тестирования")
        return 1
    
    print(f"\n📋 Найдено файлов для тестирования: {len(test_files)}")
    
    # Тестируем каждый файл
    results = []
    for i, file_path in enumerate(test_files, 1):
        print(f"\n[{i}/{len(test_files)}]")
        success = test_file(file_path, output_dir, args.verbose)
        results.append((file_path.name, success))
    
    # Итоговая статистика
    print(f"\n{'='*80}")
    print("📊 ИТОГИ ТЕСТИРОВАНИЯ")
    print(f"{'='*80}")
    
    successful = sum(1 for _, success in results if success)
    failed = len(results) - successful
    
    print(f"\nВсего протестировано: {len(results)}")
    print(f"✅ Успешно: {successful}")
    print(f"❌ Ошибок: {failed}")
    
    if failed == 0:
        print(f"\n🎉 Все тесты пройдены успешно!")
    else:
        print(f"\n⚠️ Некоторые тесты завершились с ошибками:")
        for name, success in results:
            if not success:
                print(f"   - {name}")
    
    print(f"\n📁 Результаты сохранены в: {output_dir}")
    print("=" * 80)
    
    return 0 if failed == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
