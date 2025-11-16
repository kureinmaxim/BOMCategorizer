# -*- coding: utf-8 -*-
"""
Экспорт результатов обработки BOM в PDF формат

Конвертирует Excel файлы с результатами в PDF документы.
Поддерживает сохранение таблиц, форматирования и стилей.
"""

import os
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT


class PDFExporter:
    """Класс для экспорта BOM данных в PDF"""
    
    def __init__(self):
        self.page_size = landscape(A4)
        self.styles = getSampleStyleSheet()
        self.cyrillic_font = 'DejaVuSans'
        self.cyrillic_font_bold = 'DejaVuSans-Bold'
        self._register_fonts()
    
    def _register_fonts(self):
        """Регистрирует шрифты с поддержкой кириллицы"""
        import platform
        import sys
        
        fonts_registered = False
        error_messages = []
        
        try:
            system = platform.system()
            
            # Приоритет 1: Шрифты из папки приложения (для инсталлятора)
            # Ищем папку fonts рядом с исполняемым файлом или скриптом
            if not fonts_registered:
                try:
                    # Определяем базовую директорию приложения
                    if getattr(sys, 'frozen', False):
                        # Если запущено из exe (PyInstaller)
                        base_dir = os.path.dirname(sys.executable)
                    else:
                        # Если запущено как скрипт
                        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                    
                    fonts_dir = os.path.join(base_dir, 'fonts')
                    
                    if os.path.exists(fonts_dir):
                        dejavu_paths = {
                            'DejaVuSans': os.path.join(fonts_dir, 'DejaVuSans.ttf'),
                            'DejaVuSans-Bold': os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf'),
                        }
                        
                        if all(os.path.exists(p) for p in dejavu_paths.values()):
                            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_paths['DejaVuSans']))
                            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_paths['DejaVuSans-Bold']))
                            self.cyrillic_font = 'DejaVuSans'
                            self.cyrillic_font_bold = 'DejaVuSans-Bold'
                            fonts_registered = True
                            print(f"✓ Зарегистрированы шрифты: DejaVuSans из папки приложения (поддержка кириллицы)")
                        else:
                            error_messages.append(f"DejaVuSans: Папка fonts найдена, но файлы шрифтов отсутствуют")
                    else:
                        error_messages.append(f"DejaVuSans: Папка fonts не найдена в {base_dir}")
                
                except Exception as e:
                    error_messages.append(f"DejaVuSans (папка приложения): {str(e)}")
            
            # Приоритет 2: Для Windows - пробуем Arial
            if not fonts_registered and system == 'Windows':
                font_paths = {
                    'Arial': 'C:/Windows/Fonts/arial.ttf',
                    'Arial-Bold': 'C:/Windows/Fonts/arialbd.ttf',
                }
                
                all_exist = all(os.path.exists(path) for path in font_paths.values())
                if all_exist:
                    try:
                        pdfmetrics.registerFont(TTFont('Arial', font_paths['Arial']))
                        pdfmetrics.registerFont(TTFont('Arial-Bold', font_paths['Arial-Bold']))
                        self.cyrillic_font = 'Arial'
                        self.cyrillic_font_bold = 'Arial-Bold'
                        fonts_registered = True
                        print(f"✓ Зарегистрированы шрифты: Arial (поддержка кириллицы)")
                    except Exception as e:
                        error_messages.append(f"Arial: {str(e)}")
                else:
                    missing = [k for k, v in font_paths.items() if not os.path.exists(v)]
                    error_messages.append(f"Arial: Файлы не найдены - {missing}")
            
            # Приоритет 3: Для macOS - пробуем системные шрифты
            elif system == 'Darwin':
                font_paths = {
                    'Arial': '/System/Library/Fonts/Supplemental/Arial.ttf',
                    'Arial-Bold': '/System/Library/Fonts/Supplemental/Arial Bold.ttf',
                }
                
                all_exist = all(os.path.exists(path) for path in font_paths.values())
                if all_exist:
                    try:
                        pdfmetrics.registerFont(TTFont('Arial', font_paths['Arial']))
                        pdfmetrics.registerFont(TTFont('Arial-Bold', font_paths['Arial-Bold']))
                        self.cyrillic_font = 'Arial'
                        self.cyrillic_font_bold = 'Arial-Bold'
                        fonts_registered = True
                        print(f"✓ Зарегистрированы шрифты: Arial (поддержка кириллицы)")
                    except Exception as e:
                        error_messages.append(f"Arial: {str(e)}")
                else:
                    missing = [k for k, v in font_paths.items() if not os.path.exists(v)]
                    error_messages.append(f"Arial: Файлы не найдены - {missing}")
            
            # Приоритет 4: пробуем DejaVu в системных папках
            if not fonts_registered:
                try:
                    # Пробуем найти DejaVu в разных местах
                    dejavu_paths = [
                        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
                        'C:/Windows/Fonts/DejaVuSans.ttf',  # Windows (если установлен)
                        '/System/Library/Fonts/DejaVuSans.ttf',  # macOS (если установлен)
                    ]
                    
                    dejavu_bold_paths = [
                        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
                        'C:/Windows/Fonts/DejaVuSans-Bold.ttf',
                        '/System/Library/Fonts/DejaVuSans-Bold.ttf',
                    ]
                    
                    dejavu_normal = next((p for p in dejavu_paths if os.path.exists(p)), None)
                    dejavu_bold = next((p for p in dejavu_bold_paths if os.path.exists(p)), None)
                    
                    if dejavu_normal and dejavu_bold:
                        pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_normal))
                        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold))
                        self.cyrillic_font = 'DejaVuSans'
                        self.cyrillic_font_bold = 'DejaVuSans-Bold'
                        fonts_registered = True
                        print(f"✓ Зарегистрированы шрифты: DejaVuSans (поддержка кириллицы)")
                    else:
                        error_messages.append(f"DejaVuSans: Файлы не найдены в стандартных путях")
                
                except Exception as e:
                    error_messages.append(f"DejaVuSans: {str(e)}")
            
            # Приоритет 5: пробуем загрузить DejaVu из папки проекта или pip пакета
            if not fonts_registered:
                try:
                    # Пробуем найти шрифты в папке проекта или site-packages reportlab
                    import reportlab
                    reportlab_dir = os.path.dirname(reportlab.__file__)
                    fonts_dir = os.path.join(reportlab_dir, 'fonts')
                    
                    # Проверяем разные возможные пути
                    possible_paths = [
                        os.path.join(fonts_dir, 'DejaVuSans.ttf'),
                        os.path.join(reportlab_dir, 'lib', 'fonts', 'DejaVuSans.ttf'),
                    ]
                    
                    possible_bold_paths = [
                        os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf'),
                        os.path.join(reportlab_dir, 'lib', 'fonts', 'DejaVuSans-Bold.ttf'),
                    ]
                    
                    dejavu_normal = next((p for p in possible_paths if os.path.exists(p)), None)
                    dejavu_bold = next((p for p in possible_bold_paths if os.path.exists(p)), None)
                    
                    if dejavu_normal and dejavu_bold:
                        pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_normal))
                        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold))
                        self.cyrillic_font = 'DejaVuSans'
                        self.cyrillic_font_bold = 'DejaVuSans-Bold'
                        fonts_registered = True
                        print(f"✓ Зарегистрированы шрифты: DejaVuSans из reportlab (поддержка кириллицы)")
                    else:
                        error_messages.append(f"DejaVuSans: Не найдены в папке reportlab")
                
                except Exception as e:
                    error_messages.append(f"DejaVuSans (reportlab): {str(e)}")
            
            # Приоритет 6: пробуем Times New Roman для Windows
            if not fonts_registered and system == 'Windows':
                try:
                    times_paths = {
                        'Times': 'C:/Windows/Fonts/times.ttf',
                        'Times-Bold': 'C:/Windows/Fonts/timesbd.ttf',
                    }
                    
                    all_exist = all(os.path.exists(path) for path in times_paths.values())
                    if all_exist:
                        pdfmetrics.registerFont(TTFont('TimesNewRoman', times_paths['Times']))
                        pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', times_paths['Times-Bold']))
                        self.cyrillic_font = 'TimesNewRoman'
                        self.cyrillic_font_bold = 'TimesNewRoman-Bold'
                        fonts_registered = True
                        print(f"✓ Зарегистрированы шрифты: Times New Roman (поддержка кириллицы)")
                    else:
                        error_messages.append(f"Times New Roman: Файлы не найдены")
                
                except Exception as e:
                    error_messages.append(f"Times New Roman: {str(e)}")
            
            # Если ничего не получилось - выводим предупреждение и используем Helvetica
            if not fonts_registered:
                print("\n" + "="*70)
                print("⚠️  ВНИМАНИЕ: Не удалось зарегистрировать шрифты с поддержкой кириллицы!")
                print("="*70)
                print("Кириллические символы в PDF будут отображаться некорректно (квадратиками).")
                print("\nПричины:")
                for msg in error_messages:
                    print(f"  - {msg}")
                print("\nРЕКОМЕНДАЦИИ:")
                print("  1. Установите шрифт DejaVu Sans:")
                print("     • Windows: Скачайте с https://dejavu-fonts.github.io/")
                print("       и установите файлы DejaVuSans.ttf и DejaVuSans-Bold.ttf")
                print("       в папку C:\\Windows\\Fonts\\")
                print("  2. Или установите пакет: pip install reportlab[rlpycairo]")
                print("  3. Перезапустите программу после установки шрифтов")
                print("="*70 + "\n")
                
                self.cyrillic_font = 'Helvetica'
                self.cyrillic_font_bold = 'Helvetica-Bold'
                    
        except Exception as e:
            # Критическая ошибка - используем стандартные
            print(f"⚠️  Критическая ошибка при регистрации шрифтов: {e}")
            self.cyrillic_font = 'Helvetica'
            self.cyrillic_font_bold = 'Helvetica-Bold'
    
    def export_excel_to_pdf(self, excel_path: str, pdf_path: Optional[str] = None) -> str:
        """
        Экспортирует Excel файл в PDF
        
        Args:
            excel_path: Путь к Excel файлу
            pdf_path: Путь к выходному PDF (если None, создается автоматически)
        
        Returns:
            Путь к созданному PDF файлу
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Файл не найден: {excel_path}")
        
        # Определяем путь к PDF
        if not pdf_path:
            base_name = os.path.splitext(excel_path)[0]
            pdf_path = f"{base_name}.pdf"
        
        # Загружаем Excel
        wb = load_workbook(excel_path, data_only=True)
        
        # Функция для добавления номеров страниц
        def add_page_number(canvas, doc):
            """Добавляет номер страницы внизу каждой страницы"""
            page_num = canvas.getPageNumber()
            text = f"Страница {page_num}"
            canvas.saveState()
            canvas.setFont(self.cyrillic_font, 9)
            canvas.setFillColor(colors.HexColor('#6b7280'))
            canvas.drawCentredString(self.page_size[0] / 2, 15, text)
            canvas.restoreState()
        
        # Создаем PDF документ с минимальными отступами
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=self.page_size,
            rightMargin=10*mm,
            leftMargin=10*mm,
            topMargin=8*mm,   # Уменьшен отступ сверху
            bottomMargin=12*mm  # Увеличен для размещения номера страницы
        )
        
        # Элементы документа
        story = []
        
        # Получаем список источников из листа SOURCES для заголовка
        sources_list = []
        if 'SOURCES' in wb.sheetnames:
            sources_sheet = wb['SOURCES']
            for row in sources_sheet.iter_rows(min_row=2, max_row=sources_sheet.max_row, values_only=True):
                if row and row[0]:  # Если есть данные в первой колонке
                    sources_list.append(str(row[0]))
        
        # Заголовок документа с источниками в скобках
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontName=self.cyrillic_font_bold,
            fontSize=11,  # Уменьшен для размещения источников
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=8,
            alignment=TA_CENTER
        )
        
        # Формируем заголовок с источниками
        if sources_list:
            sources_text = ", ".join(sources_list)
            title = Paragraph(f"<b>Отчет по обработке BOM ({sources_text})</b>", title_style)
        else:
            title = Paragraph(f"<b>Отчет по обработке BOM</b>", title_style)
        story.append(title)
        story.append(Spacer(1, 3*mm))
        
        # Добавляем информацию о дате создания и количестве листов
        from datetime import datetime
        info_style = ParagraphStyle(
            'Info',
            parent=self.styles['Normal'],
            fontName=self.cyrillic_font,
            fontSize=9,
            textColor=colors.HexColor('#4b5563'),
            spaceAfter=2,
            alignment=TA_CENTER
        )
        
        # Подсчитываем общее количество позиций из всех листов (кроме SUMMARY и SOURCES)
        total_items = 0
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() not in ['SUMMARY', 'SOURCES']:
                sheet = wb[sheet_name]
                # Считаем строки с данными (минус заголовок)
                if sheet.max_row > 1:
                    total_items += sheet.max_row - 1
        
        date_str = datetime.now().strftime('%d.%m.%Y %H:%M')
        story.append(Paragraph(f"<b>Дата создания:</b> {date_str}", info_style))
        story.append(Paragraph(f"<b>Категорий:</b> {len(wb.sheetnames) - 2} | <b>Всего позиций:</b> {total_items}", info_style))
        story.append(Spacer(1, 5*mm))
        
        # Сначала выводим SUMMARY и SOURCES на первом листе
        priority_sheets = ['SUMMARY', 'SOURCES']
        processed_sheets = []
        
        for sheet_name in priority_sheets:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                processed_sheets.append(sheet_name)
                
                # Заголовок листа (компактный)
                sheet_title_style = ParagraphStyle(
                    'SheetTitle',
                    parent=self.styles['Heading2'],
                    fontName=self.cyrillic_font_bold,
                    fontSize=10,
                    textColor=colors.HexColor('#2563eb'),
                    spaceAfter=2,
                    spaceBefore=0 if sheet_name == 'SUMMARY' else 3,
                    alignment=TA_LEFT,
                    keepWithNext=True
                )
                
                sheet_title = Paragraph(f"<b>{sheet_name}</b>", sheet_title_style)
                story.append(sheet_title)
                story.append(Spacer(1, 1.5*mm))
                
                # Получаем данные из листа
                data = self._get_sheet_data(sheet)
                
                if data:
                    # Создаем таблицу (компактный режим)
                    table = self._create_table(data, sheet, is_compact=True)
                    story.append(table)
                else:
                    empty_style = ParagraphStyle('Empty', parent=self.styles['Normal'], fontName=self.cyrillic_font)
                    empty_text = Paragraph("<i>Лист пуст</i>", empty_style)
                    story.append(empty_text)
                
                # Небольшой отступ
                story.append(Spacer(1, 4*mm))
        
        # После SUMMARY и SOURCES добавляем разрыв страницы
        if processed_sheets:
            story.append(PageBreak())
        
        # Обрабатываем остальные листы (кроме SUMMARY и SOURCES)
        for idx, sheet_name in enumerate(wb.sheetnames):
            if sheet_name in processed_sheets:
                continue
            
            sheet = wb[sheet_name]
            
            # Получаем данные из листа
            data = self._get_sheet_data(sheet)
            
            if data:
                # Создаем таблицу с заголовком категории внутри (обычный режим)
                table = self._create_table(data, sheet, is_compact=False, sheet_name=sheet_name)
                story.append(table)
            else:
                # Если лист пустой - показываем заголовок отдельно
                sheet_title_style = ParagraphStyle(
                    'SheetTitle',
                    parent=self.styles['Heading2'],
                    fontName=self.cyrillic_font_bold,
                    fontSize=12,
                    textColor=colors.HexColor('#2563eb'),
                    spaceAfter=3,
                    spaceBefore=5,
                    alignment=TA_LEFT
                )
                sheet_title = Paragraph(f"<b>{sheet_name}</b>", sheet_title_style)
                story.append(sheet_title)
                story.append(Spacer(1, 2*mm))
                
                empty_style = ParagraphStyle('Empty', parent=self.styles['Normal'], fontName=self.cyrillic_font)
                empty_text = Paragraph("<i>Лист пуст</i>", empty_style)
                story.append(empty_text)
            
            # Отступ между листами
            story.append(Spacer(1, 6*mm))
        
        # Сборка PDF с нумерацией страниц
        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        return pdf_path
    
    def _get_sheet_data(self, sheet: Worksheet, max_rows: int = 1000) -> List[List]:
        """Извлекает данные из листа Excel"""
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import ParagraphStyle
        
        data = []
        
        # Определяем границы данных
        max_row = min(sheet.max_row, max_rows)
        max_col = sheet.max_column
        
        if max_row == 0 or max_col == 0:
            return data
        
        # Проверяем, это SUMMARY или SOURCES для компактности
        sheet_name = sheet.title if hasattr(sheet, 'title') else ''
        is_compact = sheet_name.upper() in ['SUMMARY', 'SOURCES']
        
        # Создаем стиль для ячеек с переносом
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=self.styles['Normal'],
            fontName=self.cyrillic_font,
            fontSize=5.5 if is_compact else 7,  # Уменьшен на 1
            leading=8 if is_compact else 10,  # межстрочный интервал
            wordWrap='CJK'  # перенос слов
        )
        
        cell_style_header = ParagraphStyle(
            'CellStyleHeader',
            parent=self.styles['Normal'],
            fontName=self.cyrillic_font_bold,
            fontSize=7 if is_compact else 9,
            leading=9 if is_compact else 11,
            wordWrap='CJK'
        )
        
        # Читаем данные
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col)):
            row_data = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ""
                else:
                    value = str(value)
                
                # Оборачиваем в Paragraph для автоматического переноса
                # Заголовки (первая строка) - жирным шрифтом
                if idx == 0:
                    row_data.append(Paragraph(value, cell_style_header))
                else:
                    row_data.append(Paragraph(value, cell_style))
            
            # Проверяем, не пустая ли строка
            if any(str(cell) for cell in row_data):
                data.append(row_data)
        
        return data
    
    def _create_table(self, data: List[List], sheet: Worksheet, is_compact: bool = False, sheet_name: str = None) -> Table:
        """Создает отформатированную таблицу для PDF
        
        Args:
            data: Данные таблицы
            sheet: Лист Excel
            is_compact: Флаг компактного режима (для SUMMARY/SOURCES)
            sheet_name: Название листа/категории (для повторения на каждой странице)
        """
        if not data:
            return None
        
        # Если передано имя листа и это не служебный лист - добавляем заголовок категории
        if sheet_name and sheet_name not in ['SUMMARY', 'SOURCES']:
            # Создаем строку заголовка категории
            num_cols = len(data[0]) if data else 1
            
            # Стиль для заголовка категории
            category_header_style = ParagraphStyle(
                'CategoryHeader',
                parent=self.styles['Heading2'],
                fontName=self.cyrillic_font_bold,
                fontSize=10,  # Уменьшен размер шрифта
                textColor=colors.HexColor('#2563eb'),
                alignment=TA_LEFT  # Выравнивание по левому краю
            )
            
            # Создаем строку с заголовком категории
            # Пустая ячейка в 1-й колонке (№ п/п), текст во 2-й колонке (Наименование ИВП)
            empty_cell = Paragraph('', category_header_style)
            category_cell = Paragraph(f"<b>{sheet_name}</b>", category_header_style)
            
            # Формируем строку: пустая ячейка, заголовок, затем остальные пустые
            if num_cols >= 2:
                category_row = [empty_cell, category_cell] + [empty_cell] * (num_cols - 2)
            else:
                category_row = [category_cell]  # На случай если колонка всего одна
            
            # Вставляем строку заголовка в начало данных
            data.insert(0, category_row)
        
        # Определяем ширину колонок
        num_cols = len(data[0]) if data else 0
        page_width = self.page_size[0] - 20*mm  # Вычитаем отступы
        
        # Индивидуальные ширины колонок в зависимости от содержимого
        # Предполагаем стандартную структуру BOM: № п/п, Наименование, ТУ, Источник, шт., Примечание, № ТРУ, Стоимость
        if num_cols == 8:
            # Оптимальное распределение ширины для 8 колонок
            col_widths = [
                12*mm,   # 1. № п/п (минимальная)
                62*mm,   # 2. Наименование ИВП (широкая, увеличена на 2mm)
                44*mm,   # 3. ТУ (увеличена на 2mm)
                63*mm,   # 4. Источник (увеличена на 1mm)
                10*mm,   # 5. шт. (минимальная)
                55*mm,   # 6. Примечание (увеличена на 2mm)
                17*mm,   # 7. № ТРУ (минимальная)
                20*mm    # 8. Стоимость (узкая)
            ]
        elif num_cols == 7:
            # Если 7 колонок (без стоимости или другого поля)
            col_widths = [
                15*mm,   # № п/п
                70*mm,   # Наименование ИВП
                55*mm,   # ТУ
                52*mm,   # Источник (увеличена за счет № ТРУ)
                20*mm,   # шт.
                60*mm,   # Примечание
                20*mm    # № ТРУ (минимальная)
            ]
        else:
            # Для других количеств колонок - равномерное распределение
            col_width = page_width / num_cols if num_cols > 0 else 50*mm
            
            # Ограничиваем ширину колонок
            if col_width > 60*mm:
                col_width = 60*mm
            elif col_width < 15*mm:
                col_width = 15*mm
            
            col_widths = [col_width] * num_cols
        
        # Определяем количество строк для повторения на новых страницах
        # Если есть заголовок категории - повторяем 2 строки (категория + заголовки), иначе 1
        repeat_rows = 2 if (sheet_name and sheet_name not in ['SUMMARY', 'SOURCES']) else 1
        
        # Создаем таблицу с повтором заголовка на каждой странице
        table = Table(data, colWidths=col_widths, repeatRows=repeat_rows, splitByRow=True)
        
        # Базовый стиль таблицы (компактный для SUMMARY/SOURCES)
        header_font_size = 7 if is_compact else 9
        body_font_size = 5.5 if is_compact else 7  # Уменьшен на 1
        v_padding = 2 if is_compact else 3
        h_padding = 2 if is_compact else 3
        
        # Определяем, есть ли строка заголовка категории
        has_category_header = sheet_name and sheet_name not in ['SUMMARY', 'SOURCES']
        header_row = 1 if has_category_header else 0  # Номер строки заголовков столбцов
        body_start_row = 2 if has_category_header else 1  # Начало тела таблицы
        
        style = TableStyle([
            # Границы
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ])
        
        # Если есть заголовок категории - добавляем стили для него
        if has_category_header:
            # Применяем стили ко всей строке заголовка категории
            style.add('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4FF'))  # Светло-голубой фон
            style.add('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2563eb'))
            style.add('ALIGN', (0, 0), (-1, 0), 'LEFT')  # Выравнивание по левому краю
            style.add('VALIGN', (0, 0), (-1, 0), 'MIDDLE')
            style.add('FONTNAME', (0, 0), (-1, 0), self.cyrillic_font_bold)
            style.add('FONTSIZE', (0, 0), (-1, 0), 10)  # Уменьшенный размер шрифта
            style.add('BOTTOMPADDING', (0, 0), (-1, 0), 4)  # Немного уменьшен отступ
            style.add('TOPPADDING', (0, 0), (-1, 0), 4)
            # Убираем внутренние границы в строке заголовка для эффекта объединения
            if num_cols > 1:
                style.add('LINEBELOW', (0, 0), (-1, 0), 1, colors.black)  # Нижняя граница
                style.add('LINEABOVE', (0, 0), (-1, 0), 1, colors.black)  # Верхняя граница  
                style.add('LINEBEFORE', (0, 0), (0, 0), 1, colors.black)  # Левая граница
                style.add('LINEAFTER', (-1, 0), (-1, 0), 1, colors.black)  # Правая граница
                # Убираем вертикальные границы между ячейками в строке заголовка
                for i in range(num_cols - 1):
                    style.add('LINEAFTER', (i, 0), (i, 0), 0, colors.white)
            
            # ВАЖНО: Запрещаем разрыв между заголовком категории и заголовком столбцов
            # Это гарантирует, что они останутся вместе
            style.add('SPLITFIRST', (0, 0), (-1, 1), 0)  # Не разрывать первые 2 строки
            style.add('SPLITLAST', (0, 0), (-1, 1), 0)
        
        # Стили для строки заголовков столбцов
        style.add('BACKGROUND', (0, header_row), (-1, header_row), colors.HexColor('#E0E0E0'))
        style.add('TEXTCOLOR', (0, header_row), (-1, header_row), colors.black)
        style.add('ALIGN', (0, header_row), (-1, header_row), 'CENTER')
        style.add('VALIGN', (0, header_row), (-1, header_row), 'MIDDLE')
        style.add('FONTNAME', (0, header_row), (-1, header_row), self.cyrillic_font_bold)
        style.add('FONTSIZE', (0, header_row), (-1, header_row), header_font_size)
        style.add('BOTTOMPADDING', (0, header_row), (-1, header_row), v_padding + 2)
        style.add('TOPPADDING', (0, header_row), (-1, header_row), v_padding + 2)
        
        # Стили для тела таблицы
        style.add('BACKGROUND', (0, body_start_row), (-1, -1), colors.white)
        style.add('TEXTCOLOR', (0, body_start_row), (-1, -1), colors.black)
        style.add('ALIGN', (0, body_start_row), (-1, -1), 'LEFT')
        style.add('VALIGN', (0, body_start_row), (-1, -1), 'TOP')
        style.add('FONTNAME', (0, body_start_row), (-1, -1), self.cyrillic_font)
        style.add('FONTSIZE', (0, body_start_row), (-1, -1), body_font_size)
        style.add('TOPPADDING', (0, body_start_row), (-1, -1), v_padding)
        style.add('BOTTOMPADDING', (0, body_start_row), (-1, -1), v_padding)
        style.add('LEFTPADDING', (0, 0), (-1, -1), h_padding)
        style.add('RIGHTPADDING', (0, 0), (-1, -1), h_padding)
        
        # Чередующиеся цвета строк в теле таблицы
        style.add('ROWBACKGROUNDS', (0, body_start_row), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')])
        
        # Выравнивание по центру для числовых колонок
        # Колонка № п/п (0) - центр (для всех случаев)
        style.add('ALIGN', (0, body_start_row), (0, -1), 'CENTER')
        # Колонка шт. (обычно 4) - центр
        if num_cols > 4:
            style.add('ALIGN', (4, body_start_row), (4, -1), 'CENTER')
        # Колонка Стоимость (обычно последняя) - право
        if num_cols == 8:
            style.add('ALIGN', (7, body_start_row), (7, -1), 'RIGHT')
        
        table.setStyle(style)
        
        return table
    
    def export_with_summary(self, excel_path: str, pdf_path: Optional[str] = None, 
                          summary_info: Optional[dict] = None) -> str:
        """
        Экспортирует Excel в PDF с добавлением сводной информации
        
        Args:
            excel_path: Путь к Excel файлу
            pdf_path: Путь к выходному PDF
            summary_info: Дополнительная информация для сводки
        
        Returns:
            Путь к созданному PDF файлу
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Файл не найден: {excel_path}")
        
        if not pdf_path:
            base_name = os.path.splitext(excel_path)[0]
            pdf_path = f"{base_name}_with_summary.pdf"
        
        # Загружаем Excel
        wb = load_workbook(excel_path, data_only=True)
        
        # Функция для добавления номеров страниц
        def add_page_number(canvas, doc):
            """Добавляет номер страницы внизу каждой страницы"""
            page_num = canvas.getPageNumber()
            text = f"Страница {page_num}"
            canvas.saveState()
            canvas.setFont(self.cyrillic_font, 9)
            canvas.setFillColor(colors.HexColor('#6b7280'))
            canvas.drawCentredString(self.page_size[0] / 2, 15, text)
            canvas.restoreState()
        
        # Создаем PDF документ с минимальными отступами
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=self.page_size,
            rightMargin=10*mm,
            leftMargin=10*mm,
            topMargin=8*mm,
            bottomMargin=12*mm  # Увеличен для размещения номера страницы
        )
        
        story = []
        
        # Получаем список источников из листа SOURCES для заголовка
        sources_list = []
        if 'SOURCES' in wb.sheetnames:
            sources_sheet = wb['SOURCES']
            for row in sources_sheet.iter_rows(min_row=2, max_row=sources_sheet.max_row, values_only=True):
                if row and row[0]:
                    sources_list.append(str(row[0]))
        
        # Заголовок документа с источниками в скобках
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontName=self.cyrillic_font_bold,
            fontSize=11,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=8,
            alignment=TA_CENTER
        )
        
        # Формируем заголовок с источниками
        if sources_list:
            sources_text = ", ".join(sources_list)
            title = Paragraph(f"<b>Отчет по обработке BOM ({sources_text})</b>", title_style)
        else:
            title = Paragraph(f"<b>Отчет по обработке BOM</b>", title_style)
        story.append(title)
        story.append(Spacer(1, 3*mm))
        
        # Добавляем информацию из summary_info, если есть
        from datetime import datetime
        info_style = ParagraphStyle(
            'Info',
            parent=self.styles['Normal'],
            fontName=self.cyrillic_font,
            fontSize=9,
            textColor=colors.HexColor('#4b5563'),
            spaceAfter=2,
            alignment=TA_CENTER
        )
        
        # Если есть сводная информация из GUI, показываем её
        if summary_info:
            for key, value in summary_info.items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", info_style))
        
        # Подсчитываем общее количество позиций
        total_items = 0
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() not in ['SUMMARY', 'SOURCES']:
                sheet = wb[sheet_name]
                if sheet.max_row > 1:
                    total_items += sheet.max_row - 1
        
        date_str = datetime.now().strftime('%d.%m.%Y %H:%M')
        story.append(Paragraph(f"<b>Дата создания:</b> {date_str}", info_style))
        story.append(Paragraph(f"<b>Категорий:</b> {len(wb.sheetnames) - 2} | <b>Всего позиций:</b> {total_items}", info_style))
        story.append(Spacer(1, 5*mm))
        
        # Сначала выводим SUMMARY и SOURCES на первом листе
        priority_sheets = ['SUMMARY', 'SOURCES']
        processed_sheets = []
        
        for sheet_name in priority_sheets:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                processed_sheets.append(sheet_name)
                
                # Заголовок листа (компактный)
                sheet_title_style = ParagraphStyle(
                    'SheetTitle',
                    parent=self.styles['Heading2'],
                    fontName=self.cyrillic_font_bold,
                    fontSize=10,
                    textColor=colors.HexColor('#2563eb'),
                    spaceAfter=2,
                    spaceBefore=0 if sheet_name == 'SUMMARY' else 3,
                    alignment=TA_LEFT,
                    keepWithNext=True
                )
                
                sheet_title = Paragraph(f"<b>{sheet_name}</b>", sheet_title_style)
                story.append(sheet_title)
                story.append(Spacer(1, 1.5*mm))
                
                # Получаем данные из листа
                data = self._get_sheet_data(sheet)
                
                if data:
                    # Создаем таблицу (компактный режим)
                    table = self._create_table(data, sheet, is_compact=True)
                    story.append(table)
                else:
                    empty_style = ParagraphStyle('Empty', parent=self.styles['Normal'], fontName=self.cyrillic_font)
                    empty_text = Paragraph("<i>Лист пуст</i>", empty_style)
                    story.append(empty_text)
                
                # Небольшой отступ
                story.append(Spacer(1, 4*mm))
        
        # После SUMMARY и SOURCES добавляем разрыв страницы
        if processed_sheets:
            story.append(PageBreak())
        
        # Обрабатываем остальные листы (кроме SUMMARY и SOURCES)
        for idx, sheet_name in enumerate(wb.sheetnames):
            if sheet_name in processed_sheets:
                continue
            
            sheet = wb[sheet_name]
            
            # Получаем данные из листа
            data = self._get_sheet_data(sheet)
            
            if data:
                # Создаем таблицу с заголовком категории внутри (обычный режим)
                # Заголовок категории будет встроен в таблицу и повторяться на каждой странице
                table = self._create_table(data, sheet, is_compact=False, sheet_name=sheet_name)
                story.append(table)
            else:
                # Если лист пустой - показываем заголовок отдельно
                sheet_title_style = ParagraphStyle(
                    'SheetTitle',
                    parent=self.styles['Heading2'],
                    fontName=self.cyrillic_font_bold,
                    fontSize=12,
                    textColor=colors.HexColor('#2563eb'),
                    spaceAfter=3,
                    spaceBefore=5,
                    alignment=TA_LEFT
                )
                sheet_title = Paragraph(f"<b>{sheet_name}</b>", sheet_title_style)
                story.append(sheet_title)
                story.append(Spacer(1, 2*mm))
                
                empty_style = ParagraphStyle('Empty', parent=self.styles['Normal'], fontName=self.cyrillic_font)
                empty_text = Paragraph("<i>Лист пуст</i>", empty_style)
                story.append(empty_text)
            
            # Отступ между листами
            story.append(Spacer(1, 6*mm))
        
        # Сборка PDF с нумерацией страниц
        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        return pdf_path


def export_bom_to_pdf(excel_path: str, output_pdf: Optional[str] = None, 
                     with_summary: bool = True, summary_info: Optional[dict] = None) -> str:
    """
    Удобная функция для экспорта BOM в PDF
    
    Args:
        excel_path: Путь к Excel файлу с результатами
        output_pdf: Путь к выходному PDF (опционально)
        with_summary: Включить сводную информацию
        summary_info: Дополнительная информация для сводки
    
    Returns:
        Путь к созданному PDF файлу
    """
    exporter = PDFExporter()
    
    if with_summary:
        return exporter.export_with_summary(excel_path, output_pdf, summary_info)
    else:
        return exporter.export_excel_to_pdf(excel_path, output_pdf)

