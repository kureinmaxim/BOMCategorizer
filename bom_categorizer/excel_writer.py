# -*- coding: utf-8 -*-
"""
Запись категоризованных BOM данных в Excel

Основные функции:
- write_categorized_excel: главная функция записи
- enrich_with_mr_and_total: добавление кодов МР и общих количеств
- format_excel_output: форматирование и сортировка
- apply_excel_styles: применение стилей к ячейкам
"""

import os
import re
from typing import Dict
import pandas as pd
from openpyxl.styles import Alignment, Border, Side

from .formatters import clean_component_name, extract_nominal_value, extract_tu_code
from .utils import find_column


def remove_duplicate_suffix(text: str) -> str:
    """
    Удаляет дублирование единиц измерения и допусков в конце строки.
    
    Примеры:
        "Р1 - 12 - 0,125 - 27.4 кОм ± 1% - М кОм ± 1% - М" 
        -> "Р1 - 12 - 0,125 - 27.4 кОм ± 1% - М"
        
        "P1 - 12 - 0,125 - 1 МОм ± 1% - M кОм ± 1% - M"
        -> "P1 - 12 - 0,125 - 1 МОм ± 1% - M"
    
    Args:
        text: Исходная строка
        
    Returns:
        Строка без дублирования
    """
    if not text or not isinstance(text, str):
        return text
    
    # Паттерн для единиц измерения с возможными допусками и моделями
    # Например: " кОм ± 1% - М" или " мкФ ± 5%" или " Ом"
    unit_pattern = r'\s+(МОм|мом|кОм|ком|Ом|ом|мкФ|мкф|нФ|нф|пФ|пф|мГн|мгн|мкГн|мкгн|нГн|нгн|Гн|гн)(?:\s*[±]\s*\d+(?:[,.]\d+)?%?)?(?:\s*[-–—]\s*[А-ЯЁA-Z])?'
    
    # Находим все вхождения единиц в строке
    matches = list(re.finditer(unit_pattern, text, re.IGNORECASE))
    
    if len(matches) < 2:
        # Если единица встречается меньше 2 раз, дублирования нет
        return text
    
    # Берем последние два вхождения
    last_match = matches[-1]
    prev_match = matches[-2]
    
    # Извлекаем текст двух последних вхождений с допусками
    last_text = last_match.group(0)
    prev_text = prev_match.group(0)
    
    # Проверяем похожи ли они (могут отличаться пробелами)
    last_normalized = re.sub(r'\s+', ' ', last_text.strip().lower())
    prev_normalized = re.sub(r'\s+', ' ', prev_text.strip().lower())
    
    # Случай 1: Полное совпадение (одинаковые единицы)
    if last_normalized == prev_normalized:
        # Дублирование найдено - удаляем последнее вхождение
        result = text[:last_match.start()] + text[last_match.end():]
        return result.strip()
    
    # Случай 2: Разные единицы, но одинаковые допуски и модели
    # Например: "МОм ± 1% - M" и "кОм ± 1% - M"
    # Извлекаем допуск+модель (всё после единицы измерения)
    def extract_tolerance_model(match_text):
        """Извлекает допуск и модель из текста единицы"""
        # Удаляем саму единицу измерения, оставляем допуск и модель
        tolerance_pattern = r'(МОм|мом|кОм|ком|Ом|ом|мкФ|мкф|нФ|нф|пФ|пф|мГн|мгн|мкГн|мкгн|нГн|нгн|Гн|гн)\s*(.*)$'
        match = re.search(tolerance_pattern, match_text.strip(), re.IGNORECASE)
        if match:
            return match.group(2).strip().lower()
        return ""
    
    last_tolerance = extract_tolerance_model(last_text)
    prev_tolerance = extract_tolerance_model(prev_text)
    
    # Если допуски и модели одинаковые и не пустые - это дублирование
    if last_tolerance and prev_tolerance and last_tolerance == prev_tolerance:
        # Удаляем последнее вхождение (избыточное)
        result = text[:last_match.start()] + text[last_match.end():]
        return result.strip()
    
    return text


def add_plus_minus_to_percentages(text: str) -> str:
    """
    Добавляет знак ± перед процентами, если его там нет
    
    Примеры:
        "100 Ом 5% - Т" → "100 Ом ± 5% - Т"
        "1 кОм ± 5%" → "1 кОм ± 5%" (не меняется)
        "220 пФ 10%-М" → "220 пФ ± 10%-М"
    
    Args:
        text: Исходный текст
        
    Returns:
        Текст с добавленным знаком ± перед процентами
    """
    if not text or pd.isna(text):
        return text
    
    text_str = str(text)
    
    # Паттерн: ищем процент, перед которым НЕТ знака ±
    # Захватываем: (единица измерения или цифра)(пробел?)(цифры%)(остальное)
    # Примеры совпадений: "Ом 5%", "Ом5%", "пФ 10%", "кОм2%"
    pattern = r'(Ом|пФ|нФ|мкФ|мФ|кОм|МОм|Гн|мГн|мкГн|нГн|\d)\s*(\d+%)'
    
    # Функция замены: добавляет ± между единицей измерения и процентом
    def replace_with_plus_minus(match):
        unit_or_digit = match.group(1)
        percentage = match.group(2)
        
        # Проверяем, не стоит ли уже ± перед найденным местом
        start_pos = match.start()
        if start_pos > 0:
            # Смотрим на 3 символа назад (на случай пробелов)
            prefix = text_str[max(0, start_pos-3):start_pos]
            if '±' in prefix:
                # Знак ± уже есть, не добавляем
                return match.group(0)
        
        # Добавляем пробел ± пробел между единицей и процентом
        return f"{unit_or_digit} ± {percentage}"
    
    result = re.sub(pattern, replace_with_plus_minus, text_str, flags=re.IGNORECASE)
    
    return result


# Русские названия категорий для листов Excel
RUS_SHEET_NAMES = {
    "ics": "Микросхемы",
    "resistors": "Резисторы",
    "capacitors": "Конденсаторы",
    "inductors": "Индуктивности",
    "semiconductors": "Полупроводники",
    "connectors": "Разъемы",
    "optics": "Оптические компоненты",
    "power_modules": "Модули питания",
    "cables": "Кабели",
    "our_developments": "Наши разработки",
    "dev_boards": "Отладочные платы",
    "rf_modules": "СВЧ модули",
    "others": "Другие",
    "unclassified": "Не распределено",
}


def enrich_with_mr_and_total(df: pd.DataFrame) -> pd.DataFrame:
    """
    Добавляет колонку 'Общее количество' с суммой по группам
    
    Группировка: по коду МР, если доступен, иначе по описанию и типу
    
    Args:
        df: DataFrame с данными компонентов
        
    Returns:
        DataFrame с добавленной колонкой 'Общее количество'
    """
    if df.empty:
        df["Общее количество"] = 0
        return df

    mr_col = find_column(["код мр", "part number", "pn", "mpn"], df.columns)
    qty_col = find_column(["количество", "qty", "quantity", "_merged_qty_"], df.columns)
    
    enriched = df.copy()
    tmp = enriched.copy()

    # Обработка колонки количества: безопасное преобразование в float
    if qty_col and qty_col in tmp.columns:
        # Используем pd.to_numeric с errors='coerce' для безопасного преобразования
        qty_series = pd.to_numeric(tmp[qty_col], errors='coerce')
        # Заменяем все NaN на 1
        qty_series = qty_series.fillna(1).astype(float)
    else:
        qty_series = pd.Series([1] * len(tmp), dtype=float)

    # Определяем ключи для группировки
    # ВАЖНО: если mr_col пустой или отсутствует, используем _merged_description_
    if mr_col and mr_col in tmp.columns and tmp[mr_col].notna().any():
        group_keys = [mr_col]
    else:
        # Если нет mr_col, группируем по _merged_description_
        if '_merged_description_' in tmp.columns:
            group_keys = ['_merged_description_']
        elif 'description' in tmp.columns:
            group_keys = ['description']
        else:
            # Fallback: группируем по всем возможным колонкам
            group_keys = [c for c in tmp.columns if c in ['description', 'value', 'part']]
            if not group_keys:
                # Если ничего не нашли, не группируем
                enriched["Общее количество"] = qty_series
                return enriched

    # Конвертируем все группировочные колонки в строки для избежания проблем с mixed types
    for key in group_keys:
        if key in tmp.columns:
            tmp[key] = tmp[key].fillna('').astype(str)

    tmp["__qty__"] = qty_series
    totals = tmp.groupby(group_keys, dropna=False)["__qty__"].sum().reset_index().rename(columns={"__qty__": "Общее количество"})

    # Конвертируем те же колонки в enriched для корректного merge
    for key in group_keys:
        if key in enriched.columns:
            enriched[key] = enriched[key].fillna('').astype(str)

    enriched = enriched.merge(totals, on=group_keys, how="left")
    enriched["Общее количество"] = enriched["Общее количество"].fillna(1).astype(int)

    return enriched


def format_excel_output(df: pd.DataFrame, sheet_name: str, desc_col: str, force_reprocess: bool = False) -> pd.DataFrame:
    """
    Форматирует DataFrame для записи в Excel:
    - Очистка названий компонентов
    - Извлечение ТУ
    - Добавление примечаний
    - Сортировка
    - Нумерация
    
    Args:
        df: DataFrame с данными
        sheet_name: Название листа (категория)
        desc_col: Название колонки с описанием
        force_reprocess: Принудительно обработать даже если файл уже обработан (для сравнения)
        
    Returns:
        Отформатированный DataFrame
    """
    if df.empty:
        return df
    
    result_df = df.copy()
    
    # Переименовать столбцы
    # СНАЧАЛА ищем оригинальную колонку количества
    qty_col_candidates = ['_merged_qty_', 'qty', 'Количество', 'количество', 'Общее количество']
    for candidate in qty_col_candidates:
        if candidate in result_df.columns:
            if candidate != 'шт.':
                result_df = result_df.rename(columns={candidate: 'шт.'})
            break
    
    # Переименовать "Код МР" в стандартное написание
    kod_mr_candidates = ['код мр', 'код_мр', 'kodmr', 'Код мр', 'КОД МР']
    for candidate in kod_mr_candidates:
        if candidate in result_df.columns:
            if candidate != 'Код МР':
                result_df = result_df.rename(columns={candidate: 'Код МР'})
            break
    
    if 'наименование ивп' in result_df.columns:
        result_df = result_df.rename(columns={'наименование ивп': 'Наименование ИВП'})
    # Переименовать нормализованные английские колонки в русские
    if '_merged_description_' in result_df.columns and 'Наименование ИВП' not in result_df.columns:
        result_df = result_df.rename(columns={'_merged_description_': 'Наименование ИВП'})
    elif 'description' in result_df.columns and 'Наименование ИВП' not in result_df.columns:
        result_df = result_df.rename(columns={'description': 'Наименование ИВП'})
    
    # Обработать наименование - очистить и извлечь ТУ
    # Найти столбец с наименованием (ВАЖНО: сначала ищем переименованный столбец!)
    desc_col_name = None
    for possible_name in ['Наименование ИВП', 'наименование ивп', 'описание', 'наименование', desc_col]:
        if possible_name and possible_name in result_df.columns:
            desc_col_name = possible_name
            break
    
    if not desc_col_name:
        return result_df
    
    # Проверяем, есть ли уже колонки ТУ и Примечание (файл уже обработан)
    has_tu_column = 'ТУ' in result_df.columns or 'ту' in result_df.columns
    has_primechanie_column = 'Примечание' in result_df.columns or 'примечание' in result_df.columns
    
    # Если force_reprocess=True, удаляем старые колонки для повторной обработки
    if force_reprocess and (has_tu_column or has_primechanie_column):
        cols_to_drop = []
        if 'ТУ' in result_df.columns:
            cols_to_drop.append('ТУ')
        if 'ту' in result_df.columns:
            cols_to_drop.append('ту')
        if 'Примечание' in result_df.columns:
            cols_to_drop.append('Примечание')
        if 'примечание' in result_df.columns:
            cols_to_drop.append('примечание')
        if cols_to_drop:
            result_df = result_df.drop(columns=cols_to_drop)
        has_tu_column = False
        has_primechanie_column = False
    
    # Если колонки ТУ и Примечание уже есть, НЕ обрабатываем повторно
    if has_tu_column and has_primechanie_column:
        # Файл уже обработан, просто нормализуем имена колонок
        if 'ту' in result_df.columns:
            result_df = result_df.rename(columns={'ту': 'ТУ'})
        if 'примечание' in result_df.columns:
            result_df = result_df.rename(columns={'примечание': 'Примечание'})
        cleaned_data = []  # Пустой список, чтобы не нарушить дальнейшую логику
    else:
        # Применить функцию очистки к каждой строке (только для новых файлов)
        cleaned_data = []
        for idx, row in result_df.iterrows():
            text = str(row[desc_col_name]) if pd.notna(row[desc_col_name]) else ""
            note = str(row['note']) if 'note' in result_df.columns and pd.notna(row['note']) else ""
            
            # Извлечь ТУ из note (если есть)
            note_tu = ""
            note_manufacturer = ""
            note_type = ""
            if note:
                if '|' in note:
                    # ИСПРАВЛЕНО: Формат "ТУ | manufacturer" (например, "АЕЯР.431320.420ТУ | Texas Instruments")
                    parts = note.split('|')
                    if len(parts) >= 2:
                        # Первая часть - ТУ (для отечественных) или пустая строка
                        note_tu = parts[0].strip()
                        # Вторая часть - производитель (для импортных)
                        note_manufacturer = parts[1].strip()
                        
                        # Определяем что использовать в колонке ТУ:
                        # Если первая часть похожа на ТУ-код (содержит "ТУ"), используем её
                        # Иначе используем производителя
                        if note_tu and ('ТУ' in note_tu.upper() or 'TU' in note_tu.upper()):
                            # Это отечественный ТУ-код
                            pass  # note_tu уже содержит правильное значение
                        elif note_manufacturer:
                            # Это импортный компонент, производитель в колонку ТУ
                            note_tu = note_manufacturer
                            note_manufacturer = ""
                    else:
                        # Если не удалось разделить, используем как есть
                        note_tu = parts[0].strip()
                else:
                    # Весь note это производитель или ТУ (например, "STMicroelectronics" или "АЕЯР.431320.420ТУ")
                    note_tu = note.strip()
            
            # НЕ ПРИМЕНЯЕМ clean_component_name здесь, так как данные уже очищены в main.py!
            # Просто используем text как есть
            cleaned_text = text
            
            # Извлечь ТУ из текста
            cleaned_text, tu_code = extract_tu_code(cleaned_text)
            
            # ИСПРАВЛЕНО: Если ТУ был в note, используем его (приоритет у note), 
            # НО если note_tu пустой, а tu_code из текста есть - используем tu_code из текста
            if note_tu:
                tu_code = note_tu
            # Если note_tu пустой, но у нас уже есть tu_code из текста - оставляем его
            
            # ВАЖНО: Для категории "Наши разработки" очищаем ТУ (там не нужны производители)
            if sheet_name == "Наши разработки":
                tu_code = ""
            
            # Определить тип компонента
            comp_type = note_type if note_type else ""
            
            cleaned_data.append((cleaned_text, tu_code, comp_type))
        
        result_df[desc_col_name] = [item[0] for item in cleaned_data]
    
    # НЕ удаляем артикулы модулей! Каждый модуль с уникальным артикулом - отдельная строка
    # Группировка будет только для ПОЛНОСТЬЮ ОДИНАКОВЫХ модулей
    
    # НЕ добавляем префиксы категорий, так как clean_component_name уже удалил их
    # Компоненты должны остаться в своем "чистом" виде
    # Например: "МАТРИЦА ТРАНЗИСТОРНАЯ 1НТ251" остается как есть
    #           "Адаптер QASNL-FF" остается как есть
    
    # Для "Наших разработок" - если название пустое, взять из source_file
    if sheet_name == 'Наши разработки' and 'source_file' in result_df.columns:
        for idx in result_df.index:
            if not result_df.loc[idx, desc_col_name] or pd.isna(result_df.loc[idx, desc_col_name]) or str(result_df.loc[idx, desc_col_name]).strip() == '':
                source_file = result_df.loc[idx, 'source_file']
                if source_file and pd.notna(source_file):
                    # Извлечь название файла без расширения
                    file_name = os.path.splitext(os.path.basename(str(source_file)))[0]
                    result_df.loc[idx, desc_col_name] = file_name
    
    # Вставить ТУ и Примечание ТОЛЬКО если их еще нет (новый файл)
    if not has_tu_column and cleaned_data:
        tu_data = [item[1] for item in cleaned_data]
        desc_idx = list(result_df.columns).index(desc_col_name)
        result_df.insert(desc_idx + 1, 'ТУ', tu_data)
    
    if not has_primechanie_column:
        # Для модулей питания используем reference в качестве Примечания
        if sheet_name == 'Модули питания' and 'reference' in result_df.columns:
            # Копируем reference в Примечание
            primechanie = result_df['reference'].tolist()
        elif cleaned_data:
            component_types = [item[2] for item in cleaned_data]
            
            # Определить стандартный тип для категории
            category_standard_types = {
                'Резисторы': 'Резистор',
                'Конденсаторы': 'Конденсатор',
                'Индуктивности': 'Дроссель',
                'Микросхемы': 'Микросхема',
                'Разъемы': 'Разъем',
                'Полупроводники': '',  # Нет стандартного типа
            }
            
            standard_type = category_standard_types.get(sheet_name, '')
            
            # Если тип компонента совпадает со стандартным - оставляем пустую ячейку
            primechanie = []
            for comp_type in component_types:
                if not comp_type or comp_type == standard_type:
                    primechanie.append('')  # Пустая ячейка вместо прочерка
                else:
                    primechanie.append(comp_type)
        else:
            primechanie = []
        
        # Вставить колонку Примечание, если есть данные
        if primechanie:
            # Найти позицию после ТУ
            if 'ТУ' in result_df.columns:
                tu_idx = list(result_df.columns).index('ТУ')
                result_df.insert(tu_idx + 1, 'Примечание', primechanie)
            else:
                desc_idx = list(result_df.columns).index(desc_col_name)
                result_df.insert(desc_idx + 1, 'Примечание', primechanie)
    
    # ГРУППИРОВКА: Группируем ТОЛЬКО полностью одинаковые компоненты
    # (название + ТУ должны совпадать полностью, включая артикулы)
    if desc_col_name in result_df.columns:
        # Определяем колонки для группировки
        group_cols = [desc_col_name, 'ТУ']
        
        # Находим колонку с quantity и примечанием
        qty_col = find_column(['шт.', 'qty', 'quantity', '_merged_qty_'], result_df.columns)
        # ВАЖНО: сначала ищем "Примечание", а не "reference"
        ref_col = None
        for col_name in ['Примечание', 'примечание']:
            if col_name in result_df.columns:
                ref_col = col_name
                break
        if not ref_col:
            ref_col = find_column(['reference'], result_df.columns)
        
        if qty_col and ref_col:
            # Проверяем, есть ли дубликаты для группировки
            duplicates_mask = result_df.duplicated(subset=group_cols, keep=False)
            if duplicates_mask.any():
                # Агрегация: сумма для quantity, объединение для reference/Примечание
                agg_dict = {
                    qty_col: 'sum',
                    ref_col: lambda x: ', '.join([str(v) for v in x if pd.notna(v) and str(v).strip()])
                }
                
                # Добавляем все остальные колонки (берём первое значение)
                # НО для 'reference' тоже применяем объединение (если не совпадает с ref_col)
                for col in result_df.columns:
                    if col not in group_cols and col not in [qty_col, ref_col]:
                        # Если это 'reference' и он НЕ используется как ref_col, тоже объединяем
                        if col == 'reference' and col != ref_col:
                            agg_dict[col] = lambda x: ', '.join([str(v) for v in x if pd.notna(v) and str(v).strip()])
                        else:
                            agg_dict[col] = 'first'
                
                # Группируем
                result_df = result_df.groupby(group_cols, as_index=False, dropna=False).agg(agg_dict)
    
    # Сортировка зависит от категории
    if sheet_name in ['Конденсаторы', 'Дроссели', 'Резисторы', 'Индуктивности']:
        # Сортировка ТОЛЬКО по номиналу (игнорируя тип компонента)
        category_map = {
            'Резисторы': 'resistors',
            'Конденсаторы': 'capacitors',
            'Дроссели': 'inductors',
            'Индуктивности': 'inductors',
        }
        category_key = category_map.get(sheet_name, 'resistors')
        
        # Извлекаем номинал для каждого компонента
        def get_nominal_value(text):
            result = extract_nominal_value(str(text), category_key)
            # result может быть tuple (value, unit) или просто значение
            if isinstance(result, tuple):
                return result[0] if result[0] is not None else float('inf')
            else:
                return result if result is not None else float('inf')
        
        result_df['_nominal_value'] = result_df[desc_col_name].apply(get_nominal_value)
        
        # Сортируем ТОЛЬКО по номиналу (игнорируя тип)
        result_df = result_df.sort_values(
            by=['_nominal_value', desc_col_name],
            ascending=[True, True]
        )
        result_df = result_df.drop(columns=['_nominal_value'])
    
    elif sheet_name in ['Отладочные платы и модули', 'Модули питания', 'Оптические компоненты', 
                        'Полупроводники', 'Разъемы', 'Кабели', 'Другие']:
        # Алфавитная сортировка для этих категорий
        result_df = result_df.sort_values(by=desc_col_name, ascending=True)
    
    elif sheet_name == 'Микросхемы':
        # Специальная сортировка: латинские названия перед кириллическими
        def get_sort_key(text):
            text = str(text).strip()
            if not text:
                return (2, text)  # Пустые в конец
            
            # Найти первую букву (не цифру) для определения группы
            first_letter = None
            for char in text:
                char_upper = char.upper()
                if 'A' <= char_upper <= 'Z' or 'А' <= char_upper <= 'Я' or char_upper == 'Ё':
                    first_letter = char_upper
                    break
            
            if first_letter:
                # Латинские символы (A-Z) - группа 0
                if 'A' <= first_letter <= 'Z':
                    return (0, text.upper())
                # Кириллические символы - группа 1
                elif 'А' <= first_letter <= 'Я' or first_letter == 'Ё':
                    return (1, text.upper())
            
            # Если нет букв - группа 2
            return (2, text.upper())
        
        result_df['_sort_key'] = result_df[desc_col_name].apply(get_sort_key)
        result_df = result_df.sort_values(by='_sort_key', ascending=True)
        result_df = result_df.drop(columns=['_sort_key'])
    
    # Для остальных категорий (Наши разработки) - без сортировки
    
    result_df = result_df.reset_index(drop=True)
    
    # Переименовать source_file в Источник (только если "Источник" еще нет)
    if 'source_file' in result_df.columns and 'Источник' not in result_df.columns:
        result_df = result_df.rename(columns={'source_file': 'Источник'})
    
    # Добавить номера п/п к Источнику (только если еще не добавлены)
    if 'Источник' in result_df.columns:
        # Найти колонку № п/п или № п\п (если есть несколько с похожими названиями)
        pp_columns = [col for col in result_df.columns if str(col).startswith('№ п')]
        if pp_columns:
            # Проверить, не добавлены ли уже номера п/п (при повторной обработке)
            first_source = str(result_df['Источник'].iloc[0]) if not result_df.empty else ""
            if ', п/п ' not in first_source:
                # Взять последнюю колонку № п/п (если есть дубликаты)
                pp_col = pp_columns[-1]  # Берем последнюю
                
                def add_pp_number(row):
                    pp_val = row[pp_col]
                    # Проверяем что значение есть, не NaN и не пустая строка
                    if pd.notna(pp_val) and str(pp_val).strip():
                        try:
                            return f"{row['Источник']}, п/п {int(float(pp_val))}"
                        except (ValueError, TypeError):
                            return row['Источник']
                    return row['Источник']
                
                result_df['Источник'] = result_df.apply(add_pp_number, axis=1)
    
    # Перенести позиционные обозначения из 'reference' в 'Примечание' (для DOC/DOCX файлов)
    if 'reference' in result_df.columns:
        # Если колонка "Примечание" есть - добавляем reference туда
        if 'Примечание' in result_df.columns:
            for idx in result_df.index:
                ref_val = result_df.loc[idx, 'reference']
                prim_val = result_df.loc[idx, 'Примечание']
                
                # Если reference не пустой
                if pd.notna(ref_val) and str(ref_val).strip():
                    result_df.loc[idx, 'Примечание'] = str(ref_val).strip()
        else:
            # Если колонки "Примечание" нет - создаем её из reference
            result_df['Примечание'] = result_df['reference'].fillna('')
    
    # Удалить ненужные колонки (НЕ удаляем Код МР!)
    cols_to_remove = ['ед. изм. ктд', '_merged_qty_', 
                      'ед. изм. КТД',
                      'первоначальная цена, тыс.руб.', 'первоначальная стоимость, тыс.руб.',
                      'источник',  # дубликат "Источник" (с маленькой буквы)
                      'категория',  # дубликат категории
                      'category',  # техническая колонка категории
                      'общее количество',  # дубликат
                      'source_file', 'source_sheet',  # служебные колонки (уже в "Источник")
                      'note',  # служебная колонка
                      'zone',  # служебная колонка из DOC/DOCX (уже не нужна)
                      'reference',  # служебная колонка из DOC/DOCX (перенесена в Примечание)
                      'group_type',  # служебная колонка типа из заголовка
                      'original_note',  # оригинальное примечание (использовано для подборов)
                      'has_explicit_qty',  # техническая колонка для определения явного количества
                      '_extracted_tu_',  # техническая колонка для извлечения ТУ (не показываем пользователю)
                      '_normalized_desc_']  # техническая колонка для агрегации дубликатов
    
    # Добавить все колонки № п/п и № п\п для удаления (исходные, не новую)
    pp_columns = [col for col in result_df.columns if str(col).startswith('№ п')]
    cols_to_remove.extend(pp_columns)
    
    # Добавить все колонки "unnamed" для удаления
    unnamed_columns = [col for col in result_df.columns if 'unnamed' in str(col).lower()]
    cols_to_remove.extend(unnamed_columns)
    
    for col in cols_to_remove:
        if col in result_df.columns:
            result_df = result_df.drop(columns=[col])
    
    # Добавить № п/п в начало (ПОСЛЕ сортировки!)
    # Сначала удалим если уже есть
    if '№ п/п' in result_df.columns:
        result_df = result_df.drop(columns=['№ п/п'])
    result_df.insert(0, '№ п/п', range(1, len(result_df) + 1))
    
    # ВАЖНО: Сначала нормализуем регистр ключевых колонок (для уже обработанных файлов)
    # Это должно быть ДО создания final_cols
    cols_lower = {col.lower(): col for col in result_df.columns}
    
    # Проверяем и переименовываем № ТРУ
    if '№ тру' in cols_lower:
        existing_col = cols_lower['№ тру']
        if existing_col != '№ ТРУ':
            result_df = result_df.rename(columns={existing_col: '№ ТРУ'})
    else:
        # Колонки нет - создаем пустую
        result_df['№ ТРУ'] = ''
    
    # Проверяем и переименовываем Стоимость
    if 'стоимость' in cols_lower:
        existing_col = cols_lower['стоимость']
        if existing_col != 'Стоимость':
            result_df = result_df.rename(columns={existing_col: 'Стоимость'})
    else:
        # Колонки нет - создаем пустую
        result_df['Стоимость'] = ''
    
    # Упорядочить колонки в правильном порядке
    # Код МР после ТУ, Примечание в конце
    desired_order = ['№ п/п', 'Наименование ИВП', 'ТУ', 'Код МР', 'Источник', 'шт.']
    
    ordered_cols = [col for col in desired_order if col in result_df.columns]
    remaining_cols = [col for col in result_df.columns 
                      if col not in ordered_cols and col not in cols_to_remove and col != 'Примечание' 
                      and col not in ['№ ТРУ', 'Стоимость']]  # Исключаем, чтобы добавить в конец
    
    # Примечание добавляем в конец, затем пустые колонки № ТРУ и Стоимость
    final_cols = ordered_cols + remaining_cols
    if 'Примечание' in result_df.columns:
        final_cols.append('Примечание')
    
    # Добавляем № ТРУ и Стоимость в конец
    final_cols.append('№ ТРУ')
    final_cols.append('Стоимость')
    
    result_df = result_df[final_cols]
    
    # Добавить знак ± перед процентами в Наименовании ИВП
    if 'Наименование ИВП' in result_df.columns:
        result_df['Наименование ИВП'] = result_df['Наименование ИВП'].apply(add_plus_minus_to_percentages)
    
    # КРИТИЧНО: Удаляем дублирование единиц измерения и допусков
    # Пример: "Р1-12-0,125-27.4 кОм ± 1% - М кОм ± 1% - М" -> "Р1-12-0,125-27.4 кОм ± 1% - М"
    if 'Наименование ИВП' in result_df.columns:
        result_df['Наименование ИВП'] = result_df['Наименование ИВП'].apply(
            lambda x: remove_duplicate_suffix(str(x)) if pd.notna(x) else x
        )
    
    return result_df


def apply_excel_styles(writer: pd.ExcelWriter):
    """
    Применяет стили к Excel файлу:
    - Выравнивание (center для большинства, left для описания, ТУ, Примечание и Источник)
    - Автоматическая ширина столбцов
    
    Args:
        writer: ExcelWriter с уже записанными данными
    """
    for sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]
        
        # Найти индексы столбцов "Наименование ИВП", "ТУ", "Код МР", "Примечание" и "Источник"
        desc_col_idx = None
        tu_col_idx = None
        kod_mr_col_idx = None
        note_col_idx = None
        source_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            cell_val = str(cell.value).lower() if cell.value else ''
            if 'наименование ивп' in cell_val or 'наименование' in cell_val:
                desc_col_idx = idx
            elif cell_val == 'ту':
                tu_col_idx = idx
            elif 'код мр' in cell_val:
                kod_mr_col_idx = idx
            elif 'примечание' in cell_val:
                note_col_idx = idx
            elif 'источник' in cell_val or cell_val == 'source_file':
                source_col_idx = idx
        
        # Установить текстовый формат для колонки "Код МР" (чтобы избежать научной нотации)
        if kod_mr_col_idx:
            column_letter = ws.cell(row=1, column=kod_mr_col_idx).column_letter
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=kod_mr_col_idx)
                cell.number_format = '@'  # Текстовый формат
        
        # Создать стиль границ (тонкие черные линии со всех сторон)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Центрировать все ячейки, кроме "наименование ивп", "ТУ", "Примечание" и "Источник", и добавить границы
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                # Выравнивание
                if col_idx in (desc_col_idx, tu_col_idx, note_col_idx, source_col_idx):
                    # Наименование ИВП, ТУ, Примечание и Источник - выравнивание по левому краю
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    # Все остальные - по центру
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Границы для всех ячеек
                cell.border = thin_border
        
        # Автоматически установить ширину столбцов по содержимому
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width


def write_categorized_excel(
    outputs: Dict[str, pd.DataFrame],
    df: pd.DataFrame,
    output_xlsx: str,
    combine: bool,
    desc_col: str
):
    """
    Записывает категоризованные данные в Excel файл
    
    Args:
        outputs: Словарь {category_key: DataFrame}
        df: Полный DataFrame со всеми данными
        output_xlsx: Путь к выходному Excel файлу
        combine: Если True, добавляет SUMMARY лист
        desc_col: Название колонки с описанием
    """
    sheets_written = 0
    category_sheets = []  # Список записанных листов категорий для SUMMARY
    
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        # Сначала записываем все категории
        for key, part_df in outputs.items():
            if len(part_df) == 0:
                continue
            
            sheet_name = RUS_SHEET_NAMES.get(key, key)
            
            # Данные уже отформатированы в main.py (включая ТУ)
            # НЕ применяем format_excel_output повторно
            result_df = part_df.copy()
            
            # Проверка что есть данные
            if result_df.empty or len(result_df) == 0:
                continue
            
            # Записать в Excel
            result_df.to_excel(writer, sheet_name=sheet_name, index=False)
            category_sheets.append(sheet_name)
            sheets_written += 1
        
        # SOURCES sheet (записываем до SUMMARY)
        # Очищаем source_file от тегов замен/подборов: (зам D4), (п/б C21*) и т.д.
        unique_sources = set()
        for _, r in df.iterrows():
            source_file = r.get("source_file", "")
            source_sheet = r.get("source_sheet", "")
            
            # Очищаем source_file от тегов в скобках: (зам ...), (п/б ...), (подбор ...)
            if source_file:
                # Убираем ВСЕ теги в скобках (может быть несколько: (п/б D3*), (п/б D5*), ...)
                import re
                clean_source = source_file
                # Повторяем пока есть скобки (убираем все теги, даже если их несколько)
                while '(' in clean_source:
                    prev = clean_source
                    clean_source = re.sub(r'\s*\([^)]*\)', '', clean_source)
                    # Если ничего не изменилось - выходим (защита от бесконечного цикла)
                    if prev == clean_source:
                        break
                clean_source = clean_source.strip().rstrip(',').strip()
                unique_sources.add((clean_source, source_sheet))
            else:
                unique_sources.add((source_file, source_sheet))
        
        sources = pd.DataFrame(
            sorted(unique_sources), 
            columns=["source_file", "source_sheet"]
        )
        if not sources.empty:
            sources.to_excel(writer, sheet_name="SOURCES", index=False)
            sheets_written += 1
        
        # Проверить что хотя бы один лист записан
        if sheets_written == 0:
            # Записать пустой лист с сообщением
            empty_df = pd.DataFrame({'Сообщение': ['Нет данных для записи']})
            empty_df.to_excel(writer, sheet_name="INFO", index=False)
            sheets_written = 1
        
        # Применить стили только если есть листы
        if sheets_written > 0:
            apply_excel_styles(writer)
    
    # После записи всех листов, создаем SUMMARY, читая реальные данные из файла
    if combine and len(category_sheets) > 0:
        summary_rows = []
        
        # Читаем каждый лист категории из уже записанного файла
        for sheet_name in category_sheets:
            try:
                # Читаем лист обратно из файла
                df_sheet = pd.read_excel(output_xlsx, sheet_name=sheet_name, dtype=str)
                
                # Ищем колонку с количеством
                qty_col = find_column([
                    "шт.", "шт", "Кол-во", "qty", "quantity", "количество", "кол.", "кол-во"
                ], list(df_sheet.columns))
                
                # Считаем количество позиций и общее количество
                positions_count = len(df_sheet)
                total_qty = 0
                
                if qty_col and qty_col in df_sheet.columns:
                    for val in df_sheet[qty_col]:
                        try:
                            if pd.notna(val):
                                total_qty += int(float(val))
                        except (ValueError, TypeError):
                            pass
                else:
                    # Если колонка не найдена, используем количество строк
                    total_qty = positions_count
                
                summary_rows.append({
                    '№ п/п': len(summary_rows) + 1,
                    'Категория': sheet_name,
                    'Кол-во позиций': positions_count,
                    'Общее количество': total_qty
                })
            
            except Exception as e:
                print(f"[WARNING] Не удалось прочитать лист '{sheet_name}' для SUMMARY: {e}")
        
        # Записываем SUMMARY лист
        if summary_rows:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            
            wb = load_workbook(output_xlsx)
            
            # Создаем DataFrame для SUMMARY
            summary_df = pd.DataFrame(summary_rows)
            
            # Если лист SUMMARY уже существует, удаляем его
            if "SUMMARY" in wb.sheetnames:
                del wb["SUMMARY"]
            
            # Создаем новый лист SUMMARY
            ws = wb.create_sheet("SUMMARY", 0)  # Вставляем в начало
            
            # Записываем заголовки с жирным шрифтом
            header_font = Font(bold=True)
            for col_idx, col_name in enumerate(summary_df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Записываем данные
            for row_idx, row_data in enumerate(summary_df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Центрируем все ячейки, кроме "Категория"
                    if col_idx == 2:  # Колонка "Категория"
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Максимум 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_xlsx)
            wb.close()
            
            print(f"[SUMMARY] Создан лист SUMMARY с {len(summary_rows)} категориями")
    
    print(f"XLSX written: {output_xlsx} ({sheets_written} sheets)")
