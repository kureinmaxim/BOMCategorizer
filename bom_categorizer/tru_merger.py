# -*- coding: utf-8 -*-
"""
Модуль для объединения данных из файлов ТРУ с BOM файлами.

Функции:
- normalize_for_matching: нормализация строк для сопоставления
- merge_tru_into_bom: основная функция объединения
- apply_merge_styles: применение стилей к изменённым строкам
"""

import re
import pandas as pd
from typing import List, Dict, Tuple, Optional, Set
from difflib import SequenceMatcher
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from .formatters import extract_tu_code


# Цвета для форматирования
MERGED_ROW_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')  # Бледно-голубая
MERGED_NAME_FONT = Font(color='1a237e', bold=False)  # Тёмно-синий
OLD_QTY_FONT = Font(color='FF0000')  # Красный для старого количества


def normalize_erp_code_from_artikul(value) -> str:
    """
    Нормализует значение из колонки 'Артикул' в строковый ERP-код.
    Важно: в исходных данных иногда встречаются текстовые значения (например, повтор заголовка 'Артикул'),
    поэтому функция должна быть максимально безопасной и не выбрасывать исключения.

    Правила:
    - пусто/NaN -> ''
    - 'Артикул' -> '' (повтор заголовка)
    - '000123' -> '000123' (сохраняем ведущие нули)
    - '123.0' / '123,0' -> '123'
    - 'арт. 12345' -> '12345' (fallback на цифры)
    """
    if value is None or pd.isna(value):
        return ''

    s = str(value).strip()
    if not s:
        return ''

    if s.lower() == 'артикул':
        return ''

    # если уже чистые цифры — возвращаем как есть (сохраняем ведущие нули)
    if s.isdigit():
        return s

    # пробуем распарсить как число (поддержка запятой и NBSP)
    s_num = s.replace('\u00A0', '').replace(' ', '').replace(',', '.')
    try:
        f = float(s_num)
        if pd.isna(f):
            return ''
        return str(int(f))
    except Exception:
        m = re.search(r'\d+', s_num)
        return m.group(0) if m else ''


def normalize_for_matching(text: str) -> str:
    """
    Нормализует строку для сопоставления:
    - Удаляет лишние пробелы
    - Заменяет разные виды тире на обычный дефис
    - Убирает подчёркивания и суффиксы типа _LW
    - Приводит к нижнему регистру
    """
    if not text or pd.isna(text):
        return ""
    
    text = str(text).strip()
    
    # Заменяем разные виды тире и пробелов
    text = re.sub(r'[\u2010\u2011\u2012\u2013\u2014\u2015\u2212]', '-', text)  # Различные тире
    text = re.sub(r'\s+', ' ', text)  # Множественные пробелы → один
    text = re.sub(r'\s*-\s*', '-', text)  # Пробелы вокруг тире
    text = re.sub(r'_[A-Za-z]{1,3}$', '', text)  # Суффиксы типа _LW, _AB
    text = text.replace('_', '-')  # Подчёркивания → дефисы
    
    return text.lower()


def extract_nominal(text: str) -> str:
    """
    Извлекает номинал из названия компонента.
    Примеры: "10 кОм", "100 нФ", "4.7 мкФ"
    """
    if not text:
        return ""
    
    # Паттерн для номинала: число + единица измерения
    patterns = [
        r'(\d+[\.,]?\d*)\s*(к[Оо]м|[Мм][Оо]м|[Оо]м|н[Фф]|мк[Фф]|п[Фф]|м[Гг]н|мк[Гг]н|н[Гг]н)',
        r'(\d+[\.,]?\d*)\s*(kOhm|MOhm|Ohm|nF|uF|pF|mH|uH|nH)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, str(text), re.IGNORECASE)
        if match:
            return normalize_for_matching(match.group(0))
    
    return ""


def similarity_ratio(a: str, b: str) -> float:
    """Вычисляет степень схожести двух строк (0.0 - 1.0)"""
    return SequenceMatcher(None, a, b).ratio()


def _find_qty_col(columns) -> Optional[str]:
    """Пытается найти колонку количества по типичным названиям."""
    for c in columns:
        cl = str(c).strip().lower()
        if cl in ['шт.', 'шт', 'qty', 'количество', 'кол-во', 'кол.', 'кол']:
            return c
    return None


_QTY_PAIR_RE = re.compile(r'^\s*(\d+)\s*\(\s*(\d+)\s*\)\s*$')


def _parse_qty_pair(value) -> Optional[Tuple[int, int]]:
    """
    Парсит формат количества "TRU (BOM)" -> (tru_qty, bom_qty).
    Возвращает None если формат не совпал.
    """
    if value is None or pd.isna(value):
        return None
    m = _QTY_PAIR_RE.match(str(value))
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _concat_preserve_columns(parts: List[pd.DataFrame]) -> pd.DataFrame:
    """
    Склеивает DataFrame'ы в один, сохраняя порядок колонок:
    сначала колонки первого DF, затем добавляем новые (если встретились).
    """
    frames = [p for p in parts if isinstance(p, pd.DataFrame) and not p.empty]
    if not frames:
        return pd.DataFrame()
    col_order: List[str] = []
    seen = set()
    for f in frames:
        for c in f.columns:
            if c not in seen:
                seen.add(c)
                col_order.append(c)
    normalized = []
    for f in frames:
        ff = f.copy()
        for c in col_order:
            if c not in ff.columns:
                ff[c] = ''
        normalized.append(ff[col_order])
    return pd.concat(normalized, ignore_index=True)


def build_ostatki_and_zapas_reports(
    merged_df: pd.DataFrame,
    merged_indices: Set[int],
    unmatched_tru: Optional[pd.DataFrame] = None,
    qty_col: Optional[str] = None,
    note_col: str = 'Примечание'
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Строит два "плоских" отчёта:
    - *_ostatki: BOM позиции, которых нет в ТРУ + (BOM_qty - TRU_qty) для уменьшенных количеств
    - *_zapas:   несопоставленные ТРУ + (TRU_qty - BOM_qty) для увеличенных количеств
    """
    if merged_df is None or merged_df.empty:
        # даже если merged_df пустой, запасы могут быть из unmatched_tru
        zapas = unmatched_tru.copy() if isinstance(unmatched_tru, pd.DataFrame) else pd.DataFrame()
        return pd.DataFrame(), zapas

    qty_col = qty_col or _find_qty_col(merged_df.columns)

    ostatki_parts: List[pd.DataFrame] = []
    zapas_parts: List[pd.DataFrame] = []

    # 1) BOM позиции без совпадений (нет в ТРУ) -> ostatki
    unmatched_bom_idx = [i for i in merged_df.index if i not in merged_indices] if merged_indices else list(merged_df.index)
    if unmatched_bom_idx:
        df_unmatched_bom = merged_df.loc[unmatched_bom_idx].copy()
        if note_col not in df_unmatched_bom.columns:
            df_unmatched_bom[note_col] = ''
        # заполняем примечание только если пусто
        df_unmatched_bom[note_col] = df_unmatched_bom[note_col].astype(str)
        df_unmatched_bom.loc[df_unmatched_bom[note_col].str.strip().eq(''), note_col] = 'нет в ТРУ'
        ostatki_parts.append(df_unmatched_bom)

    # 2) Разница количества по совпавшим строкам (если qty_col найден)
    if qty_col and merged_indices:
        for i in merged_indices:
            try:
                v = merged_df.at[i, qty_col]
            except Exception:
                continue
            pair = _parse_qty_pair(v)
            if not pair:
                continue
            tru_qty, bom_qty0 = pair
            if tru_qty == bom_qty0:
                continue
            row_df = merged_df.loc[[i]].copy()
            if note_col not in row_df.columns:
                row_df[note_col] = ''
            if tru_qty < bom_qty0:
                diff = bom_qty0 - tru_qty
                row_df.at[i, qty_col] = diff
                row_df.at[i, note_col] = f'остаток (BOM {bom_qty0} > ТРУ {tru_qty})'
                ostatki_parts.append(row_df)
            else:
                diff = tru_qty - bom_qty0
                row_df.at[i, qty_col] = diff
                row_df.at[i, note_col] = f'запас (ТРУ {tru_qty} > BOM {bom_qty0})'
                zapas_parts.append(row_df)

    # 3) Несопоставленные ТРУ -> zapas
    if isinstance(unmatched_tru, pd.DataFrame) and not unmatched_tru.empty:
        zapas_parts.append(unmatched_tru.copy())

    ostatki_df = _concat_preserve_columns(ostatki_parts)
    zapas_df = _concat_preserve_columns(zapas_parts)
    if '№ п/п' in ostatki_df.columns:
        ostatki_df = ostatki_df.drop(columns=['№ п/п'])
    if '№ п/п' in zapas_df.columns:
        zapas_df = zapas_df.drop(columns=['№ п/п'])

    def _find_name_col(df: pd.DataFrame) -> Optional[str]:
        for c in ['Наименование ИВП', 'Наименование', 'наименование ивп', 'наименование']:
            if c in df.columns:
                return c
        return None

    # Сортировка плоских отчетов по приоритету категорий
    category_order = [
        'резистор',
        'конденсатор',
        'индуктивность',
        'микросхема',
        'полупроводник',
        'разъем',
        'оптический компонент',
        'модуль питания',
        'кабель',
        'свч модуль',
        'отладочная плата',
        'наши разработки',
        'другие',
        'не распределено'
    ]
    category_rank = {name: i for i, name in enumerate(category_order)}
    default_rank = len(category_order) + 1

    def _get_rank(name: str) -> int:
        if not name or pd.isna(name):
            return default_rank
        s = str(name).strip().lower()
        if not s:
            return default_rank
        first_word = s.split(' ', 1)[0]
        # Пробуем по первому слову (префиксу категории)
        if first_word in category_rank:
            return category_rank[first_word]
        # Пробуем по полному префиксу (например "свч модуль")
        for cat, rank in category_rank.items():
            if s.startswith(cat + ' '):
                return rank
        return default_rank

    def _sort_flat(df_in: pd.DataFrame) -> pd.DataFrame:
        if df_in is None or df_in.empty:
            return df_in
        name_col = _find_name_col(df_in)
        if not name_col:
            return df_in
        df = df_in.copy()
        df['_cat_rank'] = df[name_col].map(_get_rank)
        df['_name_sort'] = df[name_col].astype(str).str.lower()
        df = df.sort_values(by=['_cat_rank', '_name_sort'], ascending=[True, True]).drop(columns=['_cat_rank', '_name_sort'])
        return df.reset_index(drop=True)

    ostatki_df = _sort_flat(ostatki_df)
    zapas_df = _sort_flat(zapas_df)
    return ostatki_df, zapas_df


def extract_component_code(text: str) -> str:
    """
    Извлекает код/марку компонента из названия.
    Примеры:
    - "1564АП3У2 ЭП" → "1564АП3У2"
    - "Микросхема 1564ТП2У ЭП - АЕЯР.431200.424-07 ТУ" → "1564ТП2У"
    - "К10-17Б-Н90-0,047 мкФ ± 10%" → "К10-17Б"
    """
    if not text or pd.isna(text):
        return ""
    
    text = str(text).strip()
    
    # Паттерны для извлечения кода компонента
    patterns = [
        # Микросхемы типа 1564АП3У2, 564ИР23У1
        r'\b(\d{3,4}[А-Яа-я]{1,4}\d*[А-Яа-я]*\d*)\b',
        # Резисторы/конденсаторы типа К10-17Б, С2-29В
        r'\b([А-Я]\d{1,2}-\d{1,3}[А-Яа-я]*)\b',
        # Микросхемы типа SN74LVC8T245DWR
        r'\b([A-Z]{2,4}\d{2,4}[A-Z0-9]*)\b',
        # Общий паттерн для буквенно-цифровых кодов
        r'\b([A-Za-zА-Яа-я]{1,4}\d{2,5}[A-Za-zА-Яа-я0-9]*)\b',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    
    return ""


def extract_pure_code(text: str) -> str:
    """
    Извлекает чистый код компонента, убирая:
    - Слова-категории (Микросхема, Чип катушки индуктивности и т.п.)
    - Суффиксы типа _LW, _AB
    - Пробелы и приводит к единому формату
    
    Примеры:
    - "Чип катушки индуктивности 0603HP-47NXJ_LW" → "0603hp-47nxj"
    - "0603HP - 47NXJ" → "0603hp-47nxj"
    """
    if not text or pd.isna(text):
        return ""
    
    text = str(text).strip()
    
    # Список слов-категорий для удаления
    category_words = [
        'чип катушки индуктивности', 'чип катушка индуктивности',
        'катушка индуктивности', 'катушки индуктивности',
        'микросхема', 'микродроссель', 'дроссель',
        'конденсатор', 'резистор', 'диод', 'транзистор',
        'стабилитрон', 'индикатор единичный', 'индикатор',
        'трансформатор', 'модуль электропитания', 'модуль питания',
        'модуль передачи данных', 'модуль фильтра',
        'модуль', 'плата инструментальная', 'плата входа', 'плата', 'аттенюатор',
        'чип', 'фильтр помехоподавляющий', 'фильтр', 'реле', 'оптопара', 'кабель',
        'патч-корд', 'патч корд'
    ]
    
    # Производители для удаления (в начале и в конце строки)
    manufacturers = [
        'coilcraft', 'mini-circuits', 'texas instruments', 'analog devices',
        'hittite', 'maxim integrated', 'stmicroelectronics', 'avnet', 'ebyte',
        'api technologies', 'weinschel', 'vishay', 'yageo', 'murata', 'tdk',
        'samsung', 'intel', 'amd', 'nxp', 'infineon', 'on semiconductor',
        'microchip', 'renesas', 'broadcom', 'qualcomm', 'xilinx', 'altera',
        'hyperline', 'cablexpert'
    ]
    
    # Удаляем слова-категории (с начала)
    text_lower = text.lower()
    for word in category_words:
        if text_lower.startswith(word):
            text = text[len(word):].strip()
            text_lower = text.lower()
            
    # Заменяем кириллицу на латиницу (визуально похожие символы)
    # Это решает проблемы, когда код написан в русской раскладке (например, 0603HP vs 0603НР)
    confusables = {
        'а': 'a', 'в': 'b', 'е': 'e', 'к': 'k', 'м': 'm', 'н': 'h',
        'о': 'o', 'р': 'p', 'с': 'c', 'т': 't', 'у': 'y', 'х': 'x'
    }
    new_text_chars = []
    for char in text_lower:
        new_text_chars.append(confusables.get(char, char))
    text = "".join(new_text_chars)
    text_lower = text  # Уже в нижнем регистре
    
    # Удаляем производителей (и с начала, и с конца)
    for mfr in manufacturers:
        # С конца
        if text_lower.endswith(mfr):
            text = text[:-len(mfr)].strip()
            text_lower = text.lower()
        elif text_lower.endswith(' ' + mfr):
            text = text[:-(len(mfr) + 1)].strip()
            text_lower = text.lower()
        # С начала
        if text_lower.startswith(mfr + ' '):
            text = text[len(mfr) + 1:].strip()
            text_lower = text.lower()
        elif text_lower.startswith(mfr):
            text = text[len(mfr):].strip()
            text_lower = text.lower()
    
    # Нормализуем
    text = re.sub(r'[\u2010\u2011\u2012\u2013\u2014\u2015\u2212]', '-', text)  # Различные тире
    text = re.sub(r'\s+', '', text)  # Убираем все пробелы
    text = text.replace('_', '-')  # Подчёркивания → дефисы
    text = text.replace(':', '-')  # Двоеточие → дефис
    text = text.replace('.', '-')  # Точка → дефис
    
    # Убираем суффикс типа -LW, -AB в конце (если код > 9 символов)
    if len(text) > 9:
        text = re.sub(r'-[A-Za-z]{1,3}$', '', text)
    
    return text.lower()


def find_matching_tru_row(
    bom_name: str,
    bom_nominal: str,
    tru_df: pd.DataFrame,
    name_col: str = 'Наименование',
    min_name_similarity: float = 0.70,
    required_code: Optional[str] = None,
) -> Optional[pd.Series]:
    """
    Ищет соответствующую строку в ТРУ DataFrame.
    Использует несколько стратегий:
    1. Точное совпадение кода компонента
    2. Вхождение BOM названия в ТРУ название (или наоборот)
    3. Общая схожесть строк
    
    Args:
        bom_name: Название из BOM (Наименование ИВП)
        bom_nominal: Номинал из BOM (если есть)
        tru_df: DataFrame из файла ТРУ
        name_col: Название колонки с наименованием в ТРУ
        min_name_similarity: Минимальный порог совпадения названия (0.0-1.0)
    
    Returns:
        Строка из ТРУ или None
    """
    if name_col not in tru_df.columns:
        return None
    
    norm_bom_name = normalize_for_matching(bom_name)
    bom_code = extract_component_code(bom_name)
    required_norm = None
    if required_code:
        required_norm = str(required_code).strip().lower().replace(" ", "")
    
    # Если BOM название короткое (только код без описания типа "Микросхема" и т.п.)
    # Считаем коротким, если нет типичных слов-описаний
    description_words = ['микросхема', 'конденсатор', 'резистор', 'катушка', 'дроссель', 
                         'диод', 'транзистор', 'индикатор', 'трансформатор', 'модуль',
                         'плата', 'чип', 'стабилитрон', 'фильтр', 'реле']
    bom_lower = bom_name.lower()
    bom_is_short_code = not any(word in bom_lower for word in description_words)
    
    # Извлекаем чистый код из BOM
    bom_pure_code = extract_pure_code(bom_name)
    
    # Если в названии есть различающие слова (например, Вилка/Розетка),
    # требуем их совпадения, чтобы избежать ложных матчей по коду.
    def _extract_type_keywords(text: str) -> Set[str]:
        if not text:
            return set()
        tokens = re.findall(r'[A-Za-zА-Яа-я]+', text.lower())
        special = {'вилка', 'розетка', 'штекер', 'гнездо'}
        return {t for t in tokens if t in special}

    bom_type_keywords = _extract_type_keywords(bom_name)

    best_match = None
    best_score = 0.0
    
    for idx, row in tru_df.iterrows():
        tru_name = str(row.get(name_col, ''))
        
        # Если требуется конкретный код (например АМФИ./ГВАТ./де...), то матчим только строки,
        # где этот код реально присутствует (иначе будут ложные совпадения на "Плата 1", "Плата 2" и т.п.)
        if required_norm:
            tru_hay = tru_name.lower().replace(" ", "")
            if required_norm not in tru_hay:
                continue

        # Если есть типовые различающие слова в BOM (вилка/розетка/штекер/гнездо),
        # но в ТРУ их нет — не считаем это совпадением.
        if bom_type_keywords:
            tru_type_keywords = _extract_type_keywords(tru_name)
            if not bom_type_keywords.issubset(tru_type_keywords):
                continue

        norm_tru_name = normalize_for_matching(tru_name)
        tru_code = extract_component_code(tru_name)
        tru_pure_code = extract_pure_code(tru_name)
        
        score = 0.0
        
        # НОВАЯ Стратегия: Сравнение чистых кодов (без категорий и суффиксов)
        # Это самый надёжный способ — включает полный номер модели (напр. МДМ30-1В15ТУП)
        if bom_pure_code and tru_pure_code:
            if bom_pure_code == tru_pure_code:
                score = 0.99  # Точное совпадение — максимальный приоритет
            elif bom_pure_code in tru_pure_code or tru_pure_code in bom_pure_code:
                score = 0.97  # Один содержит другой — высокий приоритет
        
        # Стратегия 0: Если BOM — короткий код, ищем его вхождение в ТРУ
        if score < 0.9 and bom_is_short_code and norm_bom_name:
            # Проверяем нормализованную версию
            if norm_bom_name in norm_tru_name:
                score = max(score, 0.95)  # Высокий приоритет
            # Проверяем оригинальные строки (до нормализации)
            elif bom_name.strip().lower().replace(' ', '') in tru_name.lower().replace(' ', ''):
                score = max(score, 0.95)
        
        # Стратегия 1: Совпадение кода компонента (напр. МДМ30)
        # ВНИМАНИЕ: Это менее точный метод — может давать ложные совпадения!
        # Например, МДМ30-1В12 и МДМ30-1В15 оба имеют код МДМ30
        if score < 0.9 and bom_code and tru_code:
            if bom_code.lower() == tru_code.lower():
                score = max(score, 0.92)  # Ниже чем pure code чтобы избежать ложных матчей
            elif bom_code.lower() in tru_code.lower() or tru_code.lower() in bom_code.lower():
                score = max(score, 0.85)
        
        # Стратегия 2: BOM название содержится в ТРУ (или наоборот)
        if score < 0.8:
            if norm_bom_name in norm_tru_name:
                score = max(score, 0.90)
            elif norm_tru_name in norm_bom_name:
                score = max(score, 0.85)
        
        # Стратегия 3: Общая схожесть
        if score < min_name_similarity:
            sim = similarity_ratio(norm_bom_name, norm_tru_name)
            score = max(score, sim)
        
        if score < min_name_similarity:
            continue
        
        # Если есть номинал, проверяем совпадение
        if bom_nominal:
            tru_nominal = extract_nominal(tru_name)
            if tru_nominal and tru_nominal != bom_nominal:
                continue  # Номинал не совпал — пропускаем
        
        # Нашли подходящее совпадение
        if score > best_score:
            best_score = score
            best_match = row
    
    return best_match


def extract_tru_number(filename: str) -> str:
    """
    Извлекает номер ТРУ из имени файла.
    Примеры:
    - "ТРУ.953033.7471_tpy.xlsx" → "ТРУ.953033.7471"
    - "ТРУ.953033.7471.xls" → "ТРУ.953033.7471"
    - "TPY.953033.7471_тру.xlsx" → "TPY.953033.7471"
    """
    import os
    basename = os.path.splitext(os.path.basename(filename))[0]
    
    # Удаляем суффиксы типа _tpy, _тру
    basename = re.sub(r'[_-]?(tpy|тру|ТРУ)$', '', basename, flags=re.IGNORECASE)
    
    # Паттерн для ТРУ/TPY номера
    match = re.match(r'^(ТРУ|TPY|TRU)[.\s_-]?\d+[.\d]*', basename, re.IGNORECASE)
    if match:
        return match.group(0)
    
    return basename


def merge_tru_into_bom(
    bom_df: pd.DataFrame,
    tru_dfs: List[pd.DataFrame],
    tru_filenames: List[str] = None,
    bom_name_col: str = 'Наименование ИВП',
    bom_qty_col: str = 'шт.',
    tru_name_col: str = 'Наименование',
    tru_article_col: str = 'Артикул',
    tru_qty_col: str = 'Количество',
    tru_cost_col: str = 'Стоимость'
) -> Tuple[pd.DataFrame, Set[int], pd.DataFrame]:
    """
    Объединяет данные из ТРУ файлов с BOM DataFrame.
    
    Args:
        bom_df: DataFrame из BOM файла
        tru_dfs: Список DataFrame из ТРУ файлов
        tru_filenames: Список имён файлов ТРУ (для извлечения № ТРУ)
        bom_name_col: Колонка с названием в BOM
        bom_qty_col: Колонка с количеством в BOM
        tru_name_col: Колонка с названием в ТРУ
        tru_article_col: Колонка с артикулом в ТРУ
        tru_qty_col: Колонка с количеством в ТРУ
        tru_cost_col: Колонка со стоимостью в ТРУ
    
    Returns:
        (обновлённый DataFrame, множество индексов изменённых строк)
    """
    result_df = bom_df.copy()
    merged_indices: Set[int] = set()
    
    if not tru_dfs:
        return result_df, merged_indices
    
    # Добавляем имя файла в каждый ТРУ DataFrame для извлечения № ТРУ
    for i, tru_df in enumerate(tru_dfs):
        if tru_filenames and i < len(tru_filenames):
            tru_df['_tru_filename'] = tru_filenames[i]
            tru_df['_tru_number'] = extract_tru_number(tru_filenames[i])
        else:
            tru_df['_tru_filename'] = ''
            tru_df['_tru_number'] = ''
    
    # Объединяем все ТРУ в один DataFrame
    combined_tru = pd.concat(tru_dfs, ignore_index=True)
    
    if combined_tru.empty:
        return result_df, merged_indices, set()
    
    # Отслеживаем использованные индексы ТРУ
    used_tru_indices: Set[int] = set()

    # --- Подготовка: группировка строк ТРУ по ключу, чтобы корректно суммировать позиции,
    # которые встречаются в нескольких ТРУ (иначе лишнее уйдёт в *_ostatki).
    if tru_article_col in combined_tru.columns:
        combined_tru['_erp_code_norm'] = combined_tru[tru_article_col].map(normalize_erp_code_from_artikul)
    else:
        combined_tru['_erp_code_norm'] = ''

    if tru_name_col in combined_tru.columns:
        combined_tru['_pure_code'] = combined_tru[tru_name_col].map(lambda v: extract_pure_code(v))
    else:
        combined_tru['_pure_code'] = ''

    code_groups: Dict[str, List[int]] = {}
    pure_groups: Dict[str, List[int]] = {}
    try:
        for k, idxs in combined_tru.groupby('_erp_code_norm').groups.items():
            if k and str(k).strip():
                code_groups[str(k)] = list(idxs)
    except Exception:
        code_groups = {}
    try:
        for k, idxs in combined_tru.groupby('_pure_code').groups.items():
            if k and str(k).strip():
                pure_groups[str(k)] = list(idxs)
    except Exception:
        pure_groups = {}

    def _safe_float(v) -> float:
        if v is None or pd.isna(v):
            return 0.0
        if isinstance(v, (int, float)):
            try:
                return float(v)
            except Exception:
                return 0.0
        s = str(v).strip()
        if not s:
            return 0.0
        # если формат "15 (10)" -> берём первое число
        m = re.match(r'^\s*(\d+(?:[.,]\d+)?)', s)
        if m:
            s = m.group(1)
        s = s.replace('\u00A0', '').replace(' ', '').replace(',', '.')
        try:
            return float(s)
        except Exception:
            return 0.0

    def _aggregate_tru_block(tru_indices: List[int], required_code: Optional[str]) -> Tuple[pd.DataFrame, float, float, str]:
        """
        Возвращает (df_block, total_qty, total_cost, joined_tru_numbers)
        """
        if not tru_indices:
            return combined_tru.iloc[0:0], 0.0, 0.0, ''
        block = combined_tru.loc[tru_indices]

        # Если требуется конкретный код (ГВАТ/АМФИ/ИГНД/ДЕ...), оставляем только строки, где он присутствует
        if required_code:
            req = str(required_code).strip().lower().replace(' ', '')
            try:
                name_series = block.get(tru_name_col, pd.Series([''] * len(block), index=block.index))
                keep = name_series.astype(str).str.lower().str.replace(' ', '', regex=False).str.contains(req, na=False)
                block = block.loc[keep]
            except Exception:
                pass

        # qty
        qty_total = 0.0
        if tru_qty_col in block.columns:
            qty_total = float(block[tru_qty_col].map(_safe_float).sum())
        elif 'Количество' in block.columns:
            qty_total = float(block['Количество'].map(_safe_float).sum())

        # cost: sum explicit costs when present; otherwise qty*price per row
        cost_total = 0.0
        for _, r in block.iterrows():
            c_val = None
            if tru_cost_col in r.index:
                c_val = r.get(tru_cost_col, None)
            elif 'Стоимость' in r.index:
                c_val = r.get('Стоимость', None)
            c_num = _safe_float(c_val)
            if c_num > 0:
                cost_total += c_num
                continue
            q = _safe_float(r.get(tru_qty_col, r.get('Количество', None)))
            p = _safe_float(r.get('Цена', None))
            if q > 0 and p > 0:
                cost_total += q * p

        # joined TRU numbers
        tru_nums = ''
        if '_tru_number' in block.columns:
            try:
                vals = [str(v).strip() for v in block['_tru_number'].tolist() if v and not pd.isna(v) and str(v).strip()]
                uniq = sorted(set(vals))
                tru_nums = '; '.join(uniq)
            except Exception:
                tru_nums = ''

        return block, qty_total, cost_total, tru_nums
    
    # Убедимся что нужные колонки существуют в BOM
    if 'КОД ERP(МР)' not in result_df.columns:
        result_df['КОД ERP(МР)'] = ''
    if 'Стоимость' not in result_df.columns:
        result_df['Стоимость'] = ''
    if '№ ТРУ' not in result_df.columns:
        result_df['№ ТРУ'] = ''
    
    # Преобразуем колонки в object для совместимости с разными типами данных
    result_df['КОД ERP(МР)'] = result_df['КОД ERP(МР)'].astype(object)
    result_df['Стоимость'] = result_df['Стоимость'].astype(object)
    result_df['№ ТРУ'] = result_df['№ ТРУ'].astype(object)
    if bom_qty_col in result_df.columns:
        result_df[bom_qty_col] = result_df[bom_qty_col].astype(object)
    
    # Обрабатываем каждую строку BOM
    for idx, bom_row in result_df.iterrows():
        bom_name = bom_row.get(bom_name_col, '')
        
        if not bom_name or pd.isna(bom_name):
            continue
        
        # Если в BOM есть "наш код" в колонке ТУ/Производитель — используем его как обязательный ключ матчинга.
        required_code = None
        try:
            for tu_col in ['ТУ/Производитель', 'ТУ', 'ту', '_extracted_tu_']:
                if tu_col in result_df.columns:
                    v = bom_row.get(tu_col, '')
                    if v and not pd.isna(v):
                        vv = str(v).strip()
                        if re.search(r'\b(гват|амфи|игнд)\.\d+(?:\.\d+)+\b|\bде\s*\d+(?:\.\d+){0,3}\b', vv, re.IGNORECASE):
                            required_code = re.sub(r'\s+', '', vv)
                            break
        except Exception:
            required_code = None

        tru_match = find_matching_tru_row(
            bom_name=str(bom_name),
            bom_nominal=extract_nominal(str(bom_name)),
            tru_df=combined_tru,
            name_col=tru_name_col,
            required_code=required_code,
        )
        
        if tru_match is None:
            continue
        
        # Нашли совпадение — обновляем данные
        merged_indices.add(idx)
        
        # Запоминаем индекс использованной строки ТРУ
        # (важно: если позиция встречается в нескольких ТРУ, помечаем использованными ВСЕ такие строки,
        # иначе "хвост" попадёт в *_zapas и/или создаст ложные *_ostatki по количеству)
        matched_tru_block = None
        tru_qty_num_agg = None
        tru_cost_agg = None
        tru_nums_joined = None

        match_norm_code = ''
        try:
            if tru_article_col in tru_match.index:
                match_norm_code = normalize_erp_code_from_artikul(tru_match.get(tru_article_col, None))
        except Exception:
            match_norm_code = ''

        match_pure_code = ''
        try:
            if tru_match.name is not None and '_pure_code' in combined_tru.columns:
                match_pure_code = str(combined_tru.at[tru_match.name, '_pure_code'] or '').strip()
            elif tru_name_col in tru_match.index:
                match_pure_code = str(extract_pure_code(tru_match.get(tru_name_col, '')) or '').strip()
        except Exception:
            match_pure_code = ''

        group_indices: List[int] = []
        if match_norm_code and match_norm_code in code_groups:
            group_indices = code_groups.get(match_norm_code, [])
        elif match_pure_code and match_pure_code in pure_groups:
            group_indices = pure_groups.get(match_pure_code, [])
        elif tru_match.name is not None:
            group_indices = [tru_match.name]

        matched_tru_block, tru_qty_num_agg, tru_cost_agg, tru_nums_joined = _aggregate_tru_block(group_indices, required_code)
        if matched_tru_block is not None and not matched_tru_block.empty:
            used_tru_indices.update(set(matched_tru_block.index))
        elif tru_match.name is not None:
            used_tru_indices.add(tru_match.name)
        
        # 1. КОД ERP(МР) ← Артикул (без .0)
        if tru_article_col in tru_match.index:
            article = tru_match.get(tru_article_col, None)
            norm_code = normalize_erp_code_from_artikul(article)
            if norm_code:
                result_df.at[idx, 'КОД ERP(МР)'] = norm_code
            elif article and not pd.isna(article):
                result_df.at[idx, 'КОД ERP(МР)'] = str(article).strip()
        
        # 2. Стоимость ← Стоимость из ТРУ (или вычисляем как Количество × Цена)
        # Ищем колонку стоимости в tru_match
        cost_value = None
        qty_value = None
        price_value = None
        
        for col in tru_match.index:
            col_lower = str(col).lower()
            if 'стоимость' in col_lower:
                cost_value = tru_match[col]
            elif col_lower in ['количество', 'qty', 'кол-во']:
                qty_value = tru_match[col]
            elif 'цена' in col_lower or col_lower == 'price':
                price_value = tru_match[col]
        
        # Если стоимость пустая, вычисляем из количества и цены
        if (cost_value is None or pd.isna(cost_value)) and qty_value is not None and price_value is not None:
            try:
                qty_num = float(qty_value) if not pd.isna(qty_value) else 0
                price_num = float(price_value) if not pd.isna(price_value) else 0
                if qty_num > 0 and price_num > 0:
                    cost_value = qty_num * price_num
            except (ValueError, TypeError):
                pass
        
        # Если есть агрегированная стоимость по нескольким ТРУ — используем её
        if tru_cost_agg is not None and isinstance(tru_cost_agg, (int, float)) and tru_cost_agg > 0:
            try:
                result_df.at[idx, 'Стоимость'] = int(round(float(tru_cost_agg)))
            except Exception:
                result_df.at[idx, 'Стоимость'] = tru_cost_agg
        elif cost_value is not None and not pd.isna(cost_value):
            try:
                cost_num = int(round(float(cost_value)))  # Округляем до целого
                result_df.at[idx, 'Стоимость'] = cost_num
            except (ValueError, TypeError):
                result_df.at[idx, 'Стоимость'] = cost_value
        
        # 3. № ТРУ ← из имени файла
        # Если позиция найдена сразу в нескольких ТРУ — сохраняем все номера
        if tru_nums_joined:
            result_df.at[idx, '№ ТРУ'] = tru_nums_joined
        elif '_tru_number' in tru_match.index:
            tru_num = tru_match['_tru_number']
            if tru_num and not pd.isna(tru_num):
                result_df.at[idx, '№ ТРУ'] = str(tru_num)
        
        # 4. Количество: TRU_qty (BOM_qty) если разное
        if bom_qty_col in result_df.columns:
            tru_qty = None
            if tru_qty_num_agg is not None and isinstance(tru_qty_num_agg, (int, float)) and tru_qty_num_agg > 0:
                tru_qty = tru_qty_num_agg
            elif tru_qty_col in tru_match.index:
                tru_qty = tru_match[tru_qty_col]
            bom_qty = bom_row.get(bom_qty_col, '')
            
            if tru_qty and not pd.isna(tru_qty):
                try:
                    tru_qty_num = _safe_float(tru_qty)
                    bom_qty_num = _safe_float(bom_qty) if bom_qty and not pd.isna(bom_qty) else 0
                    
                    if tru_qty_num != bom_qty_num and bom_qty_num > 0:
                        # Формат: "15 (10)" где 15 — из ТРУ, 10 — исходное
                        new_qty_str = f"{int(tru_qty_num)} ({int(bom_qty_num)})"
                        result_df.at[idx, bom_qty_col] = new_qty_str
                    elif bom_qty_num == 0:
                        result_df.at[idx, bom_qty_col] = int(tru_qty_num)
                except (ValueError, TypeError):
                    pass
    
    # Находим несопоставленные строки ТРУ
    all_tru_indices = set(combined_tru.index)
    unmatched_indices = all_tru_indices - used_tru_indices
    
    # Создаём DataFrame с несопоставленными ТРУ в формате BOM
    if unmatched_indices:
        unmatched_raw = combined_tru.loc[list(unmatched_indices)].copy()
        
        # Фильтруем пустые строки (где нет наименования)
        if 'Наименование' in unmatched_raw.columns:
            unmatched_raw = unmatched_raw[
                unmatched_raw['Наименование'].notna() & 
                (unmatched_raw['Наименование'].astype(str).str.strip() != '')
            ]
        
        # Если после фильтрации ничего не осталось
        if unmatched_raw.empty:
            return result_df, merged_indices, used_tru_indices
        
        # Известные производители для извлечения
        known_manufacturers = [
            'Coilcraft', 'Mini-Circuits', 'Texas Instruments', 'Analog Devices',
            'HITTITE', 'Maxim Integrated', 'STMicroelectronics', 'Avnet', 'Ebyte',
            'API Technologies', 'WEINSCHEL', 'Vishay', 'Yageo', 'Murata', 'TDK',
            'Samsung', 'Intel', 'AMD', 'NXP', 'Infineon', 'ON Semiconductor',
            'Microchip', 'Renesas', 'Broadcom', 'Qualcomm', 'Xilinx', 'Altera'
        ]
        
        def extract_manufacturer(name):
            """Извлекает производителя из наименования"""
            if not name or pd.isna(name):
                return ''
            name_str = str(name)
            for mfr in known_manufacturers:
                if mfr.lower() in name_str.lower():
                    return mfr
            return ''
        
        def clean_name(name, manufacturer):
            """Удаляет производителя из наименования"""
            if not name or pd.isna(name) or not manufacturer:
                return name
            # Удаляем производителя из названия
            import re
            pattern = re.compile(re.escape(manufacturer), re.IGNORECASE)
            cleaned = pattern.sub('', str(name)).strip()
            # Убираем лишние пробелы
            cleaned = ' '.join(cleaned.split())
            return cleaned
            
    return result_df, merged_indices, used_tru_indices


def generate_unmatched_report(
    tru_dfs: List[pd.DataFrame],
    used_tru_indices: Set[int],
    tru_name_col: str = 'Наименование'
) -> pd.DataFrame:
    """
    Генерирует DataFrame с несопоставленными элементами ТРУ.
    """
    # Хелперы
    def calc_cost(row):
        """Вычисляет стоимость как Количество × Цена"""
        qty = row.get('Количество', 0)
        price = row.get('Цена', 0)
        cost = row.get('Стоимость', None)
        
        # Если стоимость уже есть и не пустая
        if cost and not pd.isna(cost):
            try:
                return int(round(float(cost)))
            except (ValueError, TypeError):
                pass
        
        # Вычисляем
        try:
            qty_num = float(qty) if qty and not pd.isna(qty) else 0
            price_num = float(price) if price and not pd.isna(price) else 0
            if qty_num > 0 and price_num > 0:
                return int(round(qty_num * price_num))
        except (ValueError, TypeError):
            pass
        return ''

    def get_tu_or_manufacturer(name):
        """Извлекает ТУ код или производителя"""
        if not name or pd.isna(name):
            return '', name
        
        # Сначала пробуем извлечь ТУ код
        clean_name_result, tu_code = extract_tu_code(str(name))
        if tu_code:
            return tu_code, clean_name_result
        
        # Ищем производителя (нужны known_manufacturers - возьмем глобально или определим тут)
        # Для простоты повторим логику
        known_manufacturers = [
            'Murata', 'Yageo', 'Samsung', 'TDK', 'AVX', 'Kemet', 'Vishay', 'Coilcraft',
            'Bourns', 'Panasonic', 'Rubycon', 'Nichicon', 'Epcos', 'Wurth', 'Molex',
            'TE Connectivity', 'Samtec', 'Harting', 'Phoenix Contact', 'Degson',
            'Omron', 'Song Chuan', 'Hongfa', 'Relpol', 'Finder', 'STM', 'STMicroelectronics',
            'TI', 'Texas Instruments', 'NXP', 'Microchip', 'Analog Devices', 'Maxim',
            'Infineon', 'Onsemi', 'Diodes', 'Nexperia', 'Toshiba', 'Renesas', 'Xilinx',
            'Altera', 'Intel', 'Lattice', 'Cypress', 'Micron', 'Winbond', 'ISSI',
            'Mean Well', 'Mornsun', 'Traco Power', 'Recom', 'Vicor', 'TDK-Lambda',
            'Chinfa', 'Delta', 'Arch', 'Hi-Link', 'Kingbright', 'Everlight', 'Lite-On',
            'Osram', 'Cree', 'Seoul', 'Edison', 'Lequn', 'Hittite', 'Mini-Circuits'
        ]
        
        def clean_name_mfr(name, manufacturer):
            import re
            pattern = re.compile(re.escape(manufacturer), re.IGNORECASE)
            cleaned = pattern.sub('', str(name)).strip()
            cleaned = ' '.join(cleaned.split())
            return cleaned

        name_str = str(name)
        for mfr in known_manufacturers:
            if mfr.lower() in name_str.lower():
                return mfr, clean_name_mfr(name_str, mfr)
        return '', name

    # Объединяем ТРУ файлы
    dfs = []
    for df in tru_dfs:
        dfs.append(df.copy())
    
    if not dfs:
        return pd.DataFrame()
        
    combined_tru = pd.concat(dfs, ignore_index=True)
    
    # Находим несопоставленные строки
    all_tru_indices = set(combined_tru.index)
    unmatched_indices = all_tru_indices - used_tru_indices
    
    if unmatched_indices:
        unmatched_raw = combined_tru.loc[list(unmatched_indices)].copy()
        
        # Фильтруем пустые строки
        if tru_name_col in unmatched_raw.columns:
            unmatched_raw = unmatched_raw[unmatched_raw[tru_name_col].notna() & (unmatched_raw[tru_name_col] != '')]
        
        tu_or_mfr_list = []
        clean_names_list = []
        
        for name in unmatched_raw.get(tru_name_col, []):
            tu_or_mfr, clean_nm = get_tu_or_manufacturer(name)
            tu_or_mfr_list.append(tu_or_mfr)
            clean_names_list.append(clean_nm)
        
        # Преобразуем в формат BOM: Наименование ИВП, ТУ, шт., КОД ERP(МР), Источник, Примечание, № ТРУ, Стоимость
        unmatched_tru = pd.DataFrame()
        unmatched_tru['Наименование ИВП'] = clean_names_list
        unmatched_tru['ТУ'] = tu_or_mfr_list
        unmatched_tru['шт.'] = unmatched_raw.get('Количество', '').values
        
        # Обработка артикула (КОД ERP)
        if 'Артикул' in unmatched_raw.columns:
            unmatched_tru['КОД ERP(МР)'] = unmatched_raw['Артикул'].map(normalize_erp_code_from_artikul)
        else:
            unmatched_tru['КОД ERP(МР)'] = ''
            
        unmatched_tru['Источник'] = 'ТРУ (не найден в BOM)'
        unmatched_tru['Примечание'] = ''
        unmatched_tru['№ ТРУ'] = unmatched_raw.get('_tru_number', '').values if '_tru_number' in unmatched_raw.columns else ''
        
        # Вычисляем стоимость
        unmatched_tru['Стоимость'] = unmatched_raw.apply(calc_cost, axis=1).values
        
        return unmatched_tru
    else:
        return pd.DataFrame()


def apply_merge_styles(
    worksheet,
    merged_rows: Set[int],
    name_col_idx: int = 2,
    qty_col_idx: int = 4,
    header_row: int = 1
):
    """
    Применяет стили к Excel worksheet:
    - Заголовки: светлая заливка, чёрный жирный шрифт
    - Все ячейки: сплошные границы
    - Объединённые строки: бледно-голубая заливка
    
    Args:
        worksheet: openpyxl worksheet
        merged_rows: Множество индексов строк (0-based из pandas)
        name_col_idx: Индекс колонки "Наименование ИВП" (1-based)
        qty_col_idx: Индекс колонки "шт." (1-based)
        header_row: Номер строки заголовка
    """
    # Стили
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Заголовок: светло-синяя заливка, чёрный жирный шрифт
    header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    header_font_normal = Font(bold=True, color='000000', size=11)  # Увеличен на 1
    header_font_small = Font(bold=True, color='000000', size=10)   # Для КОД
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Выравнивание для данных
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # Определяем индексы колонок для центрирования и выравнивания по правому краю
    center_cols = set()  # №, шт., № ТРУ, КОД ERP(МР), ОКПД
    right_cols = set()   # (не используется)
    left_cols = set()    # Стоимость
    code_col_idx = None  # Индекс колонки КОД
    
    # Находим нужные колонки по заголовкам
    for col_idx in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=header_row, column=col_idx)
        header_value = str(header_cell.value).strip().lower() if header_cell.value else ''
        
        if header_value in ['№ п/п', '№', 'n']:
            center_cols.add(col_idx)
        elif header_value in ['шт.', 'шт', 'qty', 'количество']:
            center_cols.add(col_idx)
        elif header_value in ['№ тру', 'тру']:
            center_cols.add(col_idx)
        elif 'окпд' in header_value:
            center_cols.add(col_idx)
        elif 'код' in header_value or 'erp' in header_value:
            center_cols.add(col_idx)
            code_col_idx = col_idx
        elif 'стоимость' in header_value:
            left_cols.add(col_idx)  # Стоимость — влево
    
    # Применяем стили к заголовкам
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_alignment
        
        # Уменьшенный шрифт для КОД
        if col_idx == code_col_idx:
            cell.font = header_font_small
        else:
            cell.font = header_font_normal
    
    # Применяем границы и выравнивание ко ВСЕМ строкам данных
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            # Выравнивание
            if col_idx in center_cols:
                cell.alignment = center_alignment
            elif col_idx in left_cols:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col_idx in right_cols:
                cell.alignment = right_alignment
    
    # Применяем стили к объединённым строкам (из ТРУ)
    for pandas_idx in merged_rows:
        # Конвертируем pandas index (0-based) в Excel row (1-based + header)
        excel_row = pandas_idx + header_row + 1
        
        # Применяем стили ко всей строке
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=excel_row, column=col_idx)
            
            # Заливка — бледно-голубая
            cell.fill = MERGED_ROW_FILL
            cell.border = thin_border
            
            # Шрифт наименования — тёмно-синий
            if col_idx == name_col_idx:
                cell.font = MERGED_NAME_FONT
            
            # Количество с форматом "X (Y)" — частичное форматирование
            if col_idx == qty_col_idx:
                cell_value = str(cell.value) if cell.value else ''
                if '(' in cell_value and ')' in cell_value:
                    # Разбиваем на части: "20 (2)" → "20 " и "(2)"
                    match = re.match(r'^(\d+)\s*(\([^)]+\))$', cell_value)
                    if match:
                        new_qty = match.group(1) + " "  # "20 "
                        old_qty = match.group(2)        # "(2)"
                        
                        # Создаём RichText с разными стилями
                        try:
                            black_font = InlineFont(color='000000')
                            red_font = InlineFont(color='8B0000', b=True)  # Тёмно-красный, жирный
                            
                            rich_text = CellRichText(
                                TextBlock(black_font, new_qty),
                                TextBlock(red_font, old_qty)
                            )
                            cell.value = rich_text
                        except Exception:
                            # Fallback — весь текст красным
                            cell.font = Font(color='8B0000', bold=True)
    
    # Установка ширины колонок (заданные ширины для стандартных колонок BOM)
    standard_widths = {
        '№ п/п': 6,
        '№': 6,
        'наименование ивп': 55,
        'наименование': 55,
        'ту': 35,
        'ту/производитель': 35,
        'шт.': 8,
        'шт': 8,
        'код erp(мр)': 15,
        'код erp': 15,
        # Эти колонки часто содержат длинные значения (списки файлов/позиций),
        # поэтому делаем шире, чтобы не резало при печати/PDF.
        'источник': 40,
        'примечание': 30,
        'документы': 49,
        'поставщик': 32,
        '№ тру': 20,
        'стоимость': 15
    }

    for col_idx in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=header_row, column=col_idx)
        header_val = str(header_cell.value).strip().lower() if header_cell.value else ''
        column_letter = header_cell.column_letter
        
        # Если колонка стандартная — ставим фиксированную ширину
        if header_val in standard_widths:
            worksheet.column_dimensions[column_letter].width = standard_widths[header_val]
        else:
            # Автоподбор ширины для неизвестных колонок
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                cell_val = worksheet.cell(row=row, column=col_idx).value
                if cell_val:
                    max_length = max(max_length, len(str(cell_val)))
            
            adjusted_width = min(max_length + 2, 80)  # Max 80 символов (для длинных "Документы/Источник")
            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)  # Min 8
