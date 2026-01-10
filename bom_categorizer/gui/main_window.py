# -*- coding: utf-8 -*-
"""
GUI для BOM Categorizer на базе PySide6

PySide6-интерфейс с поддержкой:
- Выбора входных файлов (XLSX, DOCX, TXT)
- Настройки параметров обработки
- Интерактивной классификации нераспределенных элементов
- PIN-защиты интерфейса
"""

import os
import json
import sys
import platform
import re
from datetime import datetime
from typing import Dict, Optional, List
from openpyxl.styles import Alignment, Font

from bom_categorizer.tru_rkm_processor import process_tru_rkm_files

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QPushButton, QLabel, QLineEdit,
    QListWidget, QListWidgetItem, QSpinBox, QCheckBox, QTextEdit, QTextBrowser,
    QFileDialog, QMessageBox, QScrollArea, QFrame, QDialog, QMenuBar, QMenu,
    QProgressDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView
)
from PySide6.QtCore import Qt, Signal, QThread, QSize, QUrl
from PySide6.QtGui import QFont, QColor, QPalette, QAction, QActionGroup, QKeySequence, QDragEnterEvent, QDropEvent, QCursor
import subprocess

from ..component_database import (
    add_component_to_database,
    get_database_path,
    get_database_stats,
    get_database_history,
    export_database_to_excel,
    import_database_from_excel,
    backup_database,
    clear_database,
    set_database_version,
    is_first_run,
    initialize_database_from_template,
    format_history_tooltip,
    CATEGORY_NAMES
)

from ..config_manager import initialize_all_configs

from .dialogs import (
    PinDialog,
    DatabaseStatsDialog,
    FirstRunImportDialog,
    ClassificationDialog,
    DocConversionDialog
)

from ..excel_writer import apply_excel_styles

from ..styles import DARK_THEME, LIGHT_THEME

# Импорты из новых модулей
from .workers import ProcessingWorker, ComparisonWorker, TruRkmWorker
from .search import GlobalSearchDialog
from . import sections
from . import search_methods

# Import from shared module to avoid duplication
from ..shared.config import get_config_path, load_config
from ..shared.fonts import get_system_font

# Import mixins for modular code organization
from .database_handlers import DatabaseHandlersMixin
from .file_handlers import FileHandlersMixin
from .help_dialogs import HelpDialogsMixin
from .processing_handlers import ProcessingHandlersMixin




class BOMCategorizerMainWindow(ProcessingHandlersMixin, HelpDialogsMixin, FileHandlersMixin, DatabaseHandlersMixin, QMainWindow):
    """Главное окно приложения BOM Categorizer на PySide6"""

    def __init__(self):
        super().__init__()

        # Сохраняем ссылку на QApplication для масштабирования
        self.app = QApplication.instance()

        # Загружаем конфигурацию
        self.cfg = load_config()
        self.config = self.cfg  # Псевдоним для совместимости
        ver = self.cfg.get("app_info", {}).get("version", "dev")
        name = self.cfg.get("app_info", {}).get("description", "BOM Categorizer")

        # Устанавливаем заголовок окна
        self.setWindowTitle(f"{name} v{ver}")

        # Размер окна будет установлен после определения режима

        # Переменные состояния
        self.input_files: Dict[str, int] = {}  # {путь_к_файлу: количество}
        self.tru_rkm_files: List[str] = []  # Файлы ТРУ и РКМ (только .xls)
        self.output_xlsx = "categorized.xlsx"
        self.txt_dir = ""
        self.combine = True
        self.interactive = False
        self.create_txt = False
        self.current_file_multiplier = 1
        self.selected_file_index: Optional[int] = None
        self.processing_dialog_ref = None  # Ссылка на диалог обработки (для плавного перехода)
        self.last_input_file = None  # Последний добавленный входной файл (для истории БД)
        self.last_generated_output = None  # Последний сгенерированный файл (для экспорта в PDF)

        # Сравнение файлов
        self.compare_file1 = ""
        self.compare_file2 = ""
        self.compare_output = "comparison.xlsx"

        # PIN защита
        self.unlocked = False
        self.require_pin = self.cfg.get("security", {}).get("require_pin", True)
        self.correct_pin = self.cfg.get("security", {}).get("pin", "1234")
        self.lockable_widgets = []

        # Тема интерфейса
        self.current_theme = self.cfg.get("ui", {}).get("theme", "dark")  # "dark" или "light"

        # Настройки отображения
        # На macOS используем размеры сопоставимые со стандартными приложениями
        # ВНИМАНИЕ: Глобальный шрифт уже установлен в main(), здесь только для локального использования
        if platform.system() == 'Darwin':  # macOS
            # Проверяем, есть ли Retina дисплей (devicePixelRatio >= 2)
            try:
                from PySide6.QtGui import QGuiApplication
                screens = QGuiApplication.screens()
                if screens and screens[0].devicePixelRatio() >= 2:
                    # Retina дисплей: 13pt (стандарт для macOS)
                    self.base_font_size = 13
                else:
                    # Обычный дисплей
                    self.base_font_size = 12
            except:
                # Если не удалось определить, используем стандартный размер
                self.base_font_size = 13
        else:  # Windows и Linux
            self.base_font_size = 12
        
        self.scale_levels: List[float] = [0.7, 0.8, 0.9, 1.0, 1.1, 1.25, 1.5]
        ui_settings = self.cfg.get("ui", {})
        # Дефолтный scale_factor: 1.0 для всех платформ (можно настроить в меню)
        default_scale = 1.0 if platform.system() == 'Darwin' else 0.8
        self.scale_factor = ui_settings.get("scale_factor", default_scale)
        if self.scale_factor not in self.scale_levels:
            # Если значение некорректное, используем дефолт для ОС
            self.scale_factor = default_scale

        self.current_view_mode = ui_settings.get("view_mode", "advanced")
        # Экспертный режим не сохраняется между запусками - всегда возвращаемся к расширенному
        if self.current_view_mode == "expert":
            self.current_view_mode = "advanced"
        # Проверка на корректность режима
        if self.current_view_mode not in ("simple", "advanced", "expert"):
            self.current_view_mode = "advanced"

        # Сохраняем предпочтение пользователя, но при требовании PIN блокируем доступ к продвинутым режимам
        self.preferred_view_mode = self.current_view_mode
        self.pin_forced_simple = False
        if self.require_pin and self.preferred_view_mode != "simple":
            self.pin_forced_simple = True
            self.current_view_mode = "simple"

        # Устанавливаем размер окна в зависимости от режима
        self._apply_window_size_for_mode(self.current_view_mode)

        # Дополнительные опции отображения
        self.log_with_timestamps = bool(ui_settings.get("log_timestamps", False)) if self.current_view_mode == "expert" else False
        self.auto_open_output = bool(ui_settings.get("auto_open_output", False)) if self.current_view_mode == "expert" else False
        self.auto_export_pdf = bool(ui_settings.get("auto_export_pdf", False)) if self.current_view_mode == "expert" else False
        
        # AI-подсказки
        self.ai_classifier_enabled = bool(ui_settings.get("ai_classifier_enabled", False)) if self.current_view_mode == "expert" else False
        self.ai_auto_classify = bool(ui_settings.get("ai_auto_classify", False)) if self.current_view_mode == "expert" else False

        # Плейсхолдеры для элементов меню и секций
        self.scale_actions: Dict[float, QAction] = {}
        self.view_mode_actions: Dict[str, QAction] = {}
        self.db_menu: Optional[QMenu] = None
        self.mode_label: Optional[QLabel] = None
        self.timestamp_checkbox: Optional[QCheckBox] = None
        self.auto_open_output_checkbox: Optional[QCheckBox] = None

        # Применяем стили
        self._setup_styles()

        # Создаем UI
        self._create_ui()

        # Создаем меню
        self._create_menu()

        # Применяем масштаб после создания всех виджетов
        self.apply_scale_factor()
        
        # Обновляем галочки в меню режимов (после создания меню)
        self.update_view_mode_actions()
        
        # Применяем режим работы из конфига (скрываем/показываем панели)
        self.apply_view_mode(initial=True)
        
        # Обновляем статус AI (активирует чекбоксы если AI настроен)
        self.update_ai_status()

        # Включаем поддержку Drag & Drop
        self.setAcceptDrops(True)

        # Применяем блокировку интерфейса при необходимости
        if self.require_pin:
            self.lock_interface()

    def _apply_window_size_for_mode(self, mode: str):
        """Устанавливает размер окна в зависимости от режима"""
        if mode == "simple":
            self.resize(600, 400)
        else:
            # Advanced и Expert режимы
            self.resize(800, 560)

    def _setup_styles(self):
        """Настраивает стили приложения с поддержкой темной и светлой темы"""
        # Применяем масштаб (будет применен после создания UI)
        # Тема применяется сразу
        self.apply_theme()

    def apply_theme(self):
        """Применяет выбранную тему к приложению"""
        if self.current_theme == "dark":
            theme_style = DARK_THEME
        else:
            theme_style = LIGHT_THEME
        
        # На macOS удаляем все font-size из стилей, чтобы использовались
        # программно установленные размеры (для правильной работы на Retina)
        if platform.system() == 'Darwin':  # macOS
            # Удаляем все строки с font-size из CSS
            import re
            # Удаляем font-size: XXpt; из стилей
            theme_style = re.sub(r'\s*font-size:\s*\d+pt;', '', theme_style)
        
        self.setStyleSheet(theme_style)

    def toggle_theme(self):
        """Переключает между темной и светлой темой"""
        # Переключаем тему
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        
        # Применяем новую тему
        self.apply_theme()
        
        # Сохраняем выбор в конфиг
        self.save_theme_preference()
        
        # Показываем уведомление
        theme_name = "Темная" if self.current_theme == "dark" else "Светлая"
        QMessageBox.information(
            self,
            "Тема изменена",
            f"{theme_name} тема применена успешно!"
        )

    def save_theme_preference(self):
        """Сохраняет выбор темы в конфигурационный файл"""
        self.save_ui_preferences()

    def _create_menu(self):
        """Создает меню приложения"""
        menubar = self.menuBar()
        
        # Меню "Файл"
        file_menu = menubar.addMenu("Файл")
        
        # Открыть файлы
        self.open_action = QAction("📂 Открыть файлы", self)
        self.open_action.setShortcut(QKeySequence("Ctrl+O"))
        self.open_action.triggered.connect(self.on_add_files)
        file_menu.addAction(self.open_action)
        
        file_menu.addSeparator()
        
        # Запустить обработку
        self.run_action = QAction("🚀 Запустить обработку", self)
        self.run_action.setShortcut(QKeySequence("Ctrl+R"))
        self.run_action.triggered.connect(self.on_run)
        file_menu.addAction(self.run_action)
        
        file_menu.addSeparator()
        
        # Выход
        exit_action = QAction("🚪 Выход", self)
        exit_action.setShortcut(QKeySequence("Ctrl+Q"))
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Меню "Вид"
        view_menu = menubar.addMenu("Вид")
        
        # Подменю масштаба
        scale_menu = view_menu.addMenu("Масштабирование интерфейса")
        scale_group = QActionGroup(self)
        scale_group.setExclusive(True)

        scale_labels = {
            0.7: "Масштаб 70%",
            0.8: "Масштаб 80% (по умолчанию)",
            0.9: "Масштаб 90%",
            1.0: "Масштаб 100%",
            1.1: "Масштаб 110%",
            1.25: "Масштаб 125%",
        }

        self.scale_actions.clear()
        for factor in self.scale_levels:
            label = scale_labels.get(factor, f"Масштаб {int(factor * 100)}%")
            action = QAction(label, self)
            action.setCheckable(True)
            action.triggered.connect(lambda checked, f=factor: self.set_scale_factor(f))
            scale_menu.addAction(action)
            scale_group.addAction(action)
            self.scale_actions[factor] = action

        view_menu.addSeparator()

        zoom_in_action = QAction("Увеличить масштаб (Ctrl++)", self)
        zoom_in_action.setShortcut(QKeySequence("Ctrl+="))  # = это то же, что + без Shift
        zoom_in_action.triggered.connect(self.on_zoom_in)
        view_menu.addAction(zoom_in_action)

        zoom_out_action = QAction("Уменьшить масштаб (Ctrl+-)", self)
        zoom_out_action.setShortcut(QKeySequence("Ctrl+-"))  # Только один вариант
        zoom_out_action.triggered.connect(self.on_zoom_out)
        view_menu.addAction(zoom_out_action)



        view_menu.addSeparator()

        # Подменю режимов работы
        self.mode_menu = view_menu.addMenu("Режим работы")
        mode_group = QActionGroup(self)
        mode_group.setExclusive(True)

        mode_definitions = [
            ("simple", "Простой режим"),
            ("advanced", "Расширенный режим (все функции)"),
            ("expert", "Экспертный режим (дополнительные настройки)"),
        ]

        self.view_mode_actions.clear()
        # На Windows используем Ctrl+1/2/3, на macOS — Meta (Control)
        # Meta на macOS это физическая клавиша Control, Ctrl — это Command
        if platform.system() == 'Windows':
            shortcuts = ["Ctrl+1", "Ctrl+2", "Ctrl+3"]
        else:
            shortcuts = ["Meta+1", "Meta+2", "Meta+3"]
        for i, (key, label) in enumerate(mode_definitions):
            action = QAction(label, self)
            action.setCheckable(True)
            if i < len(shortcuts):
                action.setShortcut(shortcuts[i])
            action.triggered.connect(lambda checked, m=key: self.set_view_mode(m))
            self.mode_menu.addAction(action)
            mode_group.addAction(action)
            self.view_mode_actions[key] = action

        # Ограничиваем доступ к режимам до ввода PIN
        self.update_mode_action_permissions()

        view_menu.addSeparator()

        # Пункт переключения темы
        theme_action = QAction("🌓 Переключить тему", self)
        theme_action.setShortcut("Ctrl+T")
        theme_action.triggered.connect(self.toggle_theme)
        view_menu.addAction(theme_action)
        
        # Меню "База данных"
        self.db_menu = menubar.addMenu("База данных")
        
        # Статистика БД
        stats_action = QAction("📊 Статистика", self)
        stats_action.triggered.connect(self.show_database_stats)
        self.db_menu.addAction(stats_action)
        
        # Экспорт БД
        export_action = QAction("📤 Экспорт в Excel", self)
        export_action.triggered.connect(self.export_database)
        self.db_menu.addAction(export_action)
        
        # Импорт БД
        import_action = QAction("📥 Импорт из Excel", self)
        import_action.triggered.connect(self.import_database)
        self.db_menu.addAction(import_action)
        
        self.db_menu.addSeparator()
        
        # Резервное копирование
        backup_action = QAction("💾 Резервное копирование", self)
        backup_action.triggered.connect(self.backup_database)
        self.db_menu.addAction(backup_action)
        
        # Открыть папку БД
        folder_action = QAction("📁 Открыть папку БД", self)
        folder_action.triggered.connect(self.open_database_folder)
        self.db_menu.addAction(folder_action)
        
        self.db_menu.addSeparator()
        
        # Посмотреть базу
        view_action = QAction("👁️ Посмотреть базу", self)
        view_action.triggered.connect(self.on_view_database)
        self.db_menu.addAction(view_action)
        
        # Изменить версию БД
        version_action = QAction("🔢 Изменить версию БД", self)
        version_action.triggered.connect(self.on_change_database_version)
        self.db_menu.addAction(version_action)
        
        # Очистить базу данных
        clear_action = QAction("🗑️ Очистить базу данных", self)
        clear_action.triggered.connect(self.on_clear_database)
        self.db_menu.addAction(clear_action)
        
        self.db_menu.addSeparator()
        
        # Заменить БД
        replace_action = QAction("🔄 Заменить БД", self)
        replace_action.triggered.connect(self.on_replace_database)
        self.db_menu.addAction(replace_action)
        
        # Добавить все из выходного файла
        import_output_action = QAction("📋 Добавить из выходного файла", self)
        import_output_action.triggered.connect(self.on_import_from_output)
        self.db_menu.addAction(import_output_action)
        
        # Меню "Помощь"
        help_menu = menubar.addMenu("Помощь")
        
        # Контекстная помощь
        context_help_action = QAction("❓ Контекстная помощь", self)
        context_help_action.setShortcut(QKeySequence("F1"))
        context_help_action.triggered.connect(self.show_context_help)
        help_menu.addAction(context_help_action)
        
        # База знаний
        knowledge_base_action = QAction("📚 База знаний", self)
        knowledge_base_action.triggered.connect(self.show_knowledge_base)
        help_menu.addAction(knowledge_base_action)
        
        help_menu.addSeparator()
        
        # Руководство по Drag & Drop
        dragdrop_help_action = QAction("🎯 Как использовать Drag & Drop", self)
        dragdrop_help_action.triggered.connect(self.show_dragdrop_help)
        help_menu.addAction(dragdrop_help_action)
        
        help_menu.addSeparator()
        
        # О программе
        about_action = QAction("ℹ️ О программе", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
        
        # Системная информация
        system_info_action = QAction("💻 Системная информация", self)
        system_info_action.triggered.connect(self.show_system_info)
        help_menu.addAction(system_info_action)
        
        # Меню "Поиск" (после Помощь)
        from PySide6.QtWidgets import QWidgetAction
        self.global_search_menu = menubar.addMenu("🔍 Поиск")
        
        # Создаем виджет для выпадающего меню
        search_widget = QWidget()
        search_widget.setObjectName("globalSearchWidget")
        search_widget.setFixedWidth(300)
        
        search_layout = QHBoxLayout(search_widget)
        search_layout.setContentsMargins(8, 8, 8, 8)
        search_layout.setSpacing(6)
        
        # Поле ввода
        self.global_search_input = QLineEdit()
        self.global_search_input.setObjectName("globalSearchInput")
        self.global_search_input.setPlaceholderText("Введите название ИВП или ключевое слово...")
        self.global_search_input.setClearButtonEnabled(True)
        self.global_search_input.setMinimumWidth(200)
        
        # Кнопка поиска с лупой
        search_button = QPushButton("🔎")
        search_button.setObjectName("globalSearchButton")
        search_button.setCursor(Qt.PointingHandCursor)
        search_button.setToolTip("Найти (Enter)")
        search_button.setFixedSize(32, 32)
        
        search_layout.addWidget(self.global_search_input)
        search_layout.addWidget(search_button)
        
        # Создаем действие с виджетом
        search_action = QWidgetAction(self)
        search_action.setDefaultWidget(search_widget)
        self.global_search_menu.addAction(search_action)
        
        self.global_search_menu.addSeparator()
        
        # Горячие клавиши для поиска
        # На macOS Ctrl = Command, Meta = Control
        # Пользователь хочет Command+F/P/A
        
        # Cmd+F - Фокус на поиск
        focus_search_action = QAction("Найти (Focus)", self)
        focus_search_action.setShortcut("Ctrl+F")
        focus_search_action.triggered.connect(self.focus_global_search)
        self.global_search_menu.addAction(focus_search_action)
        
        # Cmd+P - PDF Search
        pdf_search_action = QAction("Поиск в PDF", self)
        pdf_search_action.setShortcut("Ctrl+P")
        pdf_search_action.triggered.connect(lambda: self.open_pdf_search_dialog(0))
        self.global_search_menu.addAction(pdf_search_action)
        
        # Cmd+A - AI Search
        ai_search_action = QAction("AI Поиск", self)
        ai_search_action.setShortcut("Ctrl+A")
        ai_search_action.triggered.connect(lambda: self.open_pdf_search_dialog(1))
        self.global_search_menu.addAction(ai_search_action)
        
        # Подключаем сигналы
        search_button.clicked.connect(self.on_global_search_triggered)
        self.global_search_input.returnPressed.connect(self.on_global_search_triggered)
        
        # Глобальный поиск скрыт в простом режиме, виден в расширенном и экспертном
        is_advanced_or_expert = self.current_view_mode in ["advanced", "expert"]
        self.global_search_menu.menuAction().setVisible(is_advanced_or_expert)
        # Поле ввода активируется вместе с меню
        if is_advanced_or_expert:
            self.global_search_menu.setToolTip("Глобальный поиск по базе данных и файлам")
            self.global_search_input.setEnabled(True)
        else:
            self.global_search_input.setEnabled(False)
        
        # Меню "Поиск PDF" (после глобального поиска)
        self.pdf_search_menu = menubar.addMenu("📄 Поиск PDF")
        
        # Локальный поиск - доступен всегда
        self.local_pdf_action = QAction("📁 Локальный поиск PDF", self)
        self.local_pdf_action.setToolTip("Поиск PDF файлов на компьютере в папках pdf_*, pdfBZ и т.д.")
        self.local_pdf_action.triggered.connect(lambda: self.open_pdf_search_dialog(0))
        self.pdf_search_menu.addAction(self.local_pdf_action)
        
        # AI поиск - только для экспертов после разблокировки
        self.ai_pdf_action = QAction("🤖 AI поиск компонента", self)
        self.ai_pdf_action.setToolTip("Поиск информации о компоненте через Anthropic Claude или OpenAI GPT (только экспертный режим после разблокировки)")
        self.ai_pdf_action.triggered.connect(lambda: self.open_pdf_search_dialog(1))
        # Заблокирован до разблокировки приложения
        self.ai_pdf_action.setEnabled(self.current_view_mode == "expert" and self.unlocked)
        self.pdf_search_menu.addAction(self.ai_pdf_action)
        
        self.pdf_search_menu.addSeparator()
        
        # Настройки поиска PDF - только для экспертов после разблокировки
        self.pdf_settings_action = QAction("⚙️ Настройки API ключей", self)
        self.pdf_settings_action.setToolTip("Настройка API ключей для AI поиска (только экспертный режим после разблокировки)")
        self.pdf_settings_action.triggered.connect(self.open_pdf_search_settings)
        # Заблокирован до разблокировки приложения
        self.pdf_settings_action.setEnabled(self.current_view_mode == "expert" and self.unlocked)
        self.pdf_search_menu.addAction(self.pdf_settings_action)
        
        # Меню PDF доступно всегда (локальный поиск для всех, AI - только для экспертов после разблокировки)
        self.pdf_search_menu.setEnabled(True)
        self.pdf_search_menu.setToolTip("Локальный поиск PDF доступен всегда, AI поиск - в экспертном режиме после разблокировки")

    def _create_ui(self):
        """Создает элементы интерфейса"""
        # Создаем центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Создаем главный layout с прокруткой
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(6, 6, 6, 6)
        main_layout.setSpacing(6)

        # Область прокрутки
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)

        # Контейнер для содержимого
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(8)

        # Добавляем секции (используем функции из sections)
        self.main_section = sections.create_main_section(self)
        scroll_layout.addWidget(self.main_section)

        self.comparison_section = sections.create_comparison_section(self)
        scroll_layout.addWidget(self.comparison_section)

        self.expert_section = sections.create_expert_tools_section(self)
        scroll_layout.addWidget(self.expert_section)

        self.log_section = sections.create_log_section(self)
        scroll_layout.addWidget(self.log_section)

        scroll_layout.addStretch()
        scroll_layout.addWidget(sections.create_footer(self))

        scroll_area.setWidget(scroll_content)
        main_layout.addWidget(scroll_area)

    # ==================== Обработчики событий ====================

    def _build_args(self, output_file: str) -> list:
        """
        Формирует список аргументов для CLI
        
        Args:
            output_file: Путь к выходному файлу
            
        Returns:
            Список аргументов для передачи в CLI
        """
        args = []
        if self.input_files:
            # Формируем список файлов в формате "файл:количество"
            file_specs = []
            for file_path, count in self.input_files.items():
                if count > 1:
                    file_specs.append(f"{file_path}:{count}")
                else:
                    file_specs.append(file_path)
            args.extend(["--inputs"] + file_specs)
        
        sheet_txt = self.sheet_entry.text().strip()
        if sheet_txt:
            args.extend(["--sheets", sheet_txt])
        
        args.extend(["--xlsx", output_file])
        
        if self.combine_check.isChecked():
            args.append("--combine")
        
        td = self.txt_entry.text().strip()
        if td:
            args.extend(["--txt-dir", td])
        
        # Всегда отключаем автоматический интерактивный режим в GUI
        args.append("--no-interactive")
        
        # Исключить подборы и замены если чекбокс активирован
        if hasattr(self, 'exclude_podbor_checkbox') and self.exclude_podbor_checkbox.isChecked():
            args.append("--exclude-podbor")
        
        return args
    
    def check_and_convert_doc_files(self) -> bool:
        """
        Проверяет наличие .doc файлов и предлагает конвертацию
        
        Returns:
            True если можно продолжить, False если нужно остановить
        """
        # Ищем .doc файлы (старый формат)
        doc_files = [f for f in self.input_files.keys() if f.lower().endswith('.doc') and not f.lower().endswith('.docx')]
        
        if not doc_files:
            return True  # Нет .doc файлов, продолжаем
        
        # Логируем информацию о найденных .doc файлах
        self.log_text.clear()
        self.log_text.append(f"⚠️  Обнаружено .doc файлов: {len(doc_files)}\n")
        for doc_file in doc_files:
            self.log_text.append(f"   • {os.path.basename(doc_file)}")
        self.log_text.append("\n")
        
        # Показываем диалог конвертации
        dialog = DocConversionDialog(doc_files, self)
        result = dialog.exec()
        
        if result == QDialog.Rejected:
            return False  # Пользователь отменил
        
        conversion_method = dialog.conversion_method
        
        if conversion_method == 'word':
            # Конвертация через Word
            self.log_text.append("🔄 Запуск конвертации через Microsoft Word...\n")
            result = self._convert_doc_files_with_word(doc_files)
            if result:
                self.log_text.append("\n✅ Конвертация завершена успешно!")
                self.log_text.append("⏭️  Переход к обработке файлов...\n")
            return result
        elif conversion_method == 'manual':
            # Ручная конвертация - предупреждение и продолжение
            QMessageBox.warning(
                self,
                "Ручная конвертация",
                "Пожалуйста, сконвертируйте .doc файлы в .docx вручную\n"
                "и добавьте их снова через 'Добавить файлы'.\n\n"
                ".doc файлы будут пропущены при обработке."
            )
            # Удаляем .doc файлы из списка
            for doc_file in doc_files:
                if doc_file in self.input_files:
                    del self.input_files[doc_file]
            self.update_listbox()
            
            # Проверяем что остались файлы
            if not self.input_files:
                QMessageBox.critical(
                    self,
                    "Нет файлов",
                    "После удаления .doc файлов не осталось файлов для обработки"
                )
                return False
            
            return True
        
        return False
    
    def _convert_doc_files_with_word(self, doc_files: list) -> bool:
        """
        Конвертирует .doc файлы в .docx используя Microsoft Word (Windows)
        или LibreOffice (macOS/Linux)
        
        Args:
            doc_files: Список путей к .doc файлам
            
        Returns:
            True если конвертация успешна
        """
        # На macOS/Linux используем LibreOffice
        if platform.system() != 'Windows':
            return self._convert_doc_with_libreoffice(doc_files)
        
        # На Windows используем MS Word
        # Импортируем win32com только на Windows
        try:
            import win32com.client
        except ImportError:
            QMessageBox.critical(
                self,
                "Ошибка",
                "Не установлен pywin32!\n\n"
                "Установите: pip install pywin32"
            )
            return False
        
        # Создаем прогресс-диалог
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("Конвертация .doc файлов")
        progress_dialog.setMinimumSize(600, 400)
        progress_dialog.setModal(True)
        
        layout = QVBoxLayout(progress_dialog)
        
        status_label = QLabel("Подготовка...")
        status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(status_label)
        
        progress_text = QTextEdit()
        progress_text.setReadOnly(True)
        layout.addWidget(progress_text)
        
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(progress_dialog.accept)
        close_btn.setEnabled(False)
        layout.addWidget(close_btn)
        
        progress_dialog.show()
        QApplication.processEvents()
        
        # Таймер для автозакрытия
        auto_close_timer = None
        countdown_value = [3]  # Используем список для изменяемости в замыкании
        
        converted_files = []
        success = True
        
        try:
            status_label.setText("Запуск Microsoft Word...")
            progress_text.append("Открытие Microsoft Word...\n")
            QApplication.processEvents()
            
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            
            for i, doc_file in enumerate(doc_files, 1):
                status_label.setText(f"Конвертация {i}/{len(doc_files)}: {os.path.basename(doc_file)}")
                progress_text.append(f"\n[{i}/{len(doc_files)}] {os.path.basename(doc_file)}")
                QApplication.processEvents()
                
                doc_path = os.path.abspath(doc_file)
                docx_path = doc_path.replace('.doc', '.docx')
                
                try:
                    doc = word.Documents.Open(doc_path)
                    doc.SaveAs(docx_path, FileFormat=16)  # 16 = wdFormatXMLDocument
                    doc.Close()
                    
                    progress_text.append(f"  ✓ Создан: {os.path.basename(docx_path)}")
                    converted_files.append((doc_file, docx_path))
                    
                except Exception as e:
                    progress_text.append(f"  ✗ Ошибка: {str(e)}")
                    success = False
                
                QApplication.processEvents()
            
            word.Quit()
            status_label.setText("Готово!")
            progress_text.append("\n✅ Конвертация завершена.")
            progress_text.append("\n⏭️  Переход к обработке файлов...")
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка конвертации",
                f"Не удалось запустить Word:\n{str(e)}"
            )
            success = False
        
        # Обновляем список файлов
        if success and converted_files:
            # Сохраняем список для последующего удаления промежуточных файлов
            self.converted_docx_files = [new_file for old_file, new_file in converted_files]
            
            for old_file, new_file in converted_files:
                if old_file in self.input_files:
                    count = self.input_files[old_file]
                    del self.input_files[old_file]
                    self.input_files[new_file] = count
            
            self.update_listbox()
            self.update_output_filename()
            progress_text.append("\n✓ Список файлов обновлен")
        
        # Создаем диалог обработки заранее (но не показываем)
        processing_dialog = QProgressDialog(
            "Подготовка к обработке файлов...",
            None,
            0, 0,
            self
        )
        processing_dialog.setWindowTitle("Обработка BOM файлов")
        processing_dialog.setWindowModality(Qt.WindowModal)
        processing_dialog.setMinimumDuration(0)
        processing_dialog.setCancelButton(None)
        processing_dialog.setAutoClose(False)
        processing_dialog.setAutoReset(False)
        
        # Функция обратного отсчета
        def update_countdown():
            if countdown_value[0] > 1:
                close_btn.setText(f"Закрыть (автозакрытие через {countdown_value[0]} сек)")
                status_label.setText(f"Готово! Автопереход к обработке через {countdown_value[0]} сек...")
                countdown_value[0] -= 1
            elif countdown_value[0] == 1:
                # За секунду до закрытия показываем диалог обработки
                close_btn.setText(f"Закрыть (автозакрытие через {countdown_value[0]} сек)")
                status_label.setText("Подготовка к обработке...")
                progress_text.append("\n⏭️  Запуск обработки файлов...")
                QApplication.processEvents()
                
                # Показываем диалог обработки ЕЩЕ ДО закрытия этого окна
                processing_dialog.show()
                processing_dialog.setLabelText("Анализ файлов...")
                QApplication.processEvents()
                
                countdown_value[0] -= 1
            else:
                auto_close_timer.stop()
                progress_dialog.accept()
        
        # Запускаем таймер автозакрытия
        from PySide6.QtCore import QTimer
        auto_close_timer = QTimer()
        auto_close_timer.timeout.connect(update_countdown)
        
        # Сохраняем ссылку на диалог обработки
        self.processing_dialog_ref = processing_dialog
        
        close_btn.setEnabled(True)
        close_btn.setText(f"Закрыть (автозакрытие через {countdown_value[0]} сек)")
        status_label.setText(f"Готово! Автопереход к обработке через {countdown_value[0]} сек...")
        auto_close_timer.start(1000)  # Каждую секунду
        
        progress_dialog.exec()
        
        # Останавливаем таймер если пользователь закрыл вручную
        if auto_close_timer.isActive():
            auto_close_timer.stop()
        
        return success
    
    def _convert_doc_with_libreoffice(self, doc_files: list) -> bool:
        """
        Конвертирует .doc файлы в .docx используя LibreOffice (macOS/Linux)
        
        Args:
            doc_files: Список путей к .doc файлам
            
        Returns:
            True если конвертация успешна
        """
        # Проверяем наличие LibreOffice
        libreoffice_paths = [
            '/Applications/LibreOffice.app/Contents/MacOS/soffice',  # macOS
            '/usr/bin/libreoffice',  # Linux
            '/usr/bin/soffice',      # Linux альтернатива
        ]
        
        soffice_path = None
        for path in libreoffice_paths:
            if os.path.exists(path):
                soffice_path = path
                break
        
        if not soffice_path:
            # LibreOffice не найден
            reply = QMessageBox.question(
                self,
                "LibreOffice не найден",
                "LibreOffice не установлен на этом компьютере.\n\n"
                "LibreOffice - это бесплатный офисный пакет,\n"
                "который может конвертировать .doc в .docx.\n\n"
                "Хотите скачать LibreOffice?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Открываем страницу загрузки
                import webbrowser
                webbrowser.open('https://www.libreoffice.org/download/download/')
            
            return False
        
        # Создаем прогресс-диалог
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("Конвертация .doc файлов")
        progress_dialog.setMinimumSize(600, 400)
        progress_dialog.setModal(True)
        
        layout = QVBoxLayout(progress_dialog)
        
        status_label = QLabel("Подготовка...")
        status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(status_label)
        
        log_text = QTextEdit()
        log_text.setReadOnly(True)
        layout.addWidget(log_text)
        
        close_btn = QPushButton("Закрыть")
        close_btn.setEnabled(False)
        close_btn.clicked.connect(progress_dialog.accept)
        layout.addWidget(close_btn)
        
        progress_dialog.show()
        QApplication.processEvents()
        
        # Конвертация
        success = True
        converted_files = []
        
        for i, doc_file in enumerate(doc_files, 1):
            status_label.setText(f"Конвертация {i} из {len(doc_files)}...")
            log_text.append(f"📄 {os.path.basename(doc_file)}")
            QApplication.processEvents()
            
            try:
                # Определяем выходной файл
                docx_file = doc_file[:-4] + '.docx'  # .doc -> .docx
                
                # Конвертируем через LibreOffice в headless режиме
                import subprocess
                output_dir = os.path.dirname(doc_file)
                
                # Команда: soffice --headless --convert-to docx --outdir <dir> <file>
                cmd = [
                    soffice_path,
                    '--headless',
                    '--convert-to', 'docx',
                    '--outdir', output_dir,
                    doc_file
                ]
                
                log_text.append(f"   Запуск конвертации...")
                QApplication.processEvents()
                
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=60  # 60 секунд таймаут
                )
                
                if result.returncode == 0 and os.path.exists(docx_file):
                    log_text.append(f"   ✅ Успешно: {os.path.basename(docx_file)}")
                    converted_files.append((doc_file, docx_file))
                    
                    # Добавляем .docx в список файлов
                    if doc_file in self.input_files:
                        count = self.input_files[doc_file]
                        del self.input_files[doc_file]
                        self.input_files[docx_file] = count
                else:
                    log_text.append(f"   ❌ Ошибка конвертации")
                    if result.stderr:
                        log_text.append(f"   {result.stderr[:200]}")
                    success = False
                    
            except subprocess.TimeoutExpired:
                log_text.append(f"   ❌ Таймаут (файл слишком большой)")
                success = False
            except Exception as e:
                log_text.append(f"   ❌ Ошибка: {str(e)}")
                success = False
            
            QApplication.processEvents()
        
        # Обновляем список файлов
        self.update_listbox()
        
        # Финальное сообщение
        if success:
            status_label.setText("✅ Конвертация завершена успешно!")
            log_text.append("\n✅ Все файлы сконвертированы")
            log_text.append("⏭️  Можно продолжить обработку")
            
            # Таймер автозакрытия (как в Windows версии)
            countdown_value = [3]
            
            # Создаем диалог обработки заранее
            processing_dialog = QProgressDialog(
                "Подготовка к обработке файлов...",
                None,
                0, 0,
                self
            )
            processing_dialog.setWindowTitle("Обработка BOM файлов")
            processing_dialog.setWindowModality(Qt.WindowModal)
            processing_dialog.setMinimumDuration(0)
            processing_dialog.setCancelButton(None)
            processing_dialog.setAutoClose(False)
            processing_dialog.setAutoReset(False)
            
            # Сохраняем ссылку
            self.processing_dialog_ref = processing_dialog
            
            def update_countdown():
                if countdown_value[0] > 1:
                    close_btn.setText(f"Закрыть (автозакрытие через {countdown_value[0]} сек)")
                    status_label.setText(f"Готово! Автопереход к обработке через {countdown_value[0]} сек...")
                    countdown_value[0] -= 1
                elif countdown_value[0] == 1:
                    close_btn.setText(f"Закрыть (автозакрытие через {countdown_value[0]} сек)")
                    status_label.setText("Подготовка к обработке...")
                    log_text.append("\n⏭️  Запуск обработки файлов...")
                    QApplication.processEvents()
                    
                    processing_dialog.show()
                    processing_dialog.setLabelText("Анализ файлов...")
                    QApplication.processEvents()
                    
                    countdown_value[0] -= 1
                else:
                    auto_close_timer.stop()
                    progress_dialog.accept()
            
            from PySide6.QtCore import QTimer
            auto_close_timer = QTimer()
            auto_close_timer.timeout.connect(update_countdown)
            
            close_btn.setText(f"Закрыть (автозакрытие через {countdown_value[0]} сек)")
            status_label.setText(f"Готово! Автопереход к обработке через {countdown_value[0]} сек...")
            auto_close_timer.start(1000)
            
            close_btn.setEnabled(True)
            progress_dialog.exec()
            
            if auto_close_timer.isActive():
                auto_close_timer.stop()
                
            return True
            
        else:
            status_label.setText("⚠️ Конвертация завершена с ошибками")
            log_text.append("\n⚠️ Некоторые файлы не удалось сконвертировать")
            close_btn.setEnabled(True)
            progress_dialog.exec()
            return False
    
    def cleanup_converted_files(self):
        """
        Удаляет промежуточные .docx файлы, созданные при конвертации из .doc
        Вызывается после успешного завершения обработки BOM
        """
        if not hasattr(self, 'converted_docx_files') or not self.converted_docx_files:
            return
        
        deleted_count = 0
        for docx_file in self.converted_docx_files:
            try:
                if os.path.exists(docx_file):
                    os.remove(docx_file)
                    deleted_count += 1
                    if self.log_text:
                        self.log_text.append(f"🗑️  Удалён промежуточный файл: {os.path.basename(docx_file)}")
            except Exception as e:
                if self.log_text:
                    self.log_text.append(f"⚠️  Не удалось удалить {os.path.basename(docx_file)}: {e}")
        
        # Очищаем список
        self.converted_docx_files = []
        
        if deleted_count > 0 and self.log_text:
            self.log_text.append(f"✅ Удалено промежуточных файлов: {deleted_count}")
    
    def open_interactive_cli(self):
        """Открывает интерактивную командную строку"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout
        from ..cli_interactive import InteractiveCLI
        
        # Создаем диалог
        dialog = QDialog(self)
        dialog.setWindowTitle("💻 Интерактивная командная строка")
        dialog.resize(900, 600)
        
        # Создаем layout
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Добавляем CLI виджет
        cli_widget = InteractiveCLI(self, dialog)
        layout.addWidget(cli_widget)
        
        # Показываем диалог
        dialog.exec()
        
        # Логируем
        if self.log_text:
            self.log_text.append("💻 Интерактивная командная строка закрыта")

    
    def on_run(self):
        """Запуск обработки"""
        # Проверка наличия файлов
        has_bom = bool(self.input_files)
        has_tru_rkm = hasattr(self, 'tru_rkm_files') and bool(self.tru_rkm_files)
        
        if not has_bom and not has_tru_rkm:
            QMessageBox.critical(
                self,
                "Ошибка",
                "Добавьте хотя бы один входной файл (BOM или ТРУ/РКМ)"
            )
            return

        # НОВЫЙ РЕЖИМ: Объединение BOM + ТРУ
        # Когда есть и BOM файл(ы) и ТРУ файл(ы) — объединяем данные
        if has_bom and has_tru_rkm:
            self.start_bom_tru_merge()
            return

        # Обработка только ТРУ/РКМ файлов (без BOM)
        if has_tru_rkm:
            self.log_text.append(f"\n{'='*60}\n")
            self.log_text.append(f"🚀 ЗАПУСК ОБРАБОТКИ ТРУ/РКМ ФАЙЛОВ\n")
            self.log_text.append(f"{'='*60}\n")
            self.log_text.append(f"📋 Файлов для обработки: {len(self.tru_rkm_files)}\n")
            
            # Диалог прогресса для ТРУ/РКМ
            tru_progress = QProgressDialog("Обработка ТРУ/РКМ файлов...", "Отмена", 0, len(self.tru_rkm_files) + 1, self)
            tru_progress.setWindowTitle("Обработка ТРУ/РКМ")
            tru_progress.setWindowModality(Qt.WindowModal)
            tru_progress.setMinimumDuration(0)
            tru_progress.setValue(0)
            tru_progress.setAutoClose(False) # Не закрывать авто, закроем сами
            tru_progress.show()
            
            # Создаем и запускаем воркер
            self.tru_worker = TruRkmWorker(self.tru_rkm_files)
            
            def on_progress(current, total, filename, success):
                if tru_progress.wasCanceled():
                    # TODO: Implement cancellation support in processor
                    pass
                tru_progress.setValue(current)
                tru_progress.setLabelText(f"Обработка: {filename}")
                status = "✅" if success else "❌"
                self.log_text.append(f"   {status} {filename}")
                
            def on_finished(results):
                tru_progress.close()
                
                # Статистика
                success_count = sum(1 for r in results.values() if r['success'])
                
                # Сохраняем путь к последнему успешному файлу для экспорта
                for res in results.values():
                    if res['success'] and res.get('output_path'):
                        self.last_generated_output = res['output_path']
                        # Берем первый успешный файл (обычно он один для ТРУ)
                        break
                        
                self.log_text.append(f"\n🏁 Итог ТРУ/РКМ: Успешно {success_count} из {len(self.tru_rkm_files)}")
                
                # Показываем сообщения об ошибках если были
                errors = [f"{os.path.basename(p)}: {r['message']}" for p, r in results.items() if not r['success']]
                if errors:
                    self.log_text.append("\n❌ Ошибки:")
                    for err in errors:
                        self.log_text.append(f"   • {err}")
                
                # Показываем диалог завершения
                QMessageBox.information(
                    self,
                    "Обработка завершена",
                    f"Обработка ТРУ/РКМ файлов выполнена.\n\nУспешно: {success_count}\nОшибок: {len(errors)}"
                )

            self.tru_worker.progress.connect(on_progress)
            self.tru_worker.finished.connect(on_finished)
            self.tru_worker.start()
            
            return

        # Если есть только BOM файлы, запускаем сразу
        self.start_bom_processing()
    
    def start_bom_tru_merge(self):
        """
        Объединение данных BOM и ТРУ файлов.
        Переносит из ТРУ: Артикул → КОД ERP(МР), Стоимость, корректирует Количество.
        """
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from ..tru_merger import merge_tru_into_bom, apply_merge_styles
        from ..tru_merger import build_ostatki_and_zapas_reports
        from ..tru_rkm_processor import _read_tru_file
        
        self.log_text.append(f"\n{'='*60}\n")
        self.log_text.append(f"🔗 РЕЖИМ ОБЪЕДИНЕНИЯ BOM + ТРУ\n")
        self.log_text.append(f"{'='*60}\n")
        self.log_text.append(f"📋 BOM файлов: {len(self.input_files)}")
        self.log_text.append(f"📋 ТРУ файлов: {len(self.tru_rkm_files)}\n")
        
        try:
            # 1. Читаем ТРУ файлы
            self.log_text.append("📖 Чтение ТРУ файлов...")
            tru_dfs = []
            for tru_path in self.tru_rkm_files:
                tru_df = _read_tru_file(tru_path)
                if tru_df is not None and not tru_df.empty:
                    tru_dfs.append(tru_df)
                    self.log_text.append(f"   ✅ {os.path.basename(tru_path)}: {len(tru_df)} строк")
                else:
                    self.log_text.append(f"   ⚠️ {os.path.basename(tru_path)}: не удалось прочитать")
            
            if not tru_dfs:
                QMessageBox.warning(self, "Предупреждение", "Не удалось прочитать ни одного ТРУ файла")
                return
            
            # 2. Читаем BOM файл(ы)
            self.log_text.append("\n📖 Чтение BOM файлов...")
            
            for bom_path in list(self.input_files.keys()):
                self.log_text.append(f"   📄 {os.path.basename(bom_path)}")
                
                # Определяем формат файла
                ext = os.path.splitext(bom_path)[1].lower()
                
                if ext in ['.xlsx', '.xls']:
                    # Читаем Excel файл
                    try:
                        all_sheets = pd.read_excel(bom_path, sheet_name=None, engine='openpyxl')
                    except:
                        all_sheets = pd.read_excel(bom_path, sheet_name=None)
                    
                    # Ищем колонку с наименованием
                    bom_name_col = None
                    bom_qty_col = None
                    
                    # Обрабатываем каждый лист
                    total_merged = 0
                    merged_sheets = {}
                    merged_sheets = {}
                    merged_rows_per_sheet = {}
                    all_used_tru_indices = set()
                    
                    # Для дополнительных отчетов (плоские списки, без разбиения по категориям)
                    # *_ostatki.xlsx: BOM позиции без ТРУ + позиции где TRU_qty < BOM_qty (разница)
                    # *_zapas.xlsx:  несопоставленные ТРУ + позиции где TRU_qty > BOM_qty (разница)
                    ostatki_parts = []  # List[pd.DataFrame]
                    zapas_parts = []    # List[pd.DataFrame]
                    
                    for sheet_name, df in all_sheets.items():
                        # Ищем колонки
                        for col in df.columns:
                            col_lower = str(col).lower()
                            if 'наименование ивп' in col_lower or col_lower == 'наименование':
                                bom_name_col = col
                            elif col_lower in ['шт.', 'шт', 'qty', 'количество']:
                                bom_qty_col = col
                        
                        if not bom_name_col:
                            merged_sheets[sheet_name] = df
                            continue
                        
                        # Объединяем данные
                        merged_df, merged_indices, used_indices = merge_tru_into_bom(
                            bom_df=df,
                            tru_dfs=tru_dfs,
                            tru_filenames=list(self.tru_rkm_files),
                            bom_name_col=bom_name_col,
                            bom_qty_col=bom_qty_col if bom_qty_col else 'шт.'
                        )
                        
                        all_used_tru_indices.update(used_indices)
                        
                        merged_sheets[sheet_name] = merged_df
                        merged_rows_per_sheet[sheet_name] = merged_indices
                        total_merged += len(merged_indices)
                        
                        if merged_indices:
                            self.log_text.append(f"      📊 {sheet_name}: совпадений — {len(merged_indices)}")
                        
                        # === Формирование *_ostatki / *_zapas (плоские списки) ===
                        try:
                            qty_col_hint = bom_qty_col if bom_qty_col and bom_qty_col in merged_df.columns else None
                            o_df, z_df = build_ostatki_and_zapas_reports(
                                merged_df=merged_df,
                                merged_indices=merged_indices,
                                unmatched_tru=None,  # добавим единым блоком после generate_unmatched_report
                                qty_col=qty_col_hint
                            )
                            if not o_df.empty:
                                ostatki_parts.append(o_df)
                            if not z_df.empty:
                                zapas_parts.append(z_df)
                        except Exception as e:
                            # Не блокируем основной merge из-за отчетов
                            self.log_text.append(f"   ⚠️ Не удалось сформировать отчеты остатков/запаса для листа '{sheet_name}': {e}")
                    
                    # 3. Генерируем отчет о несопоставленных ТРУ (один раз для всех листов)
                    from ..tru_merger import generate_unmatched_report
                    unmatched_tru = generate_unmatched_report(
                        tru_dfs=tru_dfs,
                        used_tru_indices=all_used_tru_indices
                    )
                    
                    # Добавляем лист "Несопоставленные ТРУ"
                    if not unmatched_tru.empty:
                        # Фильтруем пустые строки на всякий случай
                        if 'Наименование ИВП' in unmatched_tru.columns:
                             unmatched_tru = unmatched_tru[unmatched_tru['Наименование ИВП'].notna() & (unmatched_tru['Наименование ИВП'] != '')]
                        
                        self.log_text.append(f"      ❗ Несопоставленных ТРУ элементов: {len(unmatched_tru)}")
                        merged_sheets['Несопоставленные ТРУ'] = unmatched_tru
                        # Несопоставленные ТРУ идут в файл *_zapas
                        try:
                            zapas_parts.append(unmatched_tru.copy())
                        except Exception:
                            pass
                    
                    # Генерируем новый Summary лист со статистикой
                    summary_rows = []
                    
                    # Порядок листов: сначала существующие категории, потом Несопоставленные
                    sheets_order = [k for k in merged_sheets.keys() if k.lower() not in ('summary', 'sources', 'несопоставленные тру')]
                    if 'Несопоставленные ТРУ' in merged_sheets:
                        sheets_order.append('Несопоставленные ТРУ')
                        
                    for sheet_name in sheets_order:
                        df_sheet = merged_sheets[sheet_name]
                        
                        # Ищем колонки
                        qty_col = None
                        cost_col = None
                        
                        for col in df_sheet.columns:
                            col_lower = str(col).lower()
                            if col_lower in ['шт.', 'шт', 'qty', 'количество', 'кол-во', 'кол.']:
                                qty_col = col
                            elif col_lower in ['стоимость', 'cost', 'сумма', 'total']:
                                cost_col = col
                        
                        # Считаем
                        positions_count = len(df_sheet)
                        total_qty = 0
                        total_cost = 0
                        
                        if qty_col:
                            for val in df_sheet[qty_col]:
                                try:
                                    if pd.notna(val):
                                        total_qty += int(float(val))
                                except (ValueError, TypeError):
                                    pass
                        else:
                            total_qty = positions_count
                            
                        if cost_col:
                            for val in df_sheet[cost_col]:
                                try:
                                    if pd.notna(val) and str(val).strip():
                                        val_str = str(val).replace(' ', '').replace(',', '.')
                                        total_cost += float(val_str)
                                except (ValueError, TypeError):
                                    pass
                                    
                        summary_rows.append({
                            '№ п/п': len(summary_rows) + 1,
                            'Категория': sheet_name,
                            'Кол-во позиций': positions_count,
                            'Общее количество': total_qty,
                            'Стоимость': int(total_cost) if total_cost > 0 else ''
                        })
                    
                    for k in list(merged_sheets.keys()):
                        if k.lower() == 'summary':
                            del merged_sheets[k]
                    
                    if summary_rows:
                        summary_df = pd.DataFrame(summary_rows)
                        merged_sheets['Summary'] = summary_df
                        
                        # Перемещаем на позицию 1 (после Summary)
                        keys = list(merged_sheets.keys())
                        if 'Несопоставленные ТРУ' in keys:
                            keys.remove('Несопоставленные ТРУ')
                            summary_idx = keys.index('Summary') if 'Summary' in keys else -1
                            keys.insert(summary_idx + 1, 'Несопоставленные ТРУ')
                            merged_sheets = {k: merged_sheets[k] for k in keys}
                    
                    # 3. Сохраняем результат
                    # Формируем имя выходного файла
                    base_name = os.path.splitext(os.path.basename(bom_path))[0]
                    output_dir = os.path.dirname(bom_path) or "."
                    output_path = os.path.join(output_dir, f"{base_name}_тру.xlsx")
                    
                    self.log_text.append(f"\n💾 Сохранение: {os.path.basename(output_path)}")
                    
                    # Записываем в Excel
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        for sheet_name, df in merged_sheets.items():
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # === Дополнительные файлы: *_ostatki.xlsx и *_zapas.xlsx ===
                    # ВАЖНО: без разбиения по категориям — один лист, элементы подряд.
                    def _concat_parts(parts: list) -> pd.DataFrame:
                        frames = [p for p in parts if isinstance(p, pd.DataFrame) and not p.empty]
                        if not frames:
                            return pd.DataFrame()
                        return pd.concat(frames, ignore_index=True)
                    
                    ostatki_df = _concat_parts(ostatki_parts)
                    zapas_df = _concat_parts(zapas_parts)
                    
                    ostatki_path = os.path.join(output_dir, f"{base_name}_ostatki.xlsx")
                    zapas_path = os.path.join(output_dir, f"{base_name}_zapas.xlsx")
                    
                    # Пишем только если есть данные (чтобы не плодить пустые файлы)
                    if not ostatki_df.empty:
                        with pd.ExcelWriter(ostatki_path, engine='openpyxl') as writer:
                            ostatki_df.to_excel(writer, sheet_name='Остатки', index=False)
                        # Лёгкие стили (границы/заголовки)
                        try:
                            wb_o = load_workbook(ostatki_path)
                            ws_o = wb_o['Остатки']
                            apply_merge_styles(worksheet=ws_o, merged_rows=set(), name_col_idx=2, qty_col_idx=4, header_row=1)
                            wb_o.save(ostatki_path)
                        except Exception:
                            pass
                        self.log_text.append(f"📄 Создан файл остатков: {os.path.basename(ostatki_path)}")
                        
                        # PDF версия для печати
                        try:
                            from ..pdf_exporter import export_bom_to_pdf
                            ostatki_pdf = os.path.splitext(ostatki_path)[0] + ".pdf"
                            export_bom_to_pdf(ostatki_path, ostatki_pdf, with_summary=False)
                            self.log_text.append(f"🖨️ PDF (остатки): {os.path.basename(ostatki_pdf)}")
                        except Exception as e:
                            self.log_text.append(f"⚠️ Не удалось создать PDF для остатков: {e}")
                    
                    if not zapas_df.empty:
                        with pd.ExcelWriter(zapas_path, engine='openpyxl') as writer:
                            zapas_df.to_excel(writer, sheet_name='Запас', index=False)
                        try:
                            wb_z = load_workbook(zapas_path)
                            ws_z = wb_z['Запас']
                            apply_merge_styles(worksheet=ws_z, merged_rows=set(), name_col_idx=2, qty_col_idx=4, header_row=1)
                            wb_z.save(zapas_path)
                        except Exception:
                            pass
                        self.log_text.append(f"📄 Создан файл запаса: {os.path.basename(zapas_path)}")
                        
                        # PDF версия для печати
                        try:
                            from ..pdf_exporter import export_bom_to_pdf
                            zapas_pdf = os.path.splitext(zapas_path)[0] + ".pdf"
                            export_bom_to_pdf(zapas_path, zapas_pdf, with_summary=False)
                            self.log_text.append(f"🖨️ PDF (запас): {os.path.basename(zapas_pdf)}")
                        except Exception as e:
                            self.log_text.append(f"⚠️ Не удалось создать PDF для запаса: {e}")
                    
                    # 4. Применяем стили к изменённым строкам
                    wb = load_workbook(output_path)
                    
                    # Проходим по ВСЕМ листам в merged_sheets, а не только по тем где были объединения
                    # Для этого используем merged_sheets.keys() и берем индексы из merged_rows_per_sheet (или пустой set)
                    for sheet_name in merged_sheets.keys():
                        if sheet_name == 'Summary' or sheet_name == 'Несопоставленные ТРУ':
                            continue # Стилизуются отдельно
                        
                        if sheet_name not in wb.sheetnames:
                            continue
                            
                        merged_indices = merged_rows_per_sheet.get(sheet_name, set())
                        
                        ws = wb[sheet_name]
                        
                        # Находим индексы колонок
                        name_col_idx = None
                        qty_col_idx = None
                        
                        for col_idx, cell in enumerate(ws[1], start=1):
                            cell_val = str(cell.value).lower() if cell.value else ''
                            if 'наименование ивп' in cell_val or cell_val == 'наименование':
                                name_col_idx = col_idx
                            elif cell_val in ['шт.', 'шт', 'qty', 'количество']:
                                qty_col_idx = col_idx
                        
                        # Применяем стили
                        apply_merge_styles(
                            worksheet=ws,
                            merged_rows=merged_indices,
                            name_col_idx=name_col_idx or 2,
                            qty_col_idx=qty_col_idx or 4,
                            header_row=1
                        )
                        
                        # Добавляем строку ИТОГО со стоимостью
                        # Ищем колонку "Стоимость"
                        cost_col_idx = None
                        for col_idx, cell in enumerate(ws[1], start=1):
                            cell_val = str(cell.value).lower() if cell.value else ''
                            if 'стоимость' in cell_val:
                                cost_col_idx = col_idx
                                break
                        
                        if cost_col_idx:
                            # Считаем сумму стоимости
                            total_cost = 0
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=cost_col_idx, max_col=cost_col_idx):
                                for cell in row:
                                    try:
                                        if cell.value and str(cell.value).strip():
                                            val = str(cell.value).replace(' ', '').replace(',', '.')
                                            total_cost += float(val)
                                    except (ValueError, TypeError):
                                        pass
                            
                            # Добавляем строку ИТОГО
                            if total_cost > 0:
                                total_row = ws.max_row + 2  # Пустая строка + ИТОГО
                                ws.cell(row=total_row, column=cost_col_idx - 1, value="ИТОГО:").font = Font(bold=True)
                                ws.cell(row=total_row, column=cost_col_idx - 1).alignment = Alignment(horizontal='right', vertical='center')
                                ws.cell(row=total_row, column=cost_col_idx, value=int(total_cost)).font = Font(bold=True)
                                ws.cell(row=total_row, column=cost_col_idx).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Стилизуем лист несопоставленных ТРУ
                    if 'Несопоставленные ТРУ' in wb.sheetnames:
                        ws_unmatched = wb['Несопоставленные ТРУ']
                        apply_merge_styles(
                            worksheet=ws_unmatched,
                            merged_rows=set(),  # Пустой набор — просто применяем заголовки и границы
                            name_col_idx=2,
                            qty_col_idx=4,  # шт. на 4 позиции в формате BOM
                            header_row=1
                        )
                        
                        # Добавляем ИТОГО для несопоставленных ТРУ
                        cost_col_idx_unm = None
                        for col_idx, cell in enumerate(ws_unmatched[1], start=1):
                            cell_val = str(cell.value).lower() if cell.value else ''
                            if 'стоимость' in cell_val:
                                cost_col_idx_unm = col_idx
                                break
                        
                        if cost_col_idx_unm:
                            total_cost_unm = 0
                            for row in ws_unmatched.iter_rows(min_row=2, max_row=ws_unmatched.max_row, min_col=cost_col_idx_unm, max_col=cost_col_idx_unm):
                                for cell in row:
                                    try:
                                        if cell.value and str(cell.value).strip():
                                            val = str(cell.value).replace(' ', '').replace(',', '.')
                                            total_cost_unm += float(val)
                                    except (ValueError, TypeError):
                                        pass
                            
                            if total_cost_unm > 0:
                                total_row_unm = ws_unmatched.max_row + 2
                                ws_unmatched.cell(row=total_row_unm, column=cost_col_idx_unm - 1, value="ИТОГО:").font = Font(bold=True)
                                ws_unmatched.cell(row=total_row_unm, column=cost_col_idx_unm - 1).alignment = Alignment(horizontal='right', vertical='center')
                                ws_unmatched.cell(row=total_row_unm, column=cost_col_idx_unm, value=int(total_cost_unm)).font = Font(bold=True)
                                ws_unmatched.cell(row=total_row_unm, column=cost_col_idx_unm).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Стилизуем лист Summary если он есть
                    if 'Summary' in wb.sheetnames:
                        ws_summary = wb['Summary']
                        
                        # 1. Заголовки жирным и центрирование
                        for cell in ws_summary[1]:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 2. Ищем колонки для суммирования
                        pos_col_idx = None
                        qty_col_idx = None
                        cost_col_idx = None
                        
                        for col_idx, cell in enumerate(ws_summary[1], start=1):
                            val = str(cell.value).lower() if cell.value else ''
                            if 'кол-во позиций' in val:
                                pos_col_idx = col_idx
                            elif 'общее количество' in val:
                                qty_col_idx = col_idx
                            elif 'стоимость' in val:
                                cost_col_idx = col_idx
                            
                            # Центрируем данные в колонках (кроме Категории)
                            if 'категория' not in val:
                                for row in range(2, ws_summary.max_row + 1):
                                    ws_summary.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                for row in range(2, ws_summary.max_row + 1):
                                    ws_summary.cell(row=row, column=col_idx).alignment = Alignment(horizontal='left', vertical='center')

                        # 3. Считаем суммы
                        total_pos = 0
                        total_qty = 0
                        total_cost = 0
                        
                        for row in range(2, ws_summary.max_row + 1):
                            if pos_col_idx:
                                val = ws_summary.cell(row=row, column=pos_col_idx).value
                                if val and isinstance(val, (int, float, str)) and str(val).isdigit():
                                    total_pos += int(val)
                            
                            if qty_col_idx:
                                val = ws_summary.cell(row=row, column=qty_col_idx).value
                                if val and isinstance(val, (int, float, str)) and str(val).isdigit():
                                    total_qty += int(val)
                                    
                            if cost_col_idx:
                                val = ws_summary.cell(row=row, column=cost_col_idx).value
                                if val:
                                    try:
                                        total_cost += float(str(val).replace(' ', '').replace(',', '.'))
                                    except: pass
                        
                        # 4. Добавляем строку ИТОГО
                        last_row = ws_summary.max_row + 2
                        
                        ws_summary.cell(row=last_row, column=2, value="ИТОГО:").font = Font(bold=True)
                        ws_summary.cell(row=last_row, column=2).alignment = Alignment(horizontal='right', vertical='center')
                        
                        if pos_col_idx:
                            ws_summary.cell(row=last_row, column=pos_col_idx, value=total_pos).font = Font(bold=True)
                            ws_summary.cell(row=last_row, column=pos_col_idx).alignment = Alignment(horizontal='center', vertical='center')
                            
                        if qty_col_idx:
                            ws_summary.cell(row=last_row, column=qty_col_idx, value=total_qty).font = Font(bold=True)
                            ws_summary.cell(row=last_row, column=qty_col_idx).alignment = Alignment(horizontal='center', vertical='center')
                            
                        if cost_col_idx and total_cost > 0:
                            ws_summary.cell(row=last_row, column=cost_col_idx, value=int(total_cost)).font = Font(bold=True)
                            ws_summary.cell(row=last_row, column=cost_col_idx).alignment = Alignment(horizontal='center', vertical='center')

                        # 5. Автоподбор ширины
                        for column in ws_summary.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except: pass
                            ws_summary.column_dimensions[column_letter].width = min(max_length + 2, 50)
                    
                    # Переупорядочиваем листы: Summary первый, в конце Другие → Несопоставленные ТРУ → SOURCES
                    sheet_names = list(wb.sheetnames)
                    
                    # Определяем правильный порядок
                    first_sheets = []  # Summary в начале
                    middle_sheets = []  # Категории в середине
                    end_sheets = []  # Другие, Несопоставленные ТРУ, SOURCES в конце
                    
                    for name in sheet_names:
                        if name == 'Summary' or name.upper() == 'SUMMARY':
                            first_sheets.append(name)
                        elif name == 'Другие':
                            end_sheets.insert(0, name)  # Другие первым в конце
                        elif name == 'Несопоставленные ТРУ':
                            # Вставляем после Другие (или первым если Другие нет)
                            if end_sheets and end_sheets[0] == 'Другие':
                                end_sheets.insert(1, name)
                            else:
                                end_sheets.insert(0, name)
                        elif name.upper() == 'SOURCES':
                            end_sheets.append(name)  # SOURCES в самом конце
                        else:
                            middle_sheets.append(name)
                    
                    # Применяем новый порядок
                    new_order = first_sheets + middle_sheets + end_sheets
                    
                    # Перемещаем листы в правильном порядке
                    for idx, name in enumerate(new_order):
                        if name in wb.sheetnames:
                            ws = wb[name]
                            current_idx = list(wb.sheetnames).index(name)
                            if current_idx != idx:
                                wb.move_sheet(ws, offset=idx - current_idx)
                    
                    # === Перенос SOURCES из листа SOURCES в SUMMARY и удаление листа SOURCES ===
                    sources_sheet_name = None
                    for name in wb.sheetnames:
                        if name.upper() == 'SOURCES':
                            sources_sheet_name = name
                            break
                    
                    if sources_sheet_name:
                        # SOURCES уже записываются в SUMMARY через excel_writer.py (после строки ИТОГО).
                        # Здесь удаляем отдельный лист SOURCES, чтобы не дублировать.
                        del wb[sources_sheet_name]
                        self.log_text.append(f"   📋 Лист SOURCES удален (данные уже в SUMMARY)")
                    
                    wb.save(output_path)
                    
                    unmatched_count = len(unmatched_tru) if unmatched_tru is not None and not unmatched_tru.empty else 0
                    self.log_text.append(f"\n✅ Объединено элементов: {total_merged}")
                    if unmatched_count > 0:
                        self.log_text.append(f"⚠️ Несопоставленных: {unmatched_count} (см. лист 'Несопоставленные ТРУ')")
                    self.log_text.append(f"📄 Результат: {output_path}")
                    
                    # Сохраняем путь для экспорта
                    self.last_generated_output = output_path
                    
                else:
                    self.log_text.append(f"   ⚠️ Формат {ext} пока не поддерживается для объединения")
            
            # Показываем диалог завершения
            unmatched_msg = ""
            if unmatched_count > 0:
                unmatched_msg = f"\nНесопоставленных элементов: {unmatched_count}\n(см. лист 'Несопоставленные ТРУ')"
            
            # Удаляем промежуточные файлы конвертации
            self.cleanup_converted_files()
            
            QMessageBox.information(
                self,
                "Объединение завершено",
                f"Данные из ТРУ успешно объединены с BOM.\n\n"
                f"Объединено элементов: {total_merged}{unmatched_msg}\n"
                f"Результат: {output_path}"
            )
            
        except Exception as e:
            import traceback
            error_msg = f"Ошибка при объединении: {str(e)}\n\n{traceback.format_exc()}"
            self.log_text.append(f"\n❌ {error_msg}")
            QMessageBox.critical(self, "Ошибка", error_msg)
    
    def start_bom_processing(self):
        """Запуск обработки BOM файлов (вынесено в отдельный метод)"""
        # Проверяем и конвертируем .doc файлы
        conversion_result = self.check_and_convert_doc_files()
        
        if not conversion_result:
            return  # Пользователь отменил или нужна ручная конвертация
        
        args = self._build_args(self.output_entry.text())
        
        # Обновляем лог (не очищаем если там уже есть информация о конвертации)
        self.log_text.append(f"\n{'='*60}\n")
        self.log_text.append(f"🚀 ЗАПУСК ОБРАБОТКИ BOM ФАЙЛОВ\n")
        self.log_text.append(f"{'='*60}\n")
        self.log_text.append(f"📋 Входных файлов: {len(self.input_files)}")
        self.log_text.append(f"📄 Выходной файл: {os.path.basename(self.output_entry.text())}\n")
        self.log_text.append(f"⚙️  Команда: split_bom {' '.join(args)}\n")
        
        # Используем уже созданный диалог или создаем новый
        if hasattr(self, 'processing_dialog_ref') and self.processing_dialog_ref:
            self.progress_dialog = self.processing_dialog_ref
            self.progress_dialog.setLabelText("Обработка файлов в процессе...")
            self.processing_dialog_ref = None  # Очищаем ссылку
        else:
            # Создаем progress dialog если не было конвертации
            self.progress_dialog = QProgressDialog(
                "Подготовка к обработке...",
                None,
                0, 0,
                self
            )
            self.progress_dialog.setWindowTitle("Обработка BOM файлов")
            self.progress_dialog.setWindowModality(Qt.WindowModal)
            self.progress_dialog.setMinimumDuration(0)
            self.progress_dialog.setCancelButton(None)
            self.progress_dialog.setAutoClose(False)
            self.progress_dialog.setAutoReset(False)
            self.progress_dialog.show()
            self.progress_dialog.setLabelText("Обработка файлов в процессе...")
        
        QApplication.processEvents()
        
        # Создаем и запускаем worker
        self.processing_worker = ProcessingWorker(args)
        self.processing_worker.progress.connect(self.on_processing_progress)
        self.processing_worker.finished.connect(self.on_processing_finished)
        self.processing_worker.start()
    

    def on_compare_files(self):
        """Сравнение файлов"""
        file1 = self.compare_entry1.text().strip()
        file2 = self.compare_entry2.text().strip()
        output = self.compare_output_entry.text().strip()
        
        # Валидация
        if not file1 or not file2:
            QMessageBox.critical(
                self,
                "Ошибка",
                "Выберите оба файла для сравнения"
            )
            return
        
        if not output:
            QMessageBox.critical(
                self,
                "Ошибка",
                "Укажите имя файла для результатов"
            )
            return
        
        # Проверяем существование файлов
        if not os.path.exists(file1):
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Первый файл не найден:\n{file1}"
            )
            return
        
        if not os.path.exists(file2):
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Второй файл не найден:\n{file2}"
            )
            return
        
        # Очищаем лог
        self.log_text.clear()
        self.log_text.append("🔄 Сравнение файлов...")
        self.log_text.append(f"  Первый:  {os.path.basename(file1)}")
        self.log_text.append(f"  Второй:  {os.path.basename(file2)}")
        self.log_text.append(f"  Результат: {os.path.basename(output)}\n")
        
        # Создаем progress dialog
        self.progress_dialog = QProgressDialog(
            "Сравнение файлов...",
            "Отмена",
            0, 0,
            self
        )
        self.progress_dialog.setWindowTitle("Обработка")
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setMinimumDuration(0)
        self.progress_dialog.setCancelButton(None)  # Убираем кнопку отмены
        self.progress_dialog.show()
        
        # Создаем и запускаем worker
        self.comparison_worker = ComparisonWorker(file1, file2, output)
        self.comparison_worker.progress.connect(self.on_comparison_progress)
        self.comparison_worker.finished.connect(self.on_comparison_finished)
        self.comparison_worker.start()
    
    def on_comparison_progress(self, message: str):
        """Обработка прогресса сравнения"""
        self.log_text.append(message)
    
    def on_comparison_finished(self, message: str, success: bool):
        """Обработка завершения сравнения"""
        # Закрываем progress dialog
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()
        
        # Добавляем сообщение в лог
        self.log_text.append("\n" + message)
        
        # Показываем результат
        if success:
            output_file = self.compare_output_entry.text().strip()
            reply = QMessageBox.question(
                self,
                "Готово",
                f"{message}\n\nОткрыть файл?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes and os.path.exists(output_file):
                try:
                    # Открываем файл в системном приложении
                    if platform.system() == 'Windows':
                        os.startfile(output_file)
                    elif platform.system() == 'Darwin':  # macOS
                        subprocess.Popen(['open', output_file])
                    else:  # Linux
                        subprocess.Popen(['xdg-open', output_file])
                except Exception as e:
                    QMessageBox.warning(
                        self,
                        "Предупреждение",
                        f"Не удалось открыть файл:\n{str(e)}"
                    )
        else:
            QMessageBox.critical(
                self,
                "Ошибка",
                message
            )

    def on_interactive_classify(self):
        """Интерактивная классификация"""
        output_file = self.output_entry.text().strip()
        
        if not output_file:
            QMessageBox.critical(
                self,
                "Ошибка",
                "Сначала обработайте файлы, затем запустите интерактивную классификацию"
            )
            return
        
        if not os.path.exists(output_file):
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Выходной файл не найден:\n{output_file}\n\nСначала обработайте входные файлы"
            )
            return
        
        self.run_interactive_classification(output_file)
    
    def run_interactive_classification(self, output_file: str):
        """
        Запускает интерактивную классификацию для выходного файла
        
        Args:
            output_file: Путь к выходному файлу с нераспределенными элементами
        """
        try:
            import pandas as pd
            
            # Проверяем наличие листа "Не распределено"
            xls = pd.ExcelFile(output_file, engine='openpyxl')
            
            if 'Не распределено' not in xls.sheet_names:
                QMessageBox.information(
                    self,
                    "Информация",
                    "В файле нет нераспределенных элементов.\n\nВсе элементы уже классифицированы!"
                )
                return
            
            df_un = pd.read_excel(output_file, sheet_name='Не распределено', engine='openpyxl')
            
            # Фильтруем пустые строки
            df_un_valid = df_un[df_un['Наименование ИВП'].notna()]
            
            if len(df_un_valid) == 0:
                QMessageBox.information(
                    self,
                    "Информация",
                    "В листе 'Не распределено' нет элементов для классификации"
                )
                return
            
            # Подготавливаем данные для диалога
            components = []
            designation_to_name = {}
            for _, row in df_un_valid.iterrows():
                # Получаем данные, обрабатывая возможные NaN
                designation = str(row.get('Обозначение', ''))
                if designation.lower() == 'nan': designation = ""
                
                name = str(row.get('Наименование ИВП', ''))
                if not name or name.lower() == 'nan':
                    name = str(row.get('Наименование', ''))
                if name.lower() == 'nan': name = ""
                
                params = str(row.get('Корпус', ''))
                if params.lower() == 'nan': params = ""
                nominal = str(row.get('Номинал', ''))
                if nominal.lower() != 'nan' and nominal:
                    params += f" {nominal}"
                
                components.append((designation, name, params.strip()))
                # Сохраняем маппинг обозначения на имя для сохранения в БД
                # В случае дубликатов обозначений будет сохранено последнее имя,
                # что является известным ограничением текущей реализации диалога
                designation_to_name[designation] = name

            # Показываем диалог классификации
            dialog = ClassificationDialog(components, self)
            dialog.exec()
            
            # После завершения диалога сохраняем результаты и обновляем лог
            if hasattr(dialog, 'classifications') and dialog.classifications:
                count = len(dialog.classifications)
                saved_count = 0
                moved_count = 0
                
                # Загружаем все листы Excel для модификации
                try:
                    # Используем openpyxl для сохранения форматирования (насколько возможно)
                    # Но pandas проще для манипуляций с данными. 
                    # Перезапись файла через pandas может сбросить форматирование.
                    # Поэтому используем append mode или просто перезаписываем данные.
                    # Для простоты и надежности перезапишем файл с помощью pandas, 
                    # так как структура простая.
                    
                    all_sheets = pd.read_excel(output_file, sheet_name=None, engine='openpyxl')
                    
                    if 'Не распределено' in all_sheets:
                        df_un = all_sheets['Не распределено']
                        
                        # Создаем список индексов для удаления из "Не распределено"
                        indices_to_drop = []
                        
                        for comp_name, category in dialog.classifications.items():
                            # Теперь ключ = имя компонента (comp_name)
                            
                            if comp_name and category and category != 's':  # 's' - пропустить
                                # 1. Сохраняем в базу данных
                                add_component_to_database(comp_name, category)
                                saved_count += 1
                                
                                # 2. Перемещаем в соответствующий лист
                                # Находим строки с этим именем в df_un
                                mask = (df_un['Наименование ИВП'].astype(str) == comp_name)
                                if 'Наименование' in df_un.columns:
                                    mask = mask | (df_un['Наименование'].astype(str) == comp_name)
                                
                                rows_to_move = df_un[mask]
                                
                                if not rows_to_move.empty:
                                    target_sheet_name = CATEGORY_NAMES.get(category, "Другие компоненты")
                                    
                                    # Если листа нет, создаем пустой
                                    if target_sheet_name not in all_sheets:
                                        all_sheets[target_sheet_name] = pd.DataFrame(columns=df_un.columns)
                                    
                                    # Добавляем строки в целевой лист
                                    all_sheets[target_sheet_name] = pd.concat([all_sheets[target_sheet_name], rows_to_move], ignore_index=True)
                                    
                                    # Добавляем индексы для удаления
                                    indices_to_drop.extend(rows_to_move.index.tolist())
                                    moved_count += len(rows_to_move)
                        
                        # Удаляем перемещенные строки из "Не распределено"
                        if indices_to_drop:
                            df_un_cleaned = df_un.drop(index=list(set(indices_to_drop)))
                            all_sheets['Не распределено'] = df_un_cleaned
                            
                            # Читаем СУЩЕСТВУЮЩИЕ SOURCES из файла ДО перезаписи
                            # ВАЖНО: если SOURCES встречается несколько раз — берем ПОСЛЕДНЮЮ строку (обычно она полная).
                            existing_sources = []
                            try:
                                from openpyxl import load_workbook as load_wb_pre
                                wb_pre = load_wb_pre(output_file)
                                if 'SUMMARY' in wb_pre.sheetnames:
                                    ws_pre = wb_pre['SUMMARY']
                                    for row_idx in range(1, ws_pre.max_row + 1):
                                        cell_val = ws_pre.cell(row=row_idx, column=1).value
                                        if cell_val and str(cell_val).strip().upper() == 'SOURCES:':
                                            # Читаем все источники из этой строки
                                            # (перезаписываем existing_sources — так мы оставим последнюю найденную строку)
                                            row_sources = []
                                            col = 2
                                            while col <= ws_pre.max_column:
                                                val = ws_pre.cell(row=row_idx, column=col).value
                                                if val:
                                                    row_sources.append(str(val))
                                                col += 1
                                            if row_sources:
                                                existing_sources = row_sources
                                wb_pre.close()
                            except Exception as e:
                                print(f"[WARNING] Не удалось прочитать существующие SOURCES: {e}")
                            
                            # Удаляем строк(и) SOURCES из SUMMARY перед перезаписью через pandas.
                            # Иначе pandas "обрежет" источники (часть уйдет в Unnamed колонки),
                            # и после добавления существующих SOURCES получится дубль: неполная + полная.
                            if 'SUMMARY' in all_sheets:
                                try:
                                    df_summary = all_sheets['SUMMARY']
                                    if df_summary is not None and not df_summary.empty:
                                        first_col = df_summary.columns[0]
                                        mask_sources = (
                                            df_summary[first_col]
                                            .astype(str)
                                            .str.strip()
                                            .str.upper()
                                            .eq('SOURCES:')
                                        )
                                        if mask_sources.any():
                                            all_sheets['SUMMARY'] = df_summary.loc[~mask_sources].copy()
                                except Exception as e:
                                    print(f"[WARNING] Не удалось удалить SOURCES из SUMMARY перед записью: {e}")

                            # Сохраняем обновленный файл с форматированием
                            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                                for sheet_name, df in all_sheets.items():
                                    # Удаляем "Unnamed" и служебные столбцы
                                    cols_to_drop = [col for col in df.columns if str(col).startswith('Unnamed') or col == 'source_multiplier']
                                    if cols_to_drop:
                                        df = df.drop(columns=cols_to_drop)
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                                # Применяем стили (ширина столбцов, границы, выравнивание)
                                apply_excel_styles(writer)
                            
                            # Добавляем SOURCES в SUMMARY (используем существующие если есть)
                            if existing_sources:
                                from openpyxl import load_workbook
                                from openpyxl.styles import Font, Alignment
                                from openpyxl.utils import get_column_letter
                                
                                wb = load_workbook(output_file)
                                if 'SUMMARY' in wb.sheetnames:
                                    ws = wb['SUMMARY']
                                    
                                    # Находим последнюю строку с данными
                                    last_row = ws.max_row
                                    sources_row = last_row + 2
                                    
                                    # Заголовок SOURCES
                                    bold_font = Font(bold=True)
                                    ws.cell(row=sources_row, column=1, value='SOURCES:').font = bold_font
                                    ws.cell(row=sources_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                                    
                                    # Записываем существующие источники
                                    for i, cell_value in enumerate(existing_sources):
                                        col = 2 + i
                                        cell = ws.cell(row=sources_row, column=col, value=cell_value)
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        
                                        # Устанавливаем ширину колонки по содержимому
                                        col_letter = get_column_letter(col)
                                        current_width = ws.column_dimensions[col_letter].width or 0
                                        new_width = len(str(cell_value)) + 2
                                        if new_width > current_width:
                                            ws.column_dimensions[col_letter].width = new_width
                                    
                                    wb.save(output_file)
                                    print(f"[SOURCES] Восстановлено {len(existing_sources)} источников после классификации")
                            
                            self.log_text.append(f"✅ Перемещено в категории: {moved_count} строк\n")
                            self.log_text.append(f"💾 Файл обновлен: {output_file}\n")
                            
                            # Регенерация PDF если включена автогенерация
                            if self.auto_export_pdf:
                                self._regenerate_pdf_after_classification(output_file)
                            
                            # Предлагаем открыть обновленный файл
                            reply = QMessageBox.question(
                                self,
                                "Готово",
                                f"Классификация завершена.\nСохранено в базу: {saved_count}\nПеремещено в файле: {moved_count}\n\nОткрыть обновленный файл?",
                                QMessageBox.Yes | QMessageBox.No,
                                QMessageBox.Yes
                            )
                            
                            if reply == QMessageBox.Yes:
                                try:
                                    if platform.system() == 'Windows':
                                        os.startfile(output_file)
                                    elif platform.system() == 'Darwin':
                                        subprocess.Popen(['open', output_file])
                                    else:
                                        subprocess.Popen(['xdg-open', output_file])
                                except Exception as e:
                                    QMessageBox.warning(self, "Ошибка", f"Не удалось открыть файл: {e}")

                except Exception as e:
                    self.log_text.append(f"❌ Ошибка при обновлении файла: {e}\n")
                    import traceback
                    print(traceback.format_exc())
                
                self.log_text.append(f"\n✅ Классифицировано элементов: {count}\n")
                self.log_text.append(f"💾 Сохранено в базу знаний: {saved_count}\n")
            
        except Exception as e:
            import traceback
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось запустить интерактивную классификацию:\n{str(e)}\n\n{traceback.format_exc()}"
            )


    def on_open_db_folder(self):
        """Открыть папку с базой данных в проводнике с выделенным файлом"""
        try:
            db_path = get_database_path()
            if not self.reveal_in_file_manager(db_path, select=True):
                raise RuntimeError("Не удалось открыть проводник.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть папку:\n{str(e)}")
    
    def on_open_install_folder(self):
        """Открыть папку установки Modern Edition (где находится config_qt.json)"""
        try:
            # Для Modern Edition открываем папку установки, а не папку базы данных
            config_path = get_config_path()
            install_dir = os.path.dirname(config_path)
            
            # Если папка не существует, создаем её
            if not os.path.exists(install_dir):
                os.makedirs(install_dir, exist_ok=True)
            
            # Открываем папку установки
            if not self.reveal_in_file_manager(install_dir, select=False):
                raise RuntimeError("Не удалось открыть проводник.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть папку установки:\n{str(e)}")

    def on_replace_database(self):
        """Заменить текущую базу данных на другую из JSON файла"""
        try:
            # Выбор файла базы данных
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Выберите файл базы данных (component_database.json)",
                "",
                "JSON файлы (*.json);;Все файлы (*.*)"
            )
            
            if not file_path:
                return
            
            # Проверяем что файл существует и валиден
            if not os.path.exists(file_path):
                QMessageBox.critical(self, "Ошибка", f"Файл не найден:\n{file_path}")
                return
            
            # Проверяем формат файла
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                # Проверяем что это база данных компонентов
                if not isinstance(data, dict):
                    QMessageBox.critical(self, "Ошибка", "Неверный формат файла!\n\nОжидается JSON с данными компонентов.")
                    return
                
                # Определяем количество компонентов
                if "components" in data:
                    component_count = len(data["components"])
                elif "metadata" in data or "categories" in data:
                    QMessageBox.critical(self, "Ошибка", "Файл не содержит компонентов!")
                    return
                else:
                    # Старый формат - прямой словарь
                    component_count = len(data)
                
                if component_count == 0:
                    reply = QMessageBox.question(
                        self,
                        "Предупреждение",
                        "⚠️ Выбранная база данных пустая (0 компонентов)!\n\n"
                        "Это удалит все компоненты из текущей базы.\n\n"
                        "Продолжить?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No
                    )
                    if reply != QMessageBox.Yes:
                        return
                
            except json.JSONDecodeError:
                QMessageBox.critical(self, "Ошибка", "Файл поврежден или имеет неверный формат JSON!")
                return
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось прочитать файл:\n{str(e)}")
                return
            
            # Получаем информацию о текущей базе
            current_db_path = get_database_path()
            current_stats = get_database_stats()
            current_count = current_stats.get('total', 0)
            
            # Подтверждение замены
            reply = QMessageBox.question(
                self,
                "Подтверждение замены",
                f"🔄 ЗАМЕНА БАЗЫ ДАННЫХ\n\n"
                f"Текущая база данных:\n"
                f"  📊 Компонентов: {current_count}\n"
                f"  📁 Расположение: ...{current_db_path[-50:]}\n\n"
                f"Новая база данных:\n"
                f"  📊 Компонентов: {component_count}\n"
                f"  📁 Файл: {os.path.basename(file_path)}\n\n"
                f"⚠️ Текущая база будет заменена!\n"
                f"Резервная копия будет создана автоматически.\n\n"
                f"Продолжить?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Создаем резервную копию текущей базы
            try:
                backup_file = backup_database()
                self.log_text.append(f"\n💾 Резервная копия создана:")
                self.log_text.append(f"   {os.path.basename(backup_file)}\n")
            except Exception as e:
                reply = QMessageBox.question(
                    self,
                    "Ошибка резервного копирования",
                    f"Не удалось создать резервную копию:\n{str(e)}\n\n"
                    f"Продолжить без резервной копии?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return
            
            # Копируем новую базу данных
            import shutil
            shutil.copy2(file_path, current_db_path)
            
            # Проверяем что копирование прошло успешно
            new_stats = get_database_stats()
            new_count = new_stats.get('total', 0)
            
            self.log_text.append(f"\n✅ База данных успешно заменена!")
            self.log_text.append(f"   Новое количество компонентов: {new_count}")
            self.log_text.append(f"   Расположение: {current_db_path}\n")
            
            # Обновляем футер после замены
            self.update_database_info()
            
            QMessageBox.information(
                self,
                "Успех",
                f"✅ База данных успешно заменена!\n\n"
                f"Компонентов в новой базе: {new_count}\n\n"
                f"Резервная копия старой базы сохранена.\n\n"
                f"Информация в футере обновлена!"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось заменить базу данных:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def on_import_from_output(self):
        """Импорт всех компонентов из выходного файла в базу данных"""
        try:
            # Проверяем есть ли выходной файл
            output_file = self.output_entry.text()
            
            if not output_file or not os.path.exists(output_file):
                QMessageBox.critical(
                    self,
                    "Ошибка",
                    "Выходной файл не найден!\n\n"
                    "Сначала обработайте входные файлы, "
                    "проверьте результат, а затем импортируйте компоненты в базу данных."
                )
                return
            
            # Подтверждение
            reply = QMessageBox.question(
                self,
                "Импорт из выходного файла",
                f"Вы хотите добавить ВСЕ компоненты из файла:\n\n"
                f"{os.path.basename(output_file)}\n\n"
                f"в базу данных?\n\n"
                f"Это позволит автоматически классифицировать эти компоненты "
                f"в будущем при обработке других файлов.\n\n"
                f"Продолжить?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Создаем диалог прогресса
            from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextEdit, QPushButton
            progress_dialog = QDialog(self)
            progress_dialog.setWindowTitle("Импорт из выходного файла")
            progress_dialog.setMinimumSize(600, 400)
            progress_dialog.setModal(True)
            
            layout = QVBoxLayout(progress_dialog)
            
            progress_text = QTextEdit()
            progress_text.setReadOnly(True)
            layout.addWidget(progress_text)
            
            close_btn = QPushButton("Закрыть")
            close_btn.clicked.connect(progress_dialog.accept)
            close_btn.setEnabled(False)
            layout.addWidget(close_btn)
            
            progress_text.append("📥 Импорт компонентов из выходного файла...")
            progress_text.append(f"Файл: {output_file}\n")
            
            progress_dialog.show()
            QApplication.processEvents()
            
            # Импортируем компоненты
            import pandas as pd
            from ..component_database import load_component_database, save_component_database
            
            # Маппинг русских названий листов на ключи категорий
            SHEET_TO_CATEGORY = {
                'Резисторы': 'resistors',
                'Конденсаторы': 'capacitors',
                'Индуктивности': 'inductors',
                'Полупроводники': 'semiconductors',
                'Микросхемы': 'ics',
                'Разъемы': 'connectors',
                'Оптика': 'optics',
                'СВЧ модули': 'rf_modules',
                'Кабели': 'cables',
                'Модули питания': 'power_modules',
                'Отладочные платы': 'dev_boards',
                'Наши разработки': 'our_developments',
                'Другие': 'others',
            }
            
            # Загружаем текущую БД один раз
            db = load_component_database()
            initial_count = len(db)
            
            # Список добавленных компонентов для истории
            added_component_names = []
            
            # Читаем файл Excel
            xl_file = pd.ExcelFile(output_file, engine='openpyxl')
            
            added_count = 0
            skipped_count = 0
            total_sheets = 0
            
            progress_text.append("📊 Обработка листов:\n")
            QApplication.processEvents()
            
            # Обрабатываем каждый лист
            for sheet_name in xl_file.sheet_names:
                # Пропускаем служебные листы
                if sheet_name in ['SOURCES', 'SUMMARY', 'Не распределено', 'INFO']:
                    continue
                
                # Проверяем что это лист категории
                if sheet_name not in SHEET_TO_CATEGORY:
                    continue
                
                category_key = SHEET_TO_CATEGORY[sheet_name]
                total_sheets += 1
                
                # Читаем данные
                df = pd.read_excel(output_file, sheet_name=sheet_name, engine='openpyxl')
                
                if df.empty:
                    continue
                
                # Ищем колонку с наименованием
                name_col = None
                for col in ['Наименование ИВП', 'Наименование', 'наименование ивп', 'наименование']:
                    if col in df.columns:
                        name_col = col
                        break
                
                if not name_col:
                    progress_text.append(f"⚠️  {sheet_name}: не найдена колонка с наименованием")
                    continue
                
                sheet_added = 0
                
                # Собираем все компоненты в память
                for idx, row in df.iterrows():
                    name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                    
                    # Пропускаем пустые названия
                    if not name or name == 'nan':
                        skipped_count += 1
                        continue
                    
                    # Добавляем в БД только если новый или категория изменилась
                    if name not in db or db[name] != category_key:
                        db[name] = category_key
                        added_component_names.append(name)
                        added_count += 1
                        sheet_added += 1
                
                progress_text.append(f"✅ {sheet_name}: добавлено {sheet_added} компонентов")
                QApplication.processEvents()
            
            # Сохраняем БД один раз со всеми изменениями
            progress_text.append(f"\n💾 Сохранение изменений в базу данных...")
            QApplication.processEvents()
            
            if added_count > 0:
                # Есть новые компоненты - сохраняем с историей
                save_component_database(
                    db, 
                    action="import_from_file", 
                    source=os.path.abspath(output_file),  # Сохраняем полный путь для истории
                    component_names=added_component_names[:50]  # Первые 50 для истории
                )
                progress_text.append(f"✅ База данных обновлена! Добавлено {added_count} новых компонентов.")
            else:
                # Нет новых компонентов, но обновляем метаданные (last_updated)
                save_component_database(
                    db, 
                    action="update", 
                    source=None,
                    component_names=[]
                )
                progress_text.append(f"✅ Метаданные базы данных обновлены (новых компонентов не найдено).")
            
            QApplication.processEvents()
            
            progress_text.append(f"\n✅ Импорт завершен!\n")
            progress_text.append(f"📈 Статистика:")
            progress_text.append(f"   Обработано листов: {total_sheets}")
            progress_text.append(f"   Добавлено компонентов: {added_count}")
            progress_text.append(f"   Пропущено (пустые): {skipped_count}\n")
            
            # Показываем обновленную статистику базы данных
            stats = get_database_stats()
            metadata = stats.get('metadata', {})
            progress_text.append(f"📊 База данных после импорта:")
            progress_text.append(f"   Всего компонентов: {stats['total']}")
            progress_text.append(f"   Версия БД: {metadata.get('version', 'N/A')}")
            
            close_btn.setEnabled(True)
            progress_dialog.exec()
            
            # Обновляем футер после импорта
            self.update_database_info()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось импортировать компоненты:\n{str(e)}")
            import traceback
            self.update_filter()
    
    def focus_global_search(self):
        """Устанавливает фокус на глобальный поиск и выделяет текст"""
        self.global_search_input.setFocus()
        self.global_search_input.selectAll()
    
    def on_global_search_triggered(self):
        """Запускает глобальный поиск по базе данных и файлам."""
        if not self.global_search_input:
            return

        query = self.global_search_input.text().strip()
        if not query:
            self.statusBar().showMessage("⚠ Введите ключевое слово для поиска", 3000)
            self.global_search_input.setFocus()
            return

        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            results = search_methods.perform_global_search(self, query)
        finally:
            QApplication.restoreOverrideCursor()

        if results["total_matches"] == 0 and not results["notes"]:
            self.statusBar().showMessage(f"ℹ Совпадений по запросу «{query}» не найдено", 4000)
            self.global_search_input.setFocus()
            self.global_search_input.selectAll()
            return

        dialog = GlobalSearchDialog(self, results)
        dialog.exec()
        self.global_search_input.setFocus()
        self.global_search_input.selectAll()
    
    def open_pdf_search_dialog(self, tab_index: int = 0):
        """
        Открывает диалог поиска PDF
        
        Args:
            tab_index: Индекс вкладки (0 - локальный поиск, 1 - AI поиск)
        """
        from .pdf_search_dialogs import PDFSearchDialog
        
        # Передаем информацию о разблокировке и режиме
        dialog = PDFSearchDialog(self, self.cfg, 
                                 unlocked=self.unlocked, 
                                 expert_mode=(self.current_view_mode == "expert"))
        dialog.tabs.setCurrentIndex(tab_index)
        dialog.show()  # Немодальный диалог
    
    def open_pdf_search_settings(self):
        """Открывает настройки поиска PDF"""
        from .pdf_search_dialogs import PDFSearchSettingsDialog
        
        dialog = PDFSearchSettingsDialog(self, self.cfg)
        if dialog.exec() == QDialog.Accepted:
            self.cfg = dialog.get_config()
            self.save_pdf_search_config(self.cfg)
    
    def save_pdf_search_config(self, config: dict):
        """Сохраняет конфигурацию поиска PDF"""
        try:
            # Используем ту же логику определения пути, что и load_config()
            config_path = get_config_path()
            
            # Загружаем текущий конфиг, чтобы сохранить все остальные настройки
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    full_config = json.load(f)
            except (FileNotFoundError, json.JSONDecodeError):
                full_config = config.copy()
            
            # Обновляем конфиг из переданного параметра
            full_config.update(config)
            
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(full_config, f, indent=2, ensure_ascii=False)
            
            # Обновляем конфиг в памяти
            self.cfg = full_config
            self.config = full_config
            
            self.log_text.append(f"✅ Настройки поиска PDF сохранены в {config_path}\n")
        except Exception as e:
            self.log_text.append(f"⚠️ Ошибка сохранения настроек: {e}\n")

    def update_database_info(self):
        """Обновляет информацию о базе данных в футере"""
        try:
            stats = get_database_stats()
            metadata = stats.get('metadata', {})
            db_version = metadata.get('version', 'N/A')
            last_updated = metadata.get('last_updated', '')
            total_components = stats.get('total', 0)
            
            # Форматируем дату для отображения
            if last_updated and last_updated != 'N/A':
                try:
                    date_part = last_updated.split()[0]  # Берем только дату без времени
                    version_text = f"{db_version} ({date_part})"
                except:
                    version_text = db_version
            else:
                version_text = db_version
            
            self.db_info_label.setText(f"БД: {version_text} ({total_components} компонентов)")
            
            # Обновляем tooltip
            self.update_database_tooltip()
            
            # Устанавливаем курсор и обработчик клика (если еще не установлены)
            if not self.db_info_label.cursor().shape() == Qt.PointingHandCursor:
                self.db_info_label.setCursor(Qt.PointingHandCursor)
                self.db_info_label.mousePressEvent = lambda event: self.on_view_database()
        except Exception as e:
            self.db_info_label.setText("БД: Ошибка загрузки")
            print(f"Ошибка обновления информации БД: {e}")
    
    def update_database_tooltip(self):
        """Обновляет tooltip для информации о базе данных"""
        try:
            from ..component_database import get_database_history
            
            stats = get_database_stats()
            metadata = stats.get('metadata', {})
            history = get_database_history()
            
            # Формируем tooltip
            tooltip_lines = []
            tooltip_lines.append(f"📊 База данных компонентов")
            tooltip_lines.append(f"═══════════════════════════")
            tooltip_lines.append(f"Версия: {metadata.get('version', 'N/A')}")
            tooltip_lines.append(f"Всего компонентов: {stats.get('total', 0)}")
            tooltip_lines.append(f"Последнее обновление: {metadata.get('last_updated', 'N/A')}")
            
            # Добавляем структуру по категориям
            by_category = stats.get('by_category', {})
            category_names = stats.get('category_names', {})
            if by_category:
                tooltip_lines.append(f"")
                tooltip_lines.append(f"📋 Структура по категориям:")
                tooltip_lines.append(f"─────────────────────────────")
                # Сортируем по количеству (от большего к меньшему)
                sorted_categories = sorted(by_category.items(), key=lambda x: x[1], reverse=True)
                for cat_key, count in sorted_categories:
                    cat_name = category_names.get(cat_key, cat_key)
                    tooltip_lines.append(f"  {cat_name}: {count}")
            
            tooltip_lines.append(f"")
            
            # Добавляем историю последних изменений
            if history and len(history) > 0:
                tooltip_lines.append(f"📜 История последних изменений:")
                tooltip_lines.append(f"─────────────────────────────")
                
                # Показываем последние 3 записи (новые записи добавляются в начало)
                for entry in history[:3]:
                    timestamp = entry.get('timestamp', 'N/A')
                    action = entry.get('action', 'unknown')
                    source = entry.get('source', 'N/A')
                    comp_count = entry.get('components_added', 0)
                    
                    action_text = {
                        'import_from_file': '📥 Импорт из файла',
                        'import_from_excel': '📊 Импорт из Excel',
                        'manual_add': '✍️ Ручное добавление',
                        'update': '🔄 Обновление',
                        'manual_version_change': '🔢 Смена версии',
                        'database_cleared': '🗑️ Очистка БД',
                        'initial_creation': '🆕 Создание БД',
                        'conversion_from_old_format': '🔄 Конвертация'
                    }.get(action, action)
                    
                    tooltip_lines.append(f"")
                    tooltip_lines.append(f"{timestamp}")
                    tooltip_lines.append(f"  {action_text}")
                    tooltip_lines.append(f"  Версия: {entry.get('version', 'N/A')}")
                    if source != 'N/A':
                        tooltip_lines.append(f"  Источник: {source}")
                    tooltip_lines.append(f"  Добавлено: {comp_count} комп.")
                    
                    # Показываем хэш изменения
                    entry_hash = entry.get('current_hash', '')
                    if entry_hash:
                        tooltip_lines.append(f"  Хэш: {entry_hash[:12]}...")
                    
                    # Показываем несколько компонентов
                    if 'component_names' in entry and entry['component_names']:
                        names = entry['component_names'][:2]  # Первые 2
                        for name in names:
                            tooltip_lines.append(f"    • {name}")
                        if len(entry['component_names']) > 2:
                            tooltip_lines.append(f"    ... и еще {len(entry['component_names']) - 2}")
            else:
                tooltip_lines.append(f"История изменений пуста")
            
            self.db_info_label.setToolTip('\n'.join(tooltip_lines))
            
        except Exception as e:
            self.db_info_label.setToolTip(f"Информация о БД недоступна: {e}")

    def on_developer_double_click(self):
        """Двойной клик на имени разработчика - PIN диалог"""
        if not self.unlocked and self.require_pin:
            dialog = PinDialog(self.correct_pin, self)
            if dialog.exec() == QDialog.Accepted and dialog.is_authenticated:
                self.unlock_interface()
                self.log_text.append("✅ Интерфейс разблокирован")
            else:
                self.log_text.append("❌ Авторизация отменена")

    def on_log_double_click(self, event):
        """Обработчик двойного клика на логе - открывает лог в текстовом редакторе"""
        try:
            import tempfile
            
            # Получаем текст лога
            log_content = self.log_text.toPlainText()
            
            if not log_content.strip():
                self.statusBar().showMessage("ℹ Лог выполнения пуст", 3000)
                return
            
            # Создаем временный файл
            with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.txt', delete=False) as f:
                f.write("=" * 80 + "\n")
                f.write("BOM Categorizer - Лог выполнения\n")
                f.write("=" * 80 + "\n\n")
                f.write(log_content)
                temp_file = f.name
            
            # Открываем в системном текстовом редакторе
            if platform.system() == 'Windows':
                os.startfile(temp_file)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.Popen(['open', temp_file])
            else:  # Linux
                subprocess.Popen(['xdg-open', temp_file])
            
            self.log_text.append(f"\n📄 Лог открыт в текстовом редакторе: {temp_file}\n")
            
        except Exception as e:
            QMessageBox.warning(
                self,
                "Предупреждение",
                f"Не удалось открыть лог в текстовом редакторе:\n{str(e)}"
            )

    def on_show_size_menu(self, event):
        """Показать меню размеров окна"""
        from PySide6.QtCore import QPoint
        
        menu = QMenu(self)
        
        # Применяем шрифт меню с учётом scale_factor (та же логика что для основных меню)
        menu_scale = max(self.scale_factor + 0.2, 0.9)
        menu_font_size = max(7, int(round(9 * menu_scale)))
        menu_font = QFont(get_system_font(), menu_font_size)
        menu.setFont(menu_font)
        
        # Предустановленные размеры
        sizes = [
            ("По умолчанию (720×900)", 720, 900),
            ("Компактный (720×792)", 720, 792),
            ("Средний (800×850)", 800, 850),
            ("Большой (900×900)", 900, 900),
            ("Широкий (1000×800)", 1000, 800),
            ("HD (1280×720)", 1280, 720),
        ]
        
        for label, w, h in sizes:
            action = QAction(label, self)
            action.triggered.connect(lambda checked=False, width=w, height=h: self.set_window_size(width, height))
            menu.addAction(action)
        
        menu.addSeparator()
        
        save_action = QAction("📌 Сохранить текущий размер", self)
        save_action.triggered.connect(self.save_current_window_size)
        menu.addAction(save_action)
        
        # Показываем меню у метки размера окна
        menu.exec(self.size_label.mapToGlobal(QPoint(0, self.size_label.height())))
    
    def set_window_size(self, width: int, height: int):
        """Устанавливает размер окна"""
        self.resize(width, height)
        self.save_window_size_to_config(width, height)
        self.statusBar().showMessage(f"✓ Размер окна изменен на {width}×{height}", 3000)
    
    def save_current_window_size(self):
        """Сохраняет текущий размер окна"""
        width = self.width()
        height = self.height()
        self.save_window_size_to_config(width, height)
        self.statusBar().showMessage(f"✓ Текущий размер окна ({width}×{height}) сохранен", 3000)
    
    def _apply_window_size_for_mode(self, mode: str):
        """Применяет размер окна для указанного режима из конфигурации"""
        if "window" not in self.cfg:
            self.cfg["window"] = {}
        
        window_cfg = self.cfg["window"]
        sizes_by_mode = window_cfg.get("sizes_by_mode", {})
        
        # Размеры по умолчанию для каждого режима
        default_sizes = {
            "simple": {"width": 820, "height": 580},
            "advanced": {"width": 820, "height": 810},
            "expert": {"width": 820, "height": 1220}
        }
        
        # Если используется старый формат конфигурации (width/height на верхнем уровне)
        if not sizes_by_mode and ("width" in window_cfg or "height" in window_cfg):
            # Мигрируем старый формат в новый
            old_width = window_cfg.get("width", default_sizes["simple"]["width"])
            old_height = window_cfg.get("height", default_sizes["simple"]["height"])
            sizes_by_mode = {
                "simple": {"width": old_width, "height": old_height},
                "advanced": default_sizes["advanced"],
                "expert": default_sizes["expert"]
            }
            # Сохраняем мигрированную конфигурацию в self.cfg
            window_cfg["sizes_by_mode"] = sizes_by_mode
            try:
                cfg_path = get_config_path()
                with open(cfg_path, "w", encoding="utf-8") as f:
                    json.dump(self.cfg, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
        
        # Получаем размеры для текущего режима
        mode_sizes = sizes_by_mode.get(mode, default_sizes.get(mode, default_sizes["simple"]))
        width = mode_sizes.get("width", default_sizes[mode]["width"])
        height = mode_sizes.get("height", default_sizes[mode]["height"])
        
        self.resize(width, height)
    
    def save_window_size_to_config(self, width: int, height: int):
        """Сохраняет размер окна для текущего режима в конфигурацию"""
        if "window" not in self.cfg:
            self.cfg["window"] = {}
        
        window_cfg = self.cfg["window"]
        if "sizes_by_mode" not in window_cfg:
            window_cfg["sizes_by_mode"] = {}
        
        mode = self.current_view_mode
        if mode not in window_cfg["sizes_by_mode"]:
            window_cfg["sizes_by_mode"][mode] = {}
        
        window_cfg["sizes_by_mode"][mode]["width"] = width
        window_cfg["sizes_by_mode"][mode]["height"] = height
        
        # Сохраняем конфигурацию
        try:
            cfg_path = get_config_path()
            with open(cfg_path, "w", encoding="utf-8") as f:
                json.dump(self.cfg, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Не удалось сохранить размер окна в конфигурацию: {e}")

    def lock_interface(self):
        """Ограничивает доступ к расширенным режимам до ввода PIN"""
        self.unlocked = False
        self.update_mode_action_permissions()
        self.apply_view_mode(initial=True)

    def unlock_interface(self):
        """Разблокировка интерфейса"""
        self.unlocked = True
        self.update_mode_action_permissions()
        self.apply_view_mode(initial=True)

        # После разблокировки автоматически возвращаем сохраненный режим пользователя
        if self.pin_forced_simple and self.preferred_view_mode != "simple":
            preferred = self.preferred_view_mode
            self.pin_forced_simple = False
            self.set_view_mode(preferred)

    def resizeEvent(self, event):
        """Обработка изменения размера окна"""
        super().resizeEvent(event)
        if hasattr(self, 'size_label'):
            self.size_label.setText(f"📐 {self.width()}×{self.height()}")
    
    def closeEvent(self, event):
        """Обработка закрытия окна - настройки НЕ сохраняются"""
        # Настройки не сохраняются - приложение всегда открывается с настройками из config_qt.json
        event.accept()

    # =======================
    # Методы меню
    # =======================
    
    # Database methods moved to database_handlers.py (DatabaseHandlersMixin)


    def show_context_help(self):
        """Показывает контекстную помощь для текущего элемента"""
        # Определяем виджет под курсором мыши (более точный способ)
        cursor_pos = QCursor.pos()
        widget_under_cursor = QApplication.widgetAt(cursor_pos)
        
        # Если виджет под курсором не найден, пробуем виджет с фокусом
        if widget_under_cursor is None:
            widget_under_cursor = self.focusWidget()
        
        # Если все еще нет, пробуем найти родительский виджет
        if widget_under_cursor is None:
            widget_under_cursor = self
        
        help_text = self._get_context_help(widget_under_cursor)
        
        if help_text:
            QMessageBox.information(
                self,
                "Контекстная помощь",
                help_text
            )
        else:
            # Общая справка, если не найдена помощь для элемента
            QMessageBox.information(
                self,
                "Контекстная помощь",
                "📖 <b>Контекстная помощь</b><br><br>"
                "Наведите курсор на элемент интерфейса и нажмите <b>F1</b> для получения справки.<br><br>"
                "Или выберите элемент и нажмите <b>F1</b> для получения подробной информации.<br><br>"
                "<b>Доступные элементы с помощью:</b><br>"
                "• Кнопки (Добавить файлы, Запустить обработку, и т.д.)<br>"
                "• Поля ввода<br>"
                "• Списки файлов<br>"
                "• Область лога<br>"
                "• Меню и пункты меню"
            )
    
    def _get_context_help(self, widget) -> str:
        """Возвращает текст помощи для конкретного виджета"""
        if widget is None:
            return ""
        
        widget_type = type(widget).__name__
        widget_text = ""
        widget_object_name = widget.objectName() if hasattr(widget, 'objectName') else ""
        
        # Пытаемся получить текст из виджета разными способами
        if hasattr(widget, 'text'):
            widget_text = widget.text()
        elif hasattr(widget, 'toolTip'):
            widget_text = widget.toolTip()
        elif hasattr(widget, 'windowTitle'):
            widget_text = widget.windowTitle()
        elif hasattr(widget, 'placeholderText'):
            widget_text = widget.placeholderText()
        
        # Если текст пустой, пробуем получить из родительского виджета (для кнопок в меню)
        if not widget_text and hasattr(widget, 'parent'):
            parent = widget.parent()
            if parent and hasattr(parent, 'text'):
                widget_text = parent.text()
        
        # Нормализуем текст (убираем эмодзи и лишние пробелы)
        widget_text_clean = widget_text.strip()
        # Убираем эмодзи для поиска
        widget_text_clean = re.sub(r'[^\w\s]', '', widget_text_clean).strip()
        
        # База знаний для различных элементов
        help_map = {
            'QPushButton': {
                'Добавить файлы': '📂 <b>Добавить файлы</b><br><br>'
                    'Добавляет BOM файлы для обработки. Поддерживаются форматы:<br>'
                    '• Excel (.xlsx) - основной формат<br>'
                    '• Word (.docx, .doc) - автоматически конвертируется<br>'
                    '• Текст (.txt) - простой текстовый формат<br><br>'
                    'Можно выбрать несколько файлов одновременно.<br>'
                    'Также можно перетащить файлы прямо в окно приложения.<br><br>'
                    '<b>Горячая клавиша:</b> Ctrl+O',
                '➕ Добавить файлы': '📂 <b>Добавить файлы</b><br><br>'
                    'Добавляет BOM файлы для обработки. Поддерживаются форматы:<br>'
                    '• Excel (.xlsx) - основной формат<br>'
                    '• Word (.docx, .doc) - автоматически конвертируется<br>'
                    '• Текст (.txt) - простой текстовый формат<br><br>'
                    'Можно выбрать несколько файлов одновременно.<br>'
                    'Также можно перетащить файлы прямо в окно приложения.<br><br>'
                    '<b>Горячая клавиша:</b> Ctrl+O',
                '🗑️ Очистить список': '🗑️ <b>Очистить список</b><br><br>'
                    'Удаляет все файлы из списка обработки.<br>'
                    'Количество экземпляров для каждого файла сбрасывается.',
                'Очистить список': '🗑️ <b>Очистить список</b><br><br>'
                    'Удаляет все файлы из списка обработки.<br>'
                    'Количество экземпляров для каждого файла сбрасывается.',
                '▶️ Запустить обработку': '🚀 <b>Запустить обработку</b><br><br>'
                    'Начинает обработку выбранных BOM файлов с автоматической классификацией компонентов.<br><br>'
                    '<b>Процесс:</b><br>'
                    '1. Конвертация .doc файлов в .docx (если нужно)<br>'
                    '2. Парсинг BOM файлов<br>'
                    '3. Автоматическая классификация по базе данных<br>'
                    '4. Создание выходного Excel файла с категориями<br><br>'
                    '<b>Горячая клавиша:</b> Ctrl+R',
                '🚀 Запустить обработку': '🚀 <b>Запустить обработку</b><br><br>'
                    'Начинает обработку выбранных BOM файлов с автоматической классификацией компонентов.<br><br>'
                    '<b>Процесс:</b><br>'
                    '1. Конвертация .doc файлов в .docx (если нужно)<br>'
                    '2. Парсинг BOM файлов<br>'
                    '3. Автоматическая классификация по базе данных<br>'
                    '4. Создание выходного Excel файла с категориями<br><br>'
                    '<b>Горячая клавиша:</b> Ctrl+R',
                'Запустить обработку': '🚀 <b>Запустить обработку</b><br><br>'
                    'Начинает обработку выбранных BOM файлов с автоматической классификацией компонентов.<br><br>'
                    '<b>Процесс:</b><br>'
                    '1. Конвертация .doc файлов в .docx (если нужно)<br>'
                    '2. Парсинг BOM файлов<br>'
                    '3. Автоматическая классификация по базе данных<br>'
                    '4. Создание выходного Excel файла с категориями<br><br>'
                    '<b>Горячая клавиша:</b> Ctrl+R',
                '🔄 Интерактивная классификация': '🎯 <b>Интерактивная классификация</b><br><br>'
                    'Открывает диалог для ручной классификации нераспределенных компонентов.<br><br>'
                    '<b>Использование:</b><br>'
                    '1. Выберите компонент из списка<br>'
                    '2. Выберите категорию<br>'
                    '3. Компонент будет добавлен в базу данных<br>'
                    '4. Повторите для всех нераспределенных компонентов',
                'Интерактивная классификация': '🎯 <b>Интерактивная классификация</b><br><br>'
                    'Открывает диалог для ручной классификации нераспределенных компонентов.<br><br>'
                    '<b>Использование:</b><br>'
                    '1. Выберите компонент из списка<br>'
                    '2. Выберите категорию<br>'
                    '3. Компонент будет добавлен в базу данных<br>'
                    '4. Повторите для всех нераспределенных компонентов',
                '⚡ Сравнить файлы': '🔍 <b>Сравнить файлы</b><br><br>'
                    'Сравнивает два BOM файла и показывает различия.<br><br>'
                    '<b>Требования:</b><br>'
                    '• Оба файла должны быть уже обработаны (с категориями)<br>'
                    '• Если файлы не обработаны, появится предупреждение<br><br>'
                    'Результат покажет добавленные, удаленные и измененные компоненты.',
                'Сравнить файлы': '🔍 <b>Сравнить файлы</b><br><br>'
                    'Сравнивает два BOM файла и показывает различия.<br><br>'
                    '<b>Требования:</b><br>'
                    '• Оба файла должны быть уже обработаны (с категориями)<br>'
                    '• Если файлы не обработаны, появится предупреждение<br><br>'
                    'Результат покажет добавленные, удаленные и измененные компоненты.',
                'Выбрать': '📁 <b>Выбрать файл</b><br><br>'
                    'Открывает диалог выбора файла для сохранения результата обработки.',
            },
            'QLineEdit': {
                'Выходной файл': '📄 <b>Выходной файл</b><br><br>'
                    'Имя файла для сохранения результата обработки.<br><br>'
                    '<b>По умолчанию:</b><br>'
                    '• Для одного файла: {имя_файла}_categorized.xlsx<br>'
                    '• Для нескольких файлов: categorized.xlsx<br>'
                    '• Сохраняется в папке первого входного файла<br>'
                    '• Если файл существует, добавляется _1, _2 и т.д.',
            },
            'QListWidget': {
                '': '📋 <b>Список файлов</b><br><br>'
                    'Список выбранных файлов для обработки.<br><br>'
                    '<b>Действия:</b><br>'
                    '• Выберите файл для изменения количества экземпляров<br>'
                    '• Двойной клик открывает диалог изменения количества<br>'
                    '• Файлы можно удалить через контекстное меню',
            },
            'QTextEdit': {
                'Лог выполнения': '📝 <b>Лог выполнения</b><br><br>'
                    'Отображает информацию о процессе обработки файлов.<br><br>'
                    '<b>Функции:</b><br>'
                    '• Показывает прогресс обработки<br>'
                    '• Отображает ошибки и предупреждения<br>'
                    '• Двойной клик открывает лог в текстовом редакторе<br>'
                    '• В экспертном режиме можно включить временные метки',
            },
            'QTextBrowser': {
                '': '📖 <b>Текстовая область</b><br><br>'
                    'Область для отображения текстовой информации с поддержкой HTML и ссылок.',
            },
            'QLabel': {
                '': '🏷️ <b>Метка</b><br><br>'
                    'Текстовая метка для отображения информации или подсказок.',
            },
        }
        
        # Ищем помощь для конкретного виджета по тексту
        if widget_type in help_map:
            # Сначала ищем по полному тексту
            if widget_text in help_map[widget_type]:
                return help_map[widget_type][widget_text]
            # Затем ищем по очищенному тексту
            if widget_text_clean in help_map[widget_type]:
                return help_map[widget_type][widget_text_clean]
            # Ищем частичное совпадение
            for key, value in help_map[widget_type].items():
                if key and (key.lower() in widget_text.lower() or widget_text.lower() in key.lower()):
                    return value
            # Если есть общая помощь для типа виджета
            if '' in help_map[widget_type]:
                return help_map[widget_type]['']
        
        # Проверяем, является ли это кнопкой меню
        if widget_type == 'QAction':
            action_text = widget.text() if hasattr(widget, 'text') else ""
            if action_text:
                # Ищем в базе знаний по тексту действия
                for key, value in help_map.get('QPushButton', {}).items():
                    if key.lower() in action_text.lower() or action_text.lower() in key.lower():
                        return value
        
        # Общая помощь по типу виджета
        general_help = {
            'QPushButton': '🔘 <b>Кнопка</b><br><br>Кнопка для выполнения действия. Нажмите для активации.',
            'QLineEdit': '📝 <b>Поле ввода</b><br><br>Поле ввода текста. Введите значение или используйте кнопку "Выбрать..." для выбора файла.',
            'QSpinBox': '🔢 <b>Числовое поле</b><br><br>Поле для ввода числового значения. Используйте стрелки или введите значение вручную.',
            'QCheckBox': '☑️ <b>Флажок</b><br><br>Флажок для включения/выключения опции.',
            'QListWidget': '📋 <b>Список</b><br><br>Список элементов. Выберите элемент для работы с ним.',
            'QTextEdit': '📄 <b>Текстовое поле</b><br><br>Текстовое поле для отображения и редактирования информации.',
            'QMenu': '📋 <b>Меню</b><br><br>Меню для доступа к функциям приложения.',
            'QMenuBar': '📋 <b>Строка меню</b><br><br>Главное меню приложения с разделами: Файл, Вид, База данных, Помощь.',
        }
        
        if widget_type in general_help:
            return general_help[widget_type]
        
        # Если ничего не найдено, возвращаем информацию о виджете
        widget_info = f"<b>{widget_type}</b>"
        if widget_text:
            widget_info += f"<br><b>Текст:</b> {widget_text}"
        if widget_object_name:
            widget_info += f"<br><b>Имя:</b> {widget_object_name}"
        widget_info += "<br><br>Для этого элемента пока нет подробной справки."
        
        return widget_info
    
    def show_knowledge_base(self):
        """Показывает базу знаний с поиском"""
        dialog = QDialog(self)
        dialog.setWindowTitle("📚 База знаний")
        dialog.resize(800, 600)
        
        layout = QVBoxLayout()
        
        # Поле поиска
        search_layout = QHBoxLayout()
        search_label = QLabel("🔍 Поиск:")
        search_input = QLineEdit()
        search_input.setPlaceholderText("Введите ключевое слово для поиска...")
        search_button = QPushButton("Найти")
        
        search_layout.addWidget(search_label)
        search_layout.addWidget(search_input)
        search_layout.addWidget(search_button)
        layout.addLayout(search_layout)
        
        # Область с результатами
        results_text = QTextEdit()
        results_text.setReadOnly(True)
        results_text.setFont(QFont("Consolas", 10))
        layout.addWidget(results_text)
        
        # Кнопки
        button_layout = QHBoxLayout()
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.accept)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        # База знаний
        knowledge_base = {
            'обработка': {
                'title': 'Обработка BOM файлов',
                'content': '''
<b>Как обработать BOM файлы:</b>
1. Нажмите "➕ Добавить файлы" и выберите файлы (XLSX, DOCX, TXT)
2. Укажите количество экземпляров для каждого файла (если нужно)
3. Выберите выходной файл (по умолчанию сохраняется в папке первого файла)
4. Нажмите "🚀 Запустить обработку"

<b>Поддерживаемые форматы:</b>
• Excel (.xlsx) - основной формат
• Word (.docx, .doc) - автоматически конвертируется
• Текст (.txt) - простой текстовый формат

<b>Результат:</b>
Создается Excel файл с листами по категориям компонентов.
'''
            },
            'классификация': {
                'title': 'Классификация компонентов',
                'content': '''
<b>Автоматическая классификация:</b>
Компоненты автоматически классифицируются по базе данных.

<b>Интерактивная классификация:</b>
Если есть нераспределенные компоненты:
1. После обработки откроется диалог
2. Выберите компонент из списка
3. Выберите категорию
4. Компонент будет добавлен в базу данных

<b>Категории:</b>
• Резисторы, Конденсаторы, Индуктивности
• Микросхемы, Диоды, Транзисторы
• Разъемы, Механика, Прочее
'''
            },
            'база данных': {
                'title': 'База данных компонентов',
                'content': '''
<b>Управление базой данных:</b>
• <b>Статистика</b> - просмотр информации о БД
• <b>Экспорт в Excel</b> - сохранение БД для редактирования
• <b>Импорт из Excel</b> - загрузка БД из файла
• <b>Резервное копирование</b> - создание бэкапа
• <b>Посмотреть базу</b> - просмотр истории изменений
• <b>Очистить базу</b> - удаление всех компонентов

<b>Версионирование:</b>
База данных использует версионирование X.Y:
• X увеличивается при импорте из файлов
• Y увеличивается при ручном добавлении
'''
            },
            'сравнение': {
                'title': 'Сравнение BOM файлов',
                'content': '''
<b>Как сравнить файлы:</b>
1. Выберите первый файл (базовый)
2. Выберите второй файл (новый)
3. Укажите файл результата
4. Нажмите "⚡ Сравнить файлы"

<b>Результат:</b>
Создается Excel файл с листами:
• "Добавлено" - новые компоненты
• "Удалено" - удаленные компоненты
• "Изменено" - измененные компоненты
'''
            },
            'масштаб': {
                'title': 'Масштабирование интерфейса',
                'content': '''
<b>Изменение масштаба:</b>
• Меню "Вид" → "Масштабирование интерфейса"
• Горячие клавиши: Ctrl+Plus, Ctrl+Minus, Ctrl+0

<b>Доступные масштабы:</b>
70%, 80%, 90%, 100%, 110%, 125%

Масштаб сохраняется в настройках.
'''
            },
            'режимы': {
                'title': 'Режимы работы',
                'content': '''
<b>Простой режим:</b>
Упрощенный интерфейс (по умолчанию).
Скрыты: сравнение файлов, лог, меню базы данных.

<b>Расширенный режим:</b>
Все функции доступны.

<b>Экспертный режим:</b>
Дополнительные настройки:
• Временные метки в логе
• Автоматическое открытие папки результата
'''
            },
        }
        
        def update_results(query=""):
            """Обновляет результаты поиска"""
            if not query.strip():
                # Показываем все статьи
                html = "<h2>📚 База знаний</h2><br>"
                html += "<p>Введите запрос в поле поиска или выберите тему ниже:</p><br>"
                for key, article in knowledge_base.items():
                    html += f'<h3>{article["title"]}</h3>'
                    html += f'<p>{article["content"]}</p>'
                    html += "<hr>"
            else:
                # Поиск по ключевым словам
                query_lower = query.lower()
                html = f"<h2>🔍 Результаты поиска: '{query}'</h2><br>"
                found = False
                for key, article in knowledge_base.items():
                    if query_lower in key.lower() or query_lower in article['title'].lower() or query_lower in article['content'].lower():
                        found = True
                        html += f'<h3>{article["title"]}</h3>'
                        html += f'<p>{article["content"]}</p>'
                        html += "<hr>"
                if not found:
                    html += "<p>Ничего не найдено. Попробуйте другие ключевые слова.</p>"
            
            results_text.setHtml(html)
        
        def on_search():
            update_results(search_input.text())
        
        search_button.clicked.connect(on_search)
        search_input.returnPressed.connect(on_search)
        
        # Показываем все статьи при открытии
        update_results()
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def show_dragdrop_help(self):
        """Показывает руководство по использованию Drag & Drop"""
        help_text = """
<h1 style="color: #89b4fa;">🎯 Улучшенный Drag & Drop</h1>

<h2 style="color: #94e2d5;">Как включить</h2>
<ol>
<li>Переключитесь в <b>Экспертный режим</b><br>
    (Вид → Режим работы → Экспертный режим)</li>
<li>В секции <b>Экспертные инструменты</b> найдите чекбокс:<br>
    <i>🎯 Улучшенный Drag & Drop</i></li>
<li>Установите галочку - функция активируется мгновенно!</li>
</ol>

<h2 style="color: #94e2d5;">Основные возможности</h2>

<h3 style="color: #f9e2af;">📁 Перетаскивание из проводника</h3>
<ul>
<li>Откройте папку с файлами в проводнике Windows</li>
<li>Выделите нужные файлы (.xlsx, .docx, .txt)</li>
<li>Перетащите их в список <b>Входные файлы</b></li>
<li>Зона подсветится синей рамкой при перетаскивании</li>
</ul>

<h3 style="color: #f9e2af;">🔄 Изменение порядка файлов</h3>
<ul>
<li>Зажмите левую кнопку мыши на файле в списке</li>
<li>Перетащите файл на нужную позицию</li>
<li>Отпустите кнопку мыши</li>
<li>Порядок обработки соответствует порядку в списке</li>
</ul>

<h3 style="color: #f9e2af;">🖱️ Контекстное меню (ПКМ)</h3>
<p>Щелкните <b>правой кнопкой мыши</b> на любом файле в списке:</p>
<ul>
<li><b>📄 Открыть файл</b> - открывает файл в Excel/Word/Notepad</li>
<li><b>📁 Показать в проводнике</b> - открывает папку и выделяет файл</li>
<li><b>📋 Копировать путь</b> - копирует полный путь к файлу</li>
<li><b>🗑️ Удалить из списка</b> - удаляет файл из списка (не физически)</li>
</ul>

<h2 style="color: #94e2d5;">Примеры использования</h2>

<h3 style="color: #cba6f7;">Пример 1: Быстрое добавление файлов</h3>
<p style="margin-left: 20px;">
1. Откройте папку с BOM-файлами<br>
2. Выделите все нужные файлы (Ctrl+Click)<br>
3. Перетащите в окно программы<br>
4. Готово! Все файлы добавлены
</p>

<h3 style="color: #cba6f7;">Пример 2: Изменение приоритета</h3>
<p style="margin-left: 20px;">
Нужно чтобы "БОМ_основной.xlsx" обработался первым:<br>
• Перетащите его в начало списка<br>
• Файлы обрабатываются сверху вниз
</p>

<h3 style="color: #cba6f7;">Пример 3: Быстрое открытие файла</h3>
<p style="margin-left: 20px;">
• ПКМ на файле → "📄 Открыть файл"<br>
• Файл откроется в Excel/Word<br>
• Удобно для быстрой проверки
</p>

<h3 style="color: #cba6f7;">Пример 4: Отправка пути коллеге</h3>
<p style="margin-left: 20px;">
• ПКМ на файле → "📋 Копировать путь"<br>
• Ctrl+V в мессенджер/email<br>
• Коллега получит точный путь к файлу
</p>

<h2 style="color: #94e2d5;">Горячие клавиши</h2>
<table style="border-collapse: collapse; width: 100%;">
<tr style="background-color: #313244;">
    <th style="padding: 8px; text-align: left; border: 1px solid #45475a;">Действие</th>
    <th style="padding: 8px; text-align: left; border: 1px solid #45475a;">Клавиша</th>
</tr>
<tr>
    <td style="padding: 8px; border: 1px solid #45475a;">Выделить все файлы</td>
    <td style="padding: 8px; border: 1px solid #45475a;"><b>Ctrl+A</b></td>
</tr>
<tr style="background-color: #1e1e2e;">
    <td style="padding: 8px; border: 1px solid #45475a;">Множественный выбор</td>
    <td style="padding: 8px; border: 1px solid #45475a;"><b>Ctrl+Click</b></td>
</tr>
<tr>
    <td style="padding: 8px; border: 1px solid #45475a;">Диапазон выбора</td>
    <td style="padding: 8px; border: 1px solid #45475a;"><b>Shift+Click</b></td>
</tr>
<tr style="background-color: #1e1e2e;">
    <td style="padding: 8px; border: 1px solid #45475a;">Удалить выбранное</td>
    <td style="padding: 8px; border: 1px solid #45475a;"><b>Delete</b></td>
</tr>
</table>

<h2 style="color: #94e2d5;">⚠️ Важные замечания</h2>
<ul>
<li>Поддерживаются только файлы: .xlsx, .docx, .doc, .txt</li>
<li>При перетаскивании из проводника файлы не перемещаются - добавляется только ссылка</li>
<li>Для отключения функции требуется перезапуск программы</li>
<li>Рекомендуется не добавлять более 100 файлов одновременно</li>
</ul>

<h2 style="color: #94e2d5;">💡 Советы</h2>
<ul>
<li>Используйте ПКМ → "Показать в проводнике" для быстрого доступа к папке</li>
<li>Копирование пути удобно для отправки локации файла другим пользователям</li>
<li>Изменяйте порядок файлов для контроля последовательности обработки</li>
<li>Визуальная подсветка показывает что файлы можно сбросить в эту область</li>
</ul>

<hr style="border: 1px solid #45475a; margin: 20px 0;">

<p style="text-align: center; color: #6c7086;">
<i>Экспериментальная функция в ветке experimental/new-feature</i><br>
Полная документация: <b>DRAG_DROP_README.md</b>
</p>
"""
        
        # Создаем диалог
        dialog = QDialog(self)
        dialog.setWindowTitle("🎯 Как использовать Drag & Drop")
        dialog.resize(800, 700)
        
        # Применяем шрифт диалога с учётом scale_factor
        dialog_font_size = int(12 * self.scale_factor)
        dialog.setFont(QFont(get_system_font(), dialog_font_size))
        
        layout = QVBoxLayout()
        
        # Текст с прокруткой
        text_widget = QTextBrowser()
        text_widget.setOpenExternalLinks(True)
        text_widget.setHtml(help_text)
        
        # Применяем шрифт с учётом scale_factor
        font_size = int(10 * self.scale_factor)
        text_widget.setFont(QFont(get_system_font(), font_size))
        layout.addWidget(text_widget)
        
        # Кнопки
        button_layout = QHBoxLayout()
        
        open_readme_btn = QPushButton("📄 Открыть полную документацию")
        open_readme_btn.clicked.connect(lambda: self._open_dragdrop_readme())
        button_layout.addWidget(open_readme_btn)
        
        button_layout.addStretch()
        
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def _open_dragdrop_readme(self):
        """Открывает файл DRAG_DROP_README.md"""
        import os
        readme_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "DRAG_DROP_README.md")
        if os.path.exists(readme_path):
            try:
                if platform.system() == 'Windows':
                    os.startfile(readme_path)
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.Popen(['open', readme_path])
                else:  # Linux
                    subprocess.Popen(['xdg-open', readme_path])
            except Exception as e:
                QMessageBox.warning(self, "Ошибка", f"Не удалось открыть файл:\n{e}")
        else:
            QMessageBox.warning(self, "Файл не найден", f"Файл DRAG_DROP_README.md не найден:\n{readme_path}")
    
    def keyPressEvent(self, event):
        """Обработка нажатий клавиш для контекстной помощи"""
        if event.key() == Qt.Key_F1:
            self.show_context_help()
            event.accept()
        else:
            super().keyPressEvent(event)
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        """Обработка входа перетаскиваемого объекта"""
        if event.mimeData().hasUrls():
            # Проверяем, что это файлы с поддерживаемыми расширениями
            urls = event.mimeData().urls()
            supported_extensions = ['.xlsx', '.docx', '.doc', '.txt']
            has_supported_file = False
            
            for url in urls:
                file_path = url.toLocalFile()
                if file_path:
                    ext = os.path.splitext(file_path)[1].lower()
                    if ext in supported_extensions:
                        has_supported_file = True
                        break
            
            if has_supported_file:
                event.acceptProposedAction()
            else:
                event.ignore()
        else:
            event.ignore()
    
    def dropEvent(self, event: QDropEvent):
        """Обработка сброса файлов"""
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            supported_extensions = ['.xlsx', '.docx', '.doc', '.txt']
            files_added = 0
            
            for url in urls:
                file_path = url.toLocalFile()
                if file_path and os.path.isfile(file_path):
                    ext = os.path.splitext(file_path)[1].lower()
                    if ext in supported_extensions:
                        # Проверяем наличие файла (без учета регистра)
                        exists = False
                        for existing_path in self.input_files:
                            if existing_path.lower() == file_path.lower():
                                exists = True
                                break
                        
                        if not exists:
                            self.input_files[file_path] = 1
                            self.last_input_file = file_path  # Сохраняем последний добавленный файл
                            files_added += 1
            
            if files_added > 0:
                self.update_listbox()
                self.update_output_filename()
                # Показываем уведомление в status bar (автоматически исчезнет через 5 секунд)
                self.statusBar().showMessage(
                    f"✓ Добавлено файлов: {files_added}. Используйте Ctrl+R для запуска обработки.",
                    5000  # 5 секунд
                )
            else:
                QMessageBox.warning(
                    self,
                    "Неподдерживаемый формат",
                    "Поддерживаются только файлы:\n"
                    "XLSX, DOCX, DOC, TXT"
                )
            
            event.acceptProposedAction()
        else:
            event.ignore()

    # ==================== Управление представлением ====================

    def apply_scale_factor(self):
        """Применяет текущий коэффициент масштабирования"""
        font_size = max(8, int(round(self.base_font_size * self.scale_factor)))
        font = QFont(get_system_font(), font_size)
        
        # Применяем масштаб глобально через QApplication (для всех новых виджетов)
        if self.app:
            self.app.setFont(font)
        
        # Применяем к главному окну
        self.setFont(font)
        
        # Применяем рекурсивно ко всем дочерним виджетам (кроме меню)
        self._apply_font_recursive(self, font)
        
        # Применяем шрифт для меню - на 20% крупнее основного интерфейса, но не меньше 90%
        # Если основной интерфейс 70%, то меню 90%; если 80%, то меню 100%
        from PySide6.QtWidgets import QMenu, QMenuBar
        menubar = self.menuBar()
        if menubar:
            # Меню всегда на 0.2 (20%) крупнее, но минимум 0.9 (90%)
            menu_scale = max(self.scale_factor + 0.2, 0.9)
            
            menu_base_size = 9  # Базовый размер для меню
            menu_font_size = max(7, int(round(menu_base_size * menu_scale)))
            menu_font = QFont(get_system_font(), menu_font_size)
            
            # Устанавливаем шрифт для самого menubar (названия "Файл", "Вид" и т.д.)
            menubar.setFont(menu_font)
            
            # ПРИНУДИТЕЛЬНО через stylesheet - это единственный способ изменить шрифт menubar
            menubar_style = f"QMenuBar {{ font-size: {menu_font_size}pt; font-family: '{get_system_font()}'; }}"
            menubar_style += f"QMenuBar::item {{ font-size: {menu_font_size}pt; font-family: '{get_system_font()}'; }}"
            menubar.setStyleSheet(menubar_style)
            
            # Устанавливаем шрифт для выпадающих меню
            for menu in self.findChildren(QMenu):
                menu.setFont(menu_font)
        
        # Обновляем размеры виджетов, заданные в пикселях
        self._update_widget_sizes()
        
        self.update_scale_actions()
    
    def _apply_font_recursive(self, widget, font):
        """Рекурсивно применяет шрифт ко всем дочерним виджетам"""
        from PySide6.QtWidgets import QMenu, QMenuBar
        
        # Применяем к текущему виджету
        current_font = widget.font()
        # Сохраняем семейство шрифта, если оно было специально задано
        if current_font.family() != font.family() and current_font.family() != get_system_font():
            # Используем существующее семейство, но обновляем размер
            current_font.setPointSize(font.pointSize())
            widget.setFont(current_font)
        else:
            widget.setFont(font)
        
        # Применяем рекурсивно ко всем дочерним виджетам
        for child in widget.findChildren(QWidget):
            # ПРОПУСКАЕМ меню - они должны сохранять системный размер шрифта
            if isinstance(child, (QMenu, QMenuBar)):
                continue
                
            child_font = child.font()
            if child_font.family() != font.family() and child_font.family() != get_system_font():
                # Сохраняем специальное семейство шрифта, но обновляем размер
                child_font.setPointSize(font.pointSize())
                child.setFont(child_font)
            else:
                child.setFont(font)
    
    def _update_widget_sizes(self):
        """Обновляет размеры виджетов в соответствии с масштабом"""
        # Базовые размеры (для масштаба 1.0)
        base_button_height = 32
        base_input_height = 28
        base_spacing = 10
        
        # Масштабированные значения
        scaled_button_height = int(base_button_height * self.scale_factor)
        scaled_input_height = int(base_input_height * self.scale_factor)
        scaled_spacing = int(base_spacing * self.scale_factor)
        
        # Обновляем высоту списка файлов
        if hasattr(self, 'files_list') and self.files_list:
            scaled_height = int(100 * self.scale_factor)
            self.files_list.setMaximumHeight(scaled_height)
            self.files_list.setMinimumHeight(int(60 * self.scale_factor))
        
        # Обновляем высоту лога
        if hasattr(self, 'log_text') and self.log_text:
            scaled_height = int(160 * self.scale_factor)
            self.log_text.setMaximumHeight(scaled_height)
            self.log_text.setMinimumHeight(int(100 * self.scale_factor))
        
        # Обновляем размеры всех кнопок
        for button in self.findChildren(QPushButton):
            button.setMinimumHeight(scaled_button_height)
            button.setMaximumHeight(scaled_button_height + 10)
        
        # Обновляем размеры полей ввода
        for line_edit in self.findChildren(QLineEdit):
            line_edit.setMinimumHeight(scaled_input_height)
            line_edit.setMaximumHeight(scaled_input_height + 10)
        
        # Обновляем размеры спинбоксов
        for spin_box in self.findChildren(QSpinBox):
            spin_box.setMinimumHeight(scaled_input_height)
            spin_box.setMaximumHeight(scaled_input_height + 10)
        
        # Обновляем интервалы в layouts
        for layout in self.findChildren(QVBoxLayout):
            if layout:
                layout.setSpacing(scaled_spacing)
        
        for layout in self.findChildren(QHBoxLayout):
            if layout:
                layout.setSpacing(scaled_spacing)
        
        # Принудительно обновляем геометрию
        self.updateGeometry()
        # НЕ вызываем adjustSize() - это автоматически уменьшает окно!
        # Размер окна должен определяться config_qt.json, а не содержимым
        QApplication.processEvents()

    def update_scale_actions(self):
        """Обновляет состояние пунктов меню масштаба"""
        if not self.scale_actions:
            return
        for factor, action in self.scale_actions.items():
            if action is None:
                continue
            blocked = action.blockSignals(True)
            action.setChecked(abs(self.scale_factor - factor) < 0.001)
            action.blockSignals(blocked)

    def set_scale_factor(self, factor: float):
        """Устанавливает масштаб интерфейса"""
        if factor not in self.scale_levels:
            factor = min(self.scale_levels, key=lambda x: abs(x - factor))
        if abs(self.scale_factor - factor) < 0.001:
            self.update_scale_actions()
            return
        self.scale_factor = factor
        self.apply_scale_factor()
        self.save_ui_preferences()

    def _current_scale_index(self) -> int:
        if self.scale_factor in self.scale_levels:
            return self.scale_levels.index(self.scale_factor)
        closest = min(range(len(self.scale_levels)), key=lambda i: abs(self.scale_levels[i] - self.scale_factor))
        self.scale_factor = self.scale_levels[closest]
        return closest

    def on_zoom_in(self):
        index = self._current_scale_index()
        if index < len(self.scale_levels) - 1:
            self.set_scale_factor(self.scale_levels[index + 1])

    def on_zoom_out(self):
        index = self._current_scale_index()
        if index > 0:
            self.set_scale_factor(self.scale_levels[index - 1])

    def reset_scale(self):
        self.set_scale_factor(0.8)  # Сброс на масштаб по умолчанию (80%)

    def update_view_mode_actions(self):
        if not self.view_mode_actions:
            return
        for key, action in self.view_mode_actions.items():
            blocked = action.blockSignals(True)
            action.setChecked(key == self.current_view_mode)
            action.blockSignals(blocked)

    def update_mode_action_permissions(self):
        """Обновляет доступность пунктов меню смены режима"""
        if not self.view_mode_actions:
            return

        locked = self.require_pin and not self.unlocked

        for key, action in self.view_mode_actions.items():
            if action is None:
                continue
            if key == "simple":
                action.setEnabled(True)
                action.setToolTip("")
            else:
                action.setEnabled(not locked)
                if locked:
                    action.setToolTip("Доступно после ввода PIN-кода")
                else:
                    action.setToolTip("")

        if self.mode_menu is not None:
            if locked:
                self.mode_menu.setToolTip("Для переключения режимов введите PIN на панели разработчика")
            else:
                self.mode_menu.setToolTip("")

    def set_view_mode(self, mode: str):
        if mode not in ("simple", "advanced", "expert"):
            return
        if self.require_pin and not self.unlocked and mode != "simple":
            QMessageBox.information(
                self,
                "Требуется PIN",
                "Переключение в расширенный или экспертный режим доступно после ввода PIN-кода."
            )
            self.update_view_mode_actions()
            return
        if mode == self.current_view_mode:
            self.update_view_mode_actions()
            return
        self.current_view_mode = mode
        if mode != "expert":
            self.log_with_timestamps = False
            self.auto_open_output = False
            self.auto_export_pdf = False
            self.ai_classifier_enabled = False
            self.ai_auto_classify = False
        self.apply_view_mode()

    def apply_view_mode(self, initial: bool = False):
        simple = self.current_view_mode == "simple"
        expert = self.current_view_mode == "expert"

        if hasattr(self, "comparison_section") and self.comparison_section:
            self.comparison_section.setVisible(not simple)
        if hasattr(self, "log_section") and self.log_section:
            self.log_section.setVisible(expert)
        if hasattr(self, "expert_section") and self.expert_section:
            self.expert_section.setVisible(expert)

        if self.db_menu is not None:
            self.db_menu.menuAction().setVisible(not simple)
        
        # PDF поиск - меню доступно всегда, но AI функции только для разблокированных экспертов
        if hasattr(self, 'pdf_search_menu') and self.pdf_search_menu is not None:
            # Меню всегда активно (для локального поиска)
            self.pdf_search_menu.setEnabled(True)
            self.pdf_search_menu.setToolTip("Локальный поиск PDF доступен всегда, AI поиск - в экспертном режиме после разблокировки")
            
            # AI поиск и настройки API только для экспертов после разблокировки
            if hasattr(self, 'ai_pdf_action'):
                self.ai_pdf_action.setEnabled(expert and self.unlocked)
            if hasattr(self, 'pdf_settings_action'):
                self.pdf_settings_action.setEnabled(expert and self.unlocked)
            
        # Глобальный поиск виден только в расширенном и экспертном режимах
        if hasattr(self, 'global_search_menu'):
            is_advanced_or_expert = self.current_view_mode in ["advanced", "expert"]
            # Скрываем меню в простом режиме
            self.global_search_menu.menuAction().setVisible(is_advanced_or_expert)
            
            # Поле ввода активно только если разблокировано И режим подходящий
            if hasattr(self, 'global_search_input'):
                is_input_enabled = is_advanced_or_expert and self.unlocked
                self.global_search_input.setEnabled(is_input_enabled)
            
            # Обновляем tooltip
            if not self.unlocked:
                self.global_search_menu.setToolTip("Глобальный поиск доступен после разблокировки")
            elif is_advanced_or_expert:
                self.global_search_menu.setToolTip("Глобальный поиск по базе данных и файлам")
            else:
                self.global_search_menu.setToolTip("Глобальный поиск доступен в расширенном и экспертном режимах")

        if self.mode_label is not None:
            mode_titles = {
                "simple": ("Режим: Простой", "#fab387"),
                "advanced": ("Режим: Расширенный", "#89b4fa"),
                "expert": ("Режим: Эксперт", "#f38ba8"),
            }
            text, color = mode_titles.get(self.current_view_mode, ("Режим: Неизвестно", "#cdd6f4"))
            self.mode_label.setText(text)
            self.mode_label.setStyleSheet(f"QLabel {{ color: {color}; font-weight: bold; }}")

        if self.timestamp_checkbox is not None:
            self.timestamp_checkbox.blockSignals(True)
            self.timestamp_checkbox.setEnabled(expert)
            self.timestamp_checkbox.setChecked(self.log_with_timestamps if expert else False)
            self.timestamp_checkbox.blockSignals(False)

        if self.auto_open_output_checkbox is not None:
            self.auto_open_output_checkbox.blockSignals(True)
            self.auto_open_output_checkbox.setEnabled(expert)
            self.auto_open_output_checkbox.setChecked(self.auto_open_output if expert else False)
            self.auto_open_output_checkbox.blockSignals(False)

        self.update_mode_action_permissions()
        self.update_view_mode_actions()

        # При смене режима обновляем размер окна
        if not initial:
            self._apply_window_size_for_mode(self.current_view_mode)
            self.save_ui_preferences()

    def on_toggle_log_timestamps(self, state: int):
        self.log_with_timestamps = bool(state)
        self.save_ui_preferences()
        if self.log_text:
            message = "🕒 Временные метки лога включены" if self.log_with_timestamps else "🕒 Временные метки лога отключены"
            self.log_text.append(message)

    def on_toggle_auto_open_output(self, state: int):
        self.auto_open_output = bool(state)
        self.save_ui_preferences()
        if self.log_text:
            message = "📂 Автооткрытие папки результата включено" if self.auto_open_output else "📂 Автооткрытие папки результата отключено"
            self.log_text.append(message)
    
    def on_toggle_combine(self, state: int):
        """Включение/выключение суммарной комплектации"""
        self.combine = bool(state == Qt.Checked)
        if self.log_text:
            message = "📦 Суммарная комплектация включена" if self.combine else "📦 Суммарная комплектация отключена"
            self.log_text.append(message)
    
    def on_toggle_enhanced_dragdrop(self, state: int):
        """Включение/выключение улучшенного Drag & Drop"""
        from .drag_drop import enable_drag_drop_improvements
        
        enabled = bool(state)
        
        if enabled:
            # Включаем улучшенный D&D
            success = enable_drag_drop_improvements(self)
            if success and self.log_text:
                self.log_text.append("🎯 Улучшенный Drag & Drop включен")
                self.log_text.append("   • Перетаскивайте файлы для изменения порядка")
                self.log_text.append("   • ПКМ на файле для контекстного меню")
        else:
            # Для отключения нужен перезапуск приложения
            if self.log_text:
                self.log_text.append("⚠️ Для отключения требуется перезапуск приложения")
        
        self.save_ui_preferences()
    
    def open_interactive_cli(self):
        """Открывает интерактивную командную строку"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout
        from ..cli_interactive import InteractiveCLI
        
        # Создаем диалог
        dialog = QDialog(self)
        dialog.setWindowTitle("💻 Интерактивная командная строка")
        dialog.resize(900, 600)
        
        # Создаем layout
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Добавляем CLI виджет
        cli_widget = InteractiveCLI(self, dialog)
        layout.addWidget(cli_widget)
        
        # Показываем диалог
        dialog.exec()
        
        # Логируем
        if self.log_text:
            self.log_text.append("💻 Интерактивная командная строка закрыта")
    
    def export_last_result_to_pdf(self):
        """Экспортирует последний выходной файл в PDF"""
        # Проверяем, есть ли последний сгенерированный файл
        output_file = ""
        if hasattr(self, 'last_generated_output') and self.last_generated_output and os.path.exists(self.last_generated_output):
            output_file = self.last_generated_output
        else:
            # Иначе берем из поля ввода
            output_file = self.output_entry.text().strip() if hasattr(self, 'output_entry') else ""
        
        if not output_file or not os.path.exists(output_file):
            # Если выходного файла нет, предлагаем экспортировать входные файлы напрямую в PDF
            if not self.input_files:
                QMessageBox.warning(
                    self,
                    "Экспорт в PDF",
                    "Нет входных файлов для экспорта.\nДобавьте файлы и повторите попытку."
                )
                return
            
            # Спрашиваем пользователя, хочет ли он экспортировать входные файлы
            reply = QMessageBox.question(
                self,
                "Экспорт входных файлов в PDF",
                f"Выходной файл еще не создан.\n\n"
                f"Хотите экспортировать {len(self.input_files)} входных файлов напрямую в PDF?\n"
                f"Каждый файл будет сохранен с расширением .pdf",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Экспортируем входные файлы
            self._export_input_files_to_pdf()
            return
        
        try:
            from ..pdf_exporter import export_bom_to_pdf
            
            # Показываем диалог выбора места сохранения
            from PySide6.QtWidgets import QFileDialog
            pdf_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить PDF",
                os.path.splitext(output_file)[0] + ".pdf",
                "PDF Files (*.pdf)"
            )
            
            if not pdf_path:
                return  # Пользователь отменил
            
            # Собираем сводную информацию
            # Собираем сводную информацию
            summary_info = {
                "Исходных файлов": len(self.input_files),
                "Выходной файл": os.path.basename(output_file),
                "Версия БД": self.db.get_version() if hasattr(self, 'db') else "N/A",
                "Программа": f"BOM Categorizer {self.cfg.get('app_info', {}).get('version', 'dev')}",
                "Учитывать подбор": "Нет" if (hasattr(self, 'exclude_podbor_checkbox') and self.exclude_podbor_checkbox.isChecked()) else "Да",
                "Создавать TXT файлы": "Да" if (hasattr(self, 'txt_entry') and self.txt_entry.text().strip()) else "Нет"
            }
            
            QApplication.setOverrideCursor(Qt.WaitCursor)
            if self.log_text:
                self.log_text.append(f"📄 Экспорт в PDF: {os.path.basename(pdf_path)}")
            
            # Выполняем экспорт
            result_pdf = export_bom_to_pdf(
                output_file,
                pdf_path,
                with_summary=True,
                summary_info=summary_info
            )
            
            QApplication.restoreOverrideCursor()
            
            if self.log_text:
                self.log_text.append(f"✅ PDF создан: {result_pdf}")
            
            # Спрашиваем, открыть ли файл
            reply = QMessageBox.question(
                self,
                "Экспорт завершен",
                f"PDF документ успешно создан:\n{result_pdf}\n\nОткрыть файл?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self._open_file(result_pdf)
        
        except ImportError as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось загрузить модуль экспорта в PDF.\n"
                f"Возможно, не установлена библиотека reportlab.\n\n"
                f"Установите: pip install reportlab\n\n"
                f"Ошибка: {e}"
            )
        except Exception as e:
            QApplication.restoreOverrideCursor()
            if self.log_text:
                self.log_text.append(f"❌ Ошибка экспорта в PDF: {e}")
            QMessageBox.critical(
                self,
                "Ошибка экспорта",
                f"Не удалось создать PDF:\n{e}"
            )
    
    def _export_input_files_to_pdf(self):
        """Экспортирует входные файлы напрямую в PDF"""
        from ..pdf_exporter import export_bom_to_pdf
        from PySide6.QtWidgets import QFileDialog
        import subprocess
        
        # Спрашиваем куда сохранять файлы
        output_dir = QFileDialog.getExistingDirectory(
            self,
            "Выберите папку для сохранения PDF файлов",
            os.path.dirname(list(self.input_files.keys())[0]) if self.input_files else os.path.expanduser("~")
        )
        
        if not output_dir:
            return  # Пользователь отменил
        
        QApplication.setOverrideCursor(Qt.WaitCursor)
        exported_count = 0
        failed_count = 0
        
        for input_file in self.input_files.keys():
            try:
                base_name = os.path.splitext(os.path.basename(input_file))[0]
                pdf_path = os.path.join(output_dir, base_name + ".pdf")
                ext = os.path.splitext(input_file)[1].lower()
                
                if self.log_text:
                    self.log_text.append(f"📄 Экспорт: {os.path.basename(input_file)} → {os.path.basename(pdf_path)}")
                
                if ext in [".xlsx", ".xls"]:
                    # Экспорт Excel файлов через pdf_exporter
                    export_bom_to_pdf(input_file, pdf_path, with_summary=False)
                    exported_count += 1
                    
                elif ext in [".docx", ".doc"]:
                    # Экспорт DOCX/DOC - пробуем Word, потом LibreOffice
                    exported_this = False
                    
                    # Пробуем Microsoft Word (Windows)
                    if sys.platform == 'win32':
                        try:
                            import win32com.client
                            word = win32com.client.Dispatch("Word.Application")
                            word.Visible = False
                            
                            doc = word.Documents.Open(os.path.abspath(input_file))
                            # 17 = wdFormatPDF
                            doc.SaveAs(os.path.abspath(pdf_path), FileFormat=17)
                            doc.Close()
                            word.Quit()
                            
                            if os.path.exists(pdf_path):
                                exported_count += 1
                                exported_this = True
                                if self.log_text:
                                    self.log_text.append(f"  ✓ Экспортировано через MS Word")
                        except Exception as word_error:
                            if self.log_text:
                                self.log_text.append(f"  ⚠️  MS Word недоступен, пробуем LibreOffice...")
                    
                    # Если Word не сработал, пробуем LibreOffice
                    if not exported_this:
                        libreoffice_paths = [
                            '/Applications/LibreOffice.app/Contents/MacOS/soffice',  # macOS
                            '/usr/bin/libreoffice',  # Linux
                            '/usr/bin/soffice',      # Linux альтернатива
                            'C:\\Program Files\\LibreOffice\\program\\soffice.exe',  # Windows
                        ]
                        
                        soffice_path = None
                        for path in libreoffice_paths:
                            if os.path.exists(path):
                                soffice_path = path
                                break
                        
                        if soffice_path:
                            # Конвертируем через LibreOffice
                            cmd = [
                                soffice_path,
                                '--headless',
                                '--convert-to', 'pdf',
                                '--outdir', output_dir,
                                input_file
                            ]
                            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                            if result.returncode == 0 and os.path.exists(pdf_path):
                                exported_count += 1
                                exported_this = True
                                if self.log_text:
                                    self.log_text.append(f"  ✓ Экспортировано через LibreOffice")
                            else:
                                failed_count += 1
                                if self.log_text:
                                    self.log_text.append(f"  ❌ Ошибка конвертации: {os.path.basename(input_file)}")
                        else:
                            # Ни Word, ни LibreOffice не найдены
                            if self.log_text:
                                self.log_text.append(f"  ⚠️  MS Word и LibreOffice не найдены, пропуск: {os.path.basename(input_file)}")
                            failed_count += 1
                
                elif ext in [".txt"]:
                    # Экспорт TXT файлов через reportlab
                    try:
                        from reportlab.lib.pagesizes import A4
                        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                        from reportlab.lib.units import mm
                        from reportlab.pdfbase import pdfmetrics
                        from reportlab.pdfbase.ttfonts import TTFont
                        
                        # Читаем текстовый файл
                        with open(input_file, 'r', encoding='utf-8') as f:
                            text_content = f.read()
                        
                        # Создаем PDF
                        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
                        story = []
                        
                        # Регистрируем шрифт с поддержкой кириллицы
                        font_name = None
                        font_paths_to_try = [
                            ('/System/Library/Fonts/Supplemental/Arial Unicode.ttf', 'ArialUnicode'),  # macOS - BEST!
                            ('/System/Library/Fonts/Supplemental/Arial.ttf', 'Arial'),  # macOS
                            ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 'DejaVuSans'),  # Linux
                            ('C:\\Windows\\Fonts\\arial.ttf', 'Arial'),  # Windows
                        ]
                        
                        for font_path, font_reg_name in font_paths_to_try:
                            if os.path.exists(font_path):
                                try:
                                    pdfmetrics.registerFont(TTFont(font_reg_name, font_path))
                                    font_name = font_reg_name
                                    if self.log_text:
                                        self.log_text.append(f"  ✓ Используется шрифт: {font_reg_name}")
                                    break
                                except Exception as e:
                                    if self.log_text:
                                        self.log_text.append(f"  ⚠️ Ошибка загрузки {font_reg_name}: {e}")
                        
                        if not font_name:
                            # Последняя попытка - стандартный Helvetica (но без кириллицы)
                            font_name = 'Helvetica'
                            if self.log_text:
                                self.log_text.append(f"  ⚠️ Системный шрифт не найден, используется Helvetica (кириллица может не работать)")
                        
                        # Стили
                        styles = getSampleStyleSheet()
                        style = ParagraphStyle(
                            'CustomStyle',
                            parent=styles['Normal'],
                            fontName=font_name,
                            fontSize=10,
                            leading=12
                        )
                        
                        # Разбиваем текст на абзацы и добавляем в PDF
                        for line in text_content.split('\n'):
                            if line.strip():
                                # Экранируем специальные символы для reportlab
                                line = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                story.append(Paragraph(line, style))
                            else:
                                story.append(Spacer(1, 3*mm))
                        
                        doc.build(story)
                        exported_count += 1
                        
                    except ImportError:
                        if self.log_text:
                            self.log_text.append(f"  ⚠️  reportlab не установлен, пропуск: {os.path.basename(input_file)}")
                        failed_count += 1
                    except Exception as txt_error:
                        if self.log_text:
                            self.log_text.append(f"  ❌ Ошибка конвертации TXT: {txt_error}")
                        failed_count += 1
                
                else:
                    # Неподдерживаемый формат
                    if self.log_text:
                        self.log_text.append(f"  ⚠️  Неподдерживаемый формат: {os.path.basename(input_file)}")
                    failed_count += 1
                    
            except Exception as e:
                failed_count += 1
                if self.log_text:
                    self.log_text.append(f"  ❌ Ошибка: {os.path.basename(input_file)} - {e}")
        
        QApplication.restoreOverrideCursor()
        
        # Показываем результат
        msg = f"Экспорт завершен:\\n\\n"
        msg += f"✅ Успешно: {exported_count} файлов\\n"
        if failed_count > 0:
            msg += f"❌ Ошибок: {failed_count} файлов"
        
        QMessageBox.information(self, "Экспорт в PDF", msg)
        
        if self.log_text:
            self.log_text.append(f"✅ Экспорт входных файлов завершен: {exported_count}/{len(self.input_files)}")
    
    def on_toggle_auto_pdf_export(self, state: int):
        """Включение/выключение автоматического экспорта в PDF"""
        self.auto_export_pdf = bool(state)
        self.save_ui_preferences()
        if self.log_text:
            message = "📄 Автоматический экспорт в PDF включен" if self.auto_export_pdf else "📄 Автоматический экспорт в PDF отключен"
            self.log_text.append(message)
    
    def on_toggle_ai_classifier(self, state: int):
        """Включение/выключение AI-подсказок"""
        self.ai_classifier_enabled = bool(state)
        self.save_ui_preferences()
        
        # Обновляем статус
        self.update_ai_status()
        
        if self.log_text:
            message = "🤖 AI-подсказки включены" if self.ai_classifier_enabled else "🤖 AI-подсказки отключены"
            self.log_text.append(message)
            
            if self.ai_classifier_enabled:
                # Проверяем наличие API ключа
                from .ai_classifier import AIClassifierSettings
                settings = AIClassifierSettings()
                api_key = settings.get_api_key()
                
                if not api_key:
                    self.log_text.append("⚠️ Для использования AI-подсказок необходимо настроить API ключ")
                    self.log_text.append("   Нажмите '⚙️ Настройки AI' для конфигурации")
    
    def on_ai_auto_classify_clicked(self, checked: bool):
        """Обработчик клика на чекбокс автоматической AI классификации"""
        if self.log_text:
            self.log_text.append(f"🔧 DEBUG: clicked, checked={checked}")
        
        # Если пользователь пытается включить
        if checked:
            from .ai_classifier import AIClassifierSettings
            settings = AIClassifierSettings()
            
            if not settings.is_enabled():
                # AI отключен - показываем предупреждение
                if self.log_text:
                    self.log_text.append("🔧 DEBUG: AI отключен, показываем предупреждение")
                QMessageBox.warning(
                    self,
                    "AI не настроен",
                    "❌ AI классификатор отключен.\n\n"
                    "Для использования автоматической AI классификации:\n"
                    "1. Откройте меню 'Поиск PDF и AI' → 'Настройки API и AI'\n"
                    "2. Включите AI классификатор\n"
                    "3. Выберите провайдера (Claude, GPT или Ollama)\n"
                    "4. Укажите API ключ"
                )
                # Отменяем включение чекбокса
                self.ai_auto_classify_checkbox.setChecked(False)
                self.ai_auto_classify = False
                return
            
            provider = settings.get_provider()
            api_key = settings.get_api_key(provider)
            
            if not api_key:
                # Нет API ключа - показываем предупреждение
                if self.log_text:
                    self.log_text.append("🔧 DEBUG: Нет API ключа, показываем предупреждение")
                QMessageBox.warning(
                    self,
                    "API ключ не указан",
                    "❌ API ключ не настроен.\n\n"
                    "Для использования автоматической AI классификации:\n"
                    "1. Откройте меню 'Поиск PDF и AI' → 'Настройки API и AI'\n"
                    "2. Укажите API ключ для выбранного провайдера"
                )
                # Отменяем включение чекбокса
                self.ai_auto_classify_checkbox.setChecked(False)
                self.ai_auto_classify = False
                return
        
        # Если дошли до сюда, значит можно изменить состояние
        self.ai_auto_classify = checked
        self.save_ui_preferences()
        
        if self.log_text:
            if self.ai_auto_classify:
                self.log_text.append("🤖 Автоматическая AI-классификация включена")
                self.log_text.append("⚠️ ВСЕ неизвестные компоненты будут отправлены на классификацию через AI")
            else:
                self.log_text.append("🤖 Автоматическая AI-классификация отключена")
    
    def open_ai_settings(self):
        """Открывает диалог настроек AI"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QFormLayout, QComboBox, QLineEdit, QDialogButtonBox, QTextEdit, QLabel
        from .ai_classifier import AIClassifierSettings
        
        # Создаем диалог
        dialog = QDialog(self)
        dialog.setWindowTitle("⚙️ Настройки AI-подсказок")
        dialog.resize(600, 500)
        
        layout = QVBoxLayout(dialog)
        
        # Описание
        desc = QLabel(
            "Настройте провайдера AI и API ключи для автоматической классификации компонентов.\n"
            "Поддерживаются: Anthropic Claude, OpenAI GPT, Ollama (локальный)."
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Форма настроек
        form = QFormLayout()
        
        # Провайдер
        provider_combo = QComboBox()
        provider_combo.addItems(["Anthropic Claude", "OpenAI GPT", "Ollama (локальный)"])
        form.addRow("Провайдер AI:", provider_combo)
        
        # Загружаем текущие настройки
        settings = AIClassifierSettings()
        current_provider = settings.get_provider()
        provider_map = {
            "anthropic": 0,
            "openai": 1,
            "ollama": 2
        }
        provider_combo.setCurrentIndex(provider_map.get(current_provider, 0))
        
        # API ключи
        anthropic_key = QLineEdit()
        anthropic_key.setPlaceholderText("sk-ant-...")
        anthropic_key.setText(settings.get_api_key("anthropic"))
        anthropic_key.setEchoMode(QLineEdit.Password)
        form.addRow("Anthropic API Key:", anthropic_key)
        
        openai_key = QLineEdit()
        openai_key.setPlaceholderText("sk-...")
        openai_key.setText(settings.get_api_key("openai"))
        openai_key.setEchoMode(QLineEdit.Password)
        form.addRow("OpenAI API Key:", openai_key)
        
        ollama_url = QLineEdit()
        ollama_url.setPlaceholderText("http://localhost:11434")
        ollama_url.setText(settings.get_api_key("ollama"))
        form.addRow("Ollama URL:", ollama_url)
        
        # Модель
        model_input = QLineEdit()
        model_input.setPlaceholderText("По умолчанию (оставьте пустым)")
        model_input.setText(settings.get_model())
        form.addRow("Модель (опционально):", model_input)
        
        layout.addLayout(form)
        
        # Справка
        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setMaximumHeight(150)
        help_text.setHtml("""
<b>Справка:</b><br>
<b>Anthropic Claude:</b> Получите API ключ на <a href="https://console.anthropic.com/">console.anthropic.com</a><br>
<b>OpenAI GPT:</b> Получите API ключ на <a href="https://platform.openai.com/api-keys">platform.openai.com</a><br>
<b>Ollama:</b> Установите локально: <a href="https://ollama.ai/">ollama.ai</a><br><br>
<b>Модели по умолчанию:</b><br>
• Anthropic: claude-3-sonnet-20240229<br>
• OpenAI: gpt-4<br>
• Ollama: llama2<br>
        """)
        help_text.setOpenExternalLinks(True)
        layout.addWidget(help_text)
        
        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        # Показываем диалог
        if dialog.exec() == QDialog.Accepted:
            # Сохраняем настройки
            provider_reverse_map = {
                0: "anthropic",
                1: "openai",
                2: "ollama"
            }
            
            new_settings = {
                "enabled": self.ai_classifier_enabled,
                "provider": provider_reverse_map[provider_combo.currentIndex()],
                "model": model_input.text().strip(),
                "api_keys": {
                    "anthropic": anthropic_key.text().strip(),
                    "openai": openai_key.text().strip(),
                    "ollama": ollama_url.text().strip()
                },
                "auto_classify": getattr(self, 'ai_auto_classify', False),
                "confidence_threshold": "medium"
            }
            
            if settings.save_settings(new_settings):
                if self.log_text:
                    self.log_text.append("✅ Настройки AI сохранены")
                
                # Обновляем статус
                self.update_ai_status()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось сохранить настройки AI")
    
    def update_ai_status(self):
        """Обновляет статус AI в UI"""
        if not hasattr(self, 'ai_status_label'):
            return
        
        from .ai_classifier import AIClassifierSettings
        settings = AIClassifierSettings()
        
        if not settings.is_enabled():
            self.ai_status_label.setText("Статус: ⚪ Отключен")
            self.ai_status_label.setStyleSheet("color: #6c7086;")
            # Чекбокс остается активным, чтобы показать подсказку при клике
            return
        
        provider = settings.get_provider()
        api_key = settings.get_api_key(provider)
        
        if not api_key:
            self.ai_status_label.setText(f"Статус: 🟡 Не настроен")
            self.ai_status_label.setStyleSheet("color: #fab387;")
            # Чекбокс остается активным, чтобы показать подсказку при клике
        else:
            provider_names = {
                "anthropic": "Claude",
                "openai": "GPT",
                "ollama": "Ollama"
            }
            provider_name = provider_names.get(provider, provider)
            self.ai_status_label.setText(f"Статус: 🟢 Готов ({provider_name})")
            self.ai_status_label.setStyleSheet("color: #a6e3a1;")
    
    def _open_file(self, file_path: str):
        """Открывает файл в системном приложении"""
        try:
            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.Popen(['open', file_path])
            else:  # Linux
                subprocess.Popen(['xdg-open', file_path])
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось открыть файл:\n{e}")

    def save_ui_preferences(self):
        """Настройки UI НЕ сохраняются - приложение всегда открывается с настройками из config_qt.json"""
        # Метод оставлен для совместимости, но ничего не делает
        pass

    def reveal_in_file_manager(self, target_path: str, select: bool = True) -> bool:
        """Открывает системный проводник и при необходимости выделяет файл."""
        if not target_path:
            return False

        try:
            abs_path = os.path.abspath(target_path)
            system = platform.system()

            if system == 'Windows':
                if select and os.path.isfile(abs_path):
                    subprocess.Popen(f'explorer /select,"{abs_path}"')
                else:
                    folder = abs_path if os.path.isdir(abs_path) else os.path.dirname(abs_path)
                    subprocess.Popen(['explorer', folder])
            elif system == 'Darwin':
                if select and os.path.isfile(abs_path):
                    subprocess.Popen(['open', '-R', abs_path])
                else:
                    folder = abs_path if os.path.isdir(abs_path) else os.path.dirname(abs_path)
                    subprocess.Popen(['open', folder])
            else:
                folder = abs_path if os.path.isdir(abs_path) else os.path.dirname(abs_path)
                subprocess.Popen(['xdg-open', folder])

            return True
        except Exception as e:
            print(f"⚠️ Не удалось открыть проводник: {e}")

def main():
    """Точка входа для PySide6 приложения"""
    # Инициализируем конфигурационные файлы из шаблонов (если их нет)
    initialize_all_configs()
    
    # Импорты для настройки Qt
    from PySide6.QtGui import QFont, QGuiApplication
    
    # ========== HIGH DPI SUPPORT ДЛЯ MACOS RETINA ==========
    # Устанавливаем переменные окружения ДО импорта/создания QApplication
    # Это критично для правильной работы на Retina дисплеях
    import os as os_env
    if platform.system() == 'Darwin':  # macOS
        # Включаем автоматическое масштабирование для Retina
        os_env.environ['QT_AUTO_SCREEN_SCALE_FACTOR'] = '1'
        os_env.environ['QT_ENABLE_HIGHDPI_SCALING'] = '1'
        # Для Qt 6
        os_env.environ['QT_SCALE_FACTOR_ROUNDING_POLICY'] = 'PassThrough'
    
    # КРИТИЧНО: эти атрибуты должны быть установлены ДО создания QApplication!
    # Без них на macOS Retina шрифты будут выглядеть в 2 раза меньше
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    
    # Для Qt 6: используем PassThrough для правильного масштабирования на Retina
    try:
        if hasattr(Qt, 'HighDpiScaleFactorRoundingPolicy'):
            QApplication.setHighDpiScaleFactorRoundingPolicy(
                Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
            )
    except Exception:
        pass  # Для совместимости со старыми версиями Qt
    # ========================================================
    
    app = QApplication(sys.argv)

    # Устанавливаем имя приложения
    app.setApplicationName("BOM Categorizer")
    app.setOrganizationName("Kurein M.N.")

    # ========== УСТАНОВКА ГЛОБАЛЬНОГО ШРИФТА ДЛЯ MACOS RETINA ==========
    # Устанавливаем шрифт ДО создания виджетов, чтобы все виджеты
    # использовали правильный размер с самого начала
    if platform.system() == 'Darwin':  # macOS
        # Определяем размер для Retina (сопоставимый с другими macOS приложениями)
        try:
            screens = QGuiApplication.screens()
            if screens and screens[0].devicePixelRatio() >= 2:
                # Retina: используем 13pt (как в стандартных macOS приложениях)
                base_size = 13
            else:
                base_size = 12
        except:
            base_size = 13  # Для надежности
        
        # Устанавливаем глобальный шрифт для приложения
        app_font = QFont(get_system_font(), base_size)
        app.setFont(app_font)
        
        print(f"🔤 macOS: Установлен глобальный шрифт {get_system_font()} размером {base_size}pt")
    # ==================================================================

    # Создаем и показываем главное окно
    window = BOMCategorizerMainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()