# -*- coding: utf-8 -*-
"""
Модуль обработки файлов ТРУ и РКМ

Обрабатывает .xls файлы ТРУ и РКМ:
- Извлекает нужные столбцы
- Переименовывает столбцы
- Объединяет столбцы
- Сохраняет в новый .xlsx файл
"""

import os
import re
from typing import List, Dict, Optional, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def detect_file_type(filename: str) -> Optional[str]:
    """
    Определяет тип файла по имени
    Поддерживает кириллицу и латиницу, а также смешанные написания.
    
    Args:
        filename: Имя файла
        
    Returns:
        'tpy' если содержит вариации TRU/TRY/TPY
        'rkm' если содержит вариации RKM/PKM
        None если не определено
    """
    filename_lower = filename.lower()
    
    # Шаблоны для ТРУ (TRU/TRY/TPY)
    # Включают кириллицу 'тру' и латиницу 'try', 'tru', 'tpy'
    # Также учитываем визуально похожие написания
    tru_patterns = [
        'try', 'tpy', 'tru',  # Latin
        'тру',                # Cyrillic
        'тpy', 'tру', 'tpу'   # Mixed variations
    ]
    
    # Шаблоны для РКМ (RKM/PKM)
    # Включают кириллицу 'ркм' и латиницу 'rkm', 'pkm'
    rkm_patterns = [
        'rkm', 'pkm',         # Latin
        'ркм',                # Cyrillic
        'pкм', 'ркm'          # Mixed variations
    ]
    
    if any(pattern in filename_lower for pattern in tru_patterns):
        return 'tpy'
    elif any(pattern in filename_lower for pattern in rkm_patterns):
        return 'rkm'
    
    return None


def generate_output_filename(input_path: str, file_type: str) -> str:
    """
    Генерирует имя выходного файла
    
    Args:
        input_path: Путь к входному файлу
        file_type: Тип файла ('tpy' или 'rkm')
        
    Returns:
        Полный путь к выходному файлу
    """
    directory = os.path.dirname(input_path)
    filename = os.path.basename(input_path)
    name_without_ext = os.path.splitext(filename)[0]
    
    # Добавляем суффикс перед расширением
    suffix = f"_{file_type}"
    output_filename = f"{name_without_ext}{suffix}.xlsx"
    output_path = os.path.join(directory, output_filename)
    
    # Проверяем существование и добавляем счетчик если нужно
    if os.path.exists(output_path):
        counter = 1
        while True:
            output_filename = f"{name_without_ext}{suffix}_{counter}.xlsx"
            output_path = os.path.join(directory, output_filename)
            if not os.path.exists(output_path):
                break
            counter += 1
    
    return output_path


def _read_tru_file(input_path: str) -> Optional[pd.DataFrame]:
    """
    Читает один ТРУ файл и возвращает сырой DataFrame с нужными колонками.
    Поддерживает форматы .xls и .xlsx, ищет колонки по названию.
    """
    def _drop_tru_noise_rows(df_in: pd.DataFrame) -> pd.DataFrame:
        """
        Удаляет "мусорные" строки, которые иногда встречаются внутри ТРУ:
        - повтор заголовков посередине таблицы (например, строка где Артикул='Артикул', Наименование='Наименование', ...)
        - строки итогов (ИТОГО)

        Это важно, т.к. такие строки потом ломают конвертацию типов и матчинг.
        """
        if df_in is None or df_in.empty:
            return df_in

        df = df_in.copy()

        def norm(v) -> str:
            if v is None or pd.isna(v):
                return ""
            return str(v).strip().lower()

        # Список канонических заголовков, которые могут повторяться строкой данных
        canon = {
            'Артикул': 'артикул',
            'Наименование': 'наименование',
            'Количество': 'количество',
            'Цена': 'цена',
            'Стоимость': 'стоимость',
            'Ответственные': 'ответственные',
            'Код ОКПД': 'код окпд',
        }

        # Предикат: строка похожа на повтор заголовков
        def is_header_repeat(row) -> bool:
            hits = 0
            for col, canon_name in canon.items():
                if col in df.columns:
                    if norm(row.get(col, "")) == canon_name:
                        hits += 1
            # Достаточно 2 совпадений, но если "Артикул" совпал — это почти наверняка заголовок
            if 'Артикул' in df.columns and norm(row.get('Артикул', "")) == canon['Артикул']:
                return True
            return hits >= 2

        # Предикат: строка итогов
        def is_total_row(row) -> bool:
            # ИТОГО может встречаться в разных колонках, чаще в Наименовании/Артикуле
            for col in ['Артикул', 'Наименование']:
                if col in df.columns and norm(row.get(col, "")).startswith('итого'):
                    return True
            return False

        mask_repeat = df.apply(is_header_repeat, axis=1) if len(df) else pd.Series([], dtype=bool)
        mask_total = df.apply(is_total_row, axis=1) if len(df) else pd.Series([], dtype=bool)

        # Также удалим полностью пустые строки (все ключевые поля пустые)
        key_cols = [c for c in ['Артикул', 'Наименование', 'Количество', 'Цена', 'Стоимость'] if c in df.columns]
        if key_cols:
            # applymap() deprecated in pandas 2.2+, use per-column Series.map instead
            mask_blank = df[key_cols].apply(lambda s: s.map(lambda v: norm(v) == "")).all(axis=1)
        else:
            mask_blank = pd.Series([False] * len(df))

        df = df.loc[~(mask_repeat | mask_total | mask_blank)].reset_index(drop=True)
        return df

    try:
        ext = os.path.splitext(input_path)[1].lower()
        df_raw = None
        
        if ext == '.xls':
            # Старый формат Excel — используем xlrd
            import xlrd
            workbook = xlrd.open_workbook(input_path, encoding_override='cp1251')
            sheet = workbook.sheet_by_index(0)
            
            # Конвертируем данные в список списков
            data = []
            for row_idx in range(sheet.nrows):
                data.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])
                
            df_raw = pd.DataFrame(data)
            
        elif ext == '.xlsx':
            # Новый формат Excel — используем openpyxl через pandas
            # Читаем БЕЗ заголовка, чтобы найти его вручную
            df_raw = pd.read_excel(input_path, sheet_name=0, header=None, engine='openpyxl')
            
        else:
            print(f"Неподдерживаемый формат файла: {ext}")
            return None
        
        if df_raw is None or len(df_raw) < 1:
            return None

        # Ищем строку заголовка
        header_row_idx = -1
        
        # Ключевые слова для поиска заголовка
        keywords = ['наименование', 'артикул', 'количество', 'цена', 'стоимость', 'ответственные', 'код']
        
        for idx in range(min(20, len(df_raw))):
            row_str = df_raw.iloc[idx].astype(str).str.lower().str.strip()
            # Проверяем наличие хотя бы двух ключевых слов
            hits = sum(1 for kw in keywords if row_str.str.contains(kw, regex=False).any())
            
            has_name = row_str.str.contains('наименование').any() or row_str.str.contains('description').any()
            
            if hits >= 2 or has_name:
                header_row_idx = idx
                break
        
        df = None
        if header_row_idx >= 0:
            # Заголовок найден
            df = df_raw.iloc[header_row_idx+1:].copy()
            df.columns = df_raw.iloc[header_row_idx]
            df = df.reset_index(drop=True)
        else:
            # Заголовок НЕ найден — считаем, что данные начинаются с 0
            # Не присваиваем column names, оставляем int индексы
            df = df_raw.copy()

        # Нормализуем названия колонок для поиска
        col_mapping = {}
        cols = list(df.columns)
        cols_lower = [str(c).lower().strip() for c in cols]
        
        # --- Поиск Наименования ---
        # Приоритет: 1. Точное "Наименование" 2. "Наименование ИВП" 3. Содержит "Наименование" (но не "заказа")
        name_idx = -1
        
        # 1. Exact match
        try:
            name_idx = cols_lower.index('наименование')
        except ValueError:
            # 2. Exact match variants
            try:
                name_idx = cols_lower.index('наименование ивп')
            except ValueError:
                pass
        
        # 3. Fuzzy match (excluding bad words)
        if name_idx < 0:
            for i, c_low in enumerate(cols_lower):
                if 'наименование' in c_low and 'заказа' not in c_low and 'подразделение' not in c_low:
                     name_idx = i
                     break
                if 'description' in c_low:
                     name_idx = i
                     break
        
        if name_idx >= 0:
            col_mapping['Наименование'] = cols[name_idx]

        # --- Поиск Количества ---
        # Приоритет: 1. "Количество заявлено в БЕИ" 2. Точное "Количество" 3. Fuzzy
        qty_idx = -1
        
        # 1. Known long headers
        for target in ['количество заявлено в беи', 'количество (беи)']:
             try:
                 qty_idx = cols_lower.index(target)
                 if qty_idx >= 0: break
             except ValueError:
                 pass
        
        # 2. Exact common headers
        if qty_idx < 0:
            for target in ['количество', 'кол-во', 'qty', 'шт', 'шт.']:
                try:
                    qty_idx = cols_lower.index(target)
                    if qty_idx >= 0: break
                except ValueError:
                    pass
        
        # 3. Fuzzy match
        if qty_idx < 0:
            for i, c_low in enumerate(cols_lower):
                if 'количество' in c_low and 'деи' not in c_low: # Избегаем ДЕИ если есть выбор
                    qty_idx = i
                    break
                if 'qty' in c_low:
                    qty_idx = i
                    break
        
        if qty_idx >= 0:
            col_mapping['Количество'] = cols[qty_idx]

        # --- Поиск Цены ---
        price_idx = -1
        # 1. Exact or starts with
        for i, c_low in enumerate(cols_lower):
            if c_low in ['цена', 'price', 'цена без ндс']:
                price_idx = i
                break
        
        # 2. Contains
        if price_idx < 0:
            for i, c_low in enumerate(cols_lower):
                if ('цена' in c_low or 'price' in c_low) and 'сумма' not in c_low:
                    price_idx = i
                    break
        
        if price_idx >= 0:
            col_mapping['Цена'] = cols[price_idx]
            
        # --- Поиск Стоимости ---
        cost_idx = -1
        for i, c_low in enumerate(cols_lower):
            if c_low in ['стоимость', 'сумма', 'сумма без ндс']:
                cost_idx = i
                break
        if cost_idx < 0:
            for i, c_low in enumerate(cols_lower):
                if 'стоимость' in c_low or 'сумма' in c_low:
                    cost_idx = i
                    break
        
        if cost_idx >= 0:
            col_mapping['Стоимость'] = cols[cost_idx]

        # --- Поиск Артикула ---
        art_idx = -1
        for i, c_low in enumerate(cols_lower):
            if 'артикул' in c_low or c_low == 'код':
                art_idx = i
                break
        if art_idx >= 0:
            col_mapping['Артикул'] = cols[art_idx]

        # --- Поиск Код ОКПД ---
        okpd_idx = -1
        for i, c_low in enumerate(cols_lower):
            if 'окпд' in c_low:
                okpd_idx = i
                break
        if okpd_idx >= 0:
            col_mapping['Код ОКПД'] = cols[okpd_idx]
            
        # --- Ответственные ---
        # Может быть "Роль ответственного" или "Группа ответственных"
        # Нам нужно то, что было раньше как "group_resp"? 
        # В старом коде: _group_resp = data_df.iloc[:, 14] (Группа ответственных?)
        # В файле: Col 14 = "Группа ответственных", Col 11 = "Роль ответственного"
        # Также ищем 'Код группы ответственных' (Col 16)
        
        # Ищем группу 
        group_idx = -1
        for i, c_low in enumerate(cols_lower):
            if 'группа ответственных' in c_low:
                group_idx = i
                break
        
        code_resp_idx = -1
        for i, c_low in enumerate(cols_lower):
            if 'код группы ответственных' in c_low:
                code_resp_idx = i
                break
        
        
        # Проверяем наличие минимальных колонок
        if 'Наименование' not in col_mapping:
            # Fallback на индексы для совсем плохих случаев (как раньше)
            if df.shape[1] >= 2:
                # ВАЖНО: Не пропускаем строки, если заголовка не было!
                data_df = df # Используем данные как есть
                
                result_df = pd.DataFrame()
                result_df['Артикул'] = data_df.iloc[:, 0] if data_df.shape[1] > 0 else ''
                # Если 0-я пустая, берем 1-ю как наименование
                result_df['Наименование'] = data_df.iloc[:, 1] if data_df.shape[1] > 1 else ''
                
                # Fallback индексы
                result_df['Количество'] = data_df.iloc[:, 4] if data_df.shape[1] > 4 else ''
                result_df['Цена'] = data_df.iloc[:, 8] if data_df.shape[1] > 8 else ''
                result_df['Стоимость'] = data_df.iloc[:, 9] if data_df.shape[1] > 9 else ''
                result_df['_group_resp'] = data_df.iloc[:, 14].fillna('') if data_df.shape[1] > 14 else ''
                result_df['_code_resp'] = data_df.iloc[:, 16].fillna('') if data_df.shape[1] > 16 else ''
                result_df['Код ОКПД'] = data_df.iloc[:, 13] if data_df.shape[1] > 13 else ''

                return _drop_tru_noise_rows(result_df)
            else:
                return None
        
        # Создаём результирующий DataFrame
        result_df = pd.DataFrame()
        result_df['Артикул'] = df[col_mapping['Артикул']] if 'Артикул' in col_mapping else ''
        result_df['Наименование'] = df[col_mapping['Наименование']]
        result_df['Количество'] = df[col_mapping['Количество']] if 'Количество' in col_mapping else ''
        result_df['Цена'] = df[col_mapping['Цена']] if 'Цена' in col_mapping else ''
        result_df['Стоимость'] = df[col_mapping['Стоимость']] if 'Стоимость' in col_mapping else ''
        
        # Ответственные: берем из найденных колонок или пытаемся по индексам если заголовки есть но странные?
        # Лучше брать по имени
        if group_idx >= 0:
            result_df['_group_resp'] = df.iloc[:, group_idx]
        else:
             # Fallback index 14 "Группа ответственных"
             result_df['_group_resp'] = df.iloc[:, 14] if df.shape[1] > 14 else ''

        if code_resp_idx >= 0:
             result_df['_code_resp'] = df.iloc[:, code_resp_idx]
        else:
             # Fallback index 16 "Код группы ответственных"
             result_df['_code_resp'] = df.iloc[:, 16] if df.shape[1] > 16 else ''
             
        result_df['Код ОКПД'] = df[col_mapping['Код ОКПД']] if 'Код ОКПД' in col_mapping else ''

        return _drop_tru_noise_rows(result_df)
        
    except Exception as e:
        print(f"Ошибка чтения {input_path}: {e}")
        import traceback
        traceback.print_exc()
        return None

def process_tru_files_batch(input_paths: List[str], output_path: str) -> Tuple[bool, str]:
    """
    Обрабатывает несколько ТРУ файлов и сохраняет в один
    """
    try:
        all_dfs = []
        
        for path in input_paths:
            df = _read_tru_file(path)
            if df is not None:
                all_dfs.append(df)
        
        if not all_dfs:
            return False, "Не удалось прочитать ни одного файла"
            
        # Объединяем все DataFrame
        result_df = pd.concat(all_dfs, ignore_index=True)
        
        # Обработка колонки "Ответственные"
        # Конвертируем коды в int где возможно
        def format_code(val):
            try:
                if pd.isna(val): return ""
                return str(int(float(val)))
            except:
                return str(val)
        
        code_resp_formatted = result_df['_code_resp'].apply(format_code)
        
        # Объединяем: Код + пробел + Группа
        result_df['Ответственные'] = (code_resp_formatted + ' ' + result_df['_group_resp'].astype(str)).str.strip()
        
        # Очистка и сортировка
        # Очистка цен и количеств
        def parse_price(val):
            try:
                if isinstance(val, (int, float)):
                    return float(val)
                val_str = str(val).replace(',', '.').replace(' ', '')
                return float(val_str)
            except:
                return 0.0
                
        result_df['Количество'] = result_df['Количество'].apply(parse_price)
        result_df['Цена'] = result_df['Цена'].apply(parse_price)
        
        # Рассчитываем стоимость для сортировки (по формуле)
        result_df['Стоимость'] = result_df['Количество'] * result_df['Цена']
        
        # Сортировка
        # 1. По коду (числовому)
        result_df['_sort_key_code'] = pd.to_numeric(result_df['_code_resp'], errors='coerce').fillna(float('inf'))
        
        # Сортируем: Код, затем Цена
        result_df = result_df.sort_values(by=['_sort_key_code', 'Цена'], ascending=[True, True])
        
        # Удаляем временные колонки
        result_df = result_df.drop(columns=['_group_resp', '_code_resp', '_sort_key_code'])

        if 'Код ОКПД' not in result_df.columns:
            result_df['Код ОКПД'] = ''

        # Упорядочиваем колонки для вывода
        cols_order = ['Артикул', 'Наименование', 'Количество', 'Цена', 'Стоимость', 'Ответственные', 'Код ОКПД']
        result_df = result_df[cols_order]
        
        # Удаляем пустые строки
        result_df = result_df.dropna(how='all')
        
        # Сохраняем (логика с openpyxl такая же как раньше)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='ТРУ')
            
            worksheet = writer.sheets['ТРУ']
            
            # Стили
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            for row in worksheet.iter_rows(min_row=2):
                def _format_code_cell(val):
                    if val is None:
                        return val
                    if isinstance(val, bool):
                        return str(val)
                    if isinstance(val, int):
                        return str(val)
                    if isinstance(val, float):
                        if val.is_integer():
                            return str(int(val))
                        s = f"{val:.10f}".rstrip('0').rstrip('.')
                        return s
                    return str(val)

                row[0].value = _format_code_cell(row[0].value)
                row[0].alignment = center_align # Артикул
                row[0].number_format = '@'
                row[1].alignment = left_align # Наименование
                row[2].alignment = center_align # Количество
                
                # Цена
                row[3].alignment = center_align
                row[3].number_format = '#,##0.00'
                
                # Стоимость (Формула)
                row_idx = row[0].row
                cell = row[4]
                cell.value = f"=C{row_idx}*D{row_idx}"
                cell.alignment = center_align
                cell.number_format = '#,##0.00'
                
                row[5].alignment = left_align # Ответственные
                row[6].value = _format_code_cell(row[6].value)
                row[6].alignment = center_align # Код ОКПД
                row[6].number_format = '@'

            column_widths = {'A': 15, 'B': 50, 'C': 12, 'D': 15, 'E': 15, 'F': 40, 'G': 20}
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # Дополнительно фиксируем выравнивание всей колонки Код ОКПД по центру
            for cell in worksheet.iter_cols(min_col=7, max_col=7, min_row=2):
                for c in cell:
                    c.alignment = center_align
            
            # ИТОГО
            last_row = len(result_df) + 1
            total_row = last_row + 2
            
            total_label_cell = worksheet.cell(row=total_row, column=4, value="ИТОГО:")
            total_label_cell.font = Font(bold=True, size=12)
            total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
            
            total_value_cell = worksheet.cell(row=total_row, column=5, value=f"=SUM(E2:E{last_row})")
            total_value_cell.font = Font(bold=True, size=12)
            total_value_cell.alignment = Alignment(horizontal='center', vertical='center')
            total_value_cell.number_format = '#,##0.00'
            
            top_border = Border(top=Side(border_style="double", color="000000"))
            total_label_cell.border = top_border
            total_value_cell.border = top_border
            
            total_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            total_label_cell.fill = total_fill
            total_value_cell.fill = total_fill
            
        # Пересохранение через COM
        try:
            import platform
            if platform.system() == 'Windows':
                _resave_with_excel_com(output_path)
        except:
            pass
            
        return True, f"Объединено и обработано {len(all_dfs)} файлов, {len(result_df)} строк"
        
    except Exception as e:
        return False, f"Ошибка пакетной обработки: {str(e)}"

def process_rkm_file(input_path: str, output_path: str) -> Tuple[bool, str]:
    """
    Обрабатывает один РКМ файл
    
    Логика:
    1. Ищет заголовок таблицы
    2. Оставляет колонки: № п/п, Наименование, Цена, Затраты, Документы, Поставщик
    3. Фильтрует строки (начиная с 2.1.1 и т.д.)
    """
    try:
        # Читаем весь файл
        df_raw = pd.read_excel(input_path, header=None)
        
        # Фиксированные индексы колонок для PKM формата (на основе анализа заголовков)
        # План (правая часть): 17-22
        # Факт (левая часть): 6-15
        PLAN_COLS = {
            'Количество': 17,
            'Цена': 18,
            'Стоимость': 19,
            'Документы': 20,
            'Поставщик': 22
        }
        FACT_COLS = {
            'Количество': 6,
            'Цена': 8,
            'Стоимость': 10,
            'Документы': 12,
            'Поставщик': 14
        }
        
        # Поиск строки с номерами колонок (1, 2, 3...)
        header_row_idx = -1
        for idx, row in df_raw.head(20).iterrows():
            row_str = row.astype(str).tolist()
            digit_count = sum(1 for s in row_str if s.strip().isdigit() and len(s.strip()) <= 2)
            if digit_count > 10:
                header_row_idx = idx
                break
        
        if header_row_idx < 0:
            # Fallback: ищем "№ п/п"
            for idx, row in df_raw.head(10).iterrows():
                if '№' in str(row.iloc[0]).lower() or 'п/п' in str(row.iloc[0]).lower():
                    header_row_idx = idx
                    break
            if header_row_idx < 0:
                header_row_idx = 4  # Дефолт
        
        # Извлечение данных
        data = []
        
        def clean_float(val):
            if pd.isna(val) or str(val).lower() == 'nan': return 0.0
            try: return float(val)
            except:
                try: return float(str(val).replace(',','.').replace(' ','').replace('\xa0', ''))
                except: return 0.0

        def get_val(row, plan_col, fact_col, is_num=False):
            # Сначала пробуем План, потом Факт
            for col_idx in [plan_col, fact_col]:
                if col_idx < len(row):
                    val = row.iloc[col_idx]
                    if pd.isna(val) or str(val).lower() == 'nan': continue
                    if is_num:
                        f_v = clean_float(val)
                        if f_v > 0: return f_v
                    else:
                        s_v = str(val).strip()
                        if s_v and len(s_v) > 1: return s_v
            return 0.0 if is_num else ""

        for idx, row in df_raw.iloc[header_row_idx+1:].iterrows():
            try:
                # Колонка 0 = №, Колонка 1 = Наименование
                no_val = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
                name_val = str(row.iloc[1]).strip() if len(row) > 1 and not pd.isna(row.iloc[1]) else ""
                
                # Фильтрация: нужен номер вида X.X.X и непустое название
                if not name_val or name_val.lower() == 'nan':
                    continue
                if not no_val or not no_val[0].isdigit():
                    continue
                if not re.match(r'^[\d\.]+$', no_val):
                    continue
                # Пропускаем категории (номера вида "1." или "2.")
                if re.match(r'^\d+\.$', no_val):
                    continue
                
                item = {
                    '№': no_val,
                    'Наименование': name_val,
                    'Количество': get_val(row, PLAN_COLS['Количество'], FACT_COLS['Количество'], is_num=True),
                    'Цена': get_val(row, PLAN_COLS['Цена'], FACT_COLS['Цена'], is_num=True),
                    'Стоимость': get_val(row, PLAN_COLS['Стоимость'], FACT_COLS['Стоимость'], is_num=True),
                    'Документы': get_val(row, PLAN_COLS['Документы'], FACT_COLS['Документы']),
                    'Поставщик': get_val(row, PLAN_COLS['Поставщик'], FACT_COLS['Поставщик']),
                }
                
                # Пропускаем подзаголовки категорий (все числовые значения = 0)
                if item['Количество'] == 0 and item['Цена'] == 0 and item['Стоимость'] == 0:
                    continue
                
                # Чистка поставщика и документов
                item['Поставщик'] = " ".join(line.strip() for line in str(item['Поставщик']).split('\n') if line.strip())
                item['Поставщик'] = re.sub(r'\s*,\s*$', '', item['Поставщик'])
                item['Документы'] = " ".join(line.strip() for line in str(item['Документы']).split('\n') if line.strip())
                
                data.append(item)
            except Exception:
                continue
                
        if not data:
            return False, "Данные не найдены"
            
        result_df = pd.DataFrame(data)
        
        # 1. Clean description names (remove technical noise)
        def clean_description(name: str) -> str:
            if not name:
                return name
            # Split and join with spaces to replace newlines
            lines = [line.strip() for line in str(name).split('\n') if line.strip()]
            if not lines:
                return name
            
            clean_name = " ".join(lines)
            
            # Remove multi-line noise
            clean_name = re.sub(r'\s*,\s*$', '', clean_name)
            # Normalize multiple spaces to one
            clean_name = re.sub(r'\s+', ' ', clean_name)
            # Ensure space before parenthesis if missing
            clean_name = re.sub(r'([^\s])\(', r'\1 (', clean_name)
            
            # Fix common run-together Russian words and manufacturers
            keywords_to_fix = [
                'отладочная плата', 'плата инструментальная', 'передачи данных',
                'индикации', 'управления', 'питания',
                'Analog Devices', 'AnalogDevices', 'Coilcraft', 'Mini-Circuits', 
                'Texas Instruments', 'Maxim Integrated', 'MaximIntegrated',
                'STMicroelectronics', 'Avnet', 'Vishay', 'Yageo', 'Murata', 'TDK',
                'Samsung', 'Infineon', 'Microchip', 'Renesas', 'Broadcom', 
                'Qualcomm', 'Xilinx', 'Altera'
            ]
            
            for kw in keywords_to_fix:
                stuck_kw = kw.replace(' ', '')
                clean_name = re.sub(rf'([a-zA-Z0-9])({stuck_kw})', rf'\1 {kw}', clean_name, flags=re.IGNORECASE)
                if ' ' in kw:
                    clean_name = re.sub(rf'\b{stuck_kw}\b', kw, clean_name, flags=re.IGNORECASE)
            
            clean_name = re.sub(r'\s+', ' ', clean_name)
            return clean_name.strip()
        
        result_df['Наименование'] = result_df['Наименование'].apply(clean_description)
        
        # Helper to pick first non-empty string
        def first_non_empty(series):
            for v in series:
                s = str(v).strip()
                if s and s.lower() != 'nan':
                    return s
            return ""

        # 2. Deduplicate: Group by Наименование
        aggregation = {
            '№': 'first',
            'Цена': 'max',      # Берем максимальную цену если есть разброс
            'Количество': 'sum',
            'Стоимость': 'sum',
            'Документы': lambda x: '; '.join(sorted(set(str(v).strip() for v in x if str(v).strip() and str(v).lower() != 'nan'))),
            'Поставщик': first_non_empty,
        }
        
        # Group by cleaned name
        result_df = result_df.groupby('Наименование', as_index=False).agg(aggregation)
        
        # 3. Re-number rows sequentially
        result_df['№'] = range(1, len(result_df) + 1)
        
        # 4. Sort by row number
        result_df = result_df.sort_values('№')
        
        # Rename column to "шт."
        result_df = result_df.rename(columns={'Количество': 'шт.'})
        
        # Упорядочиваем колонки
        cols_order = ['№', 'Наименование', 'шт.', 'Цена', 'Стоимость', 'Документы', 'Поставщик']
        # Проверяем, все ли есть, если каких-то нет в result_df, добавляем пустые
        for c in cols_order:
            if c not in result_df.columns:
                result_df[c] = ""
                
        result_df = result_df[cols_order]
        
        # Сохранение (аналогично ТРУ)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='РКМ')
            
            worksheet = writer.sheets['РКМ']
            
            # Стили
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Apply widths
            # User request: Name -20% (60->48), Docs +40% (35->49), Provider -30% (45->31.5 -> 32)
            # Plus new "шт." column
            ws_widths = {'A': 8, 'B': 48, 'C': 10, 'D': 15, 'E': 15, 'F': 49, 'G': 32}
            for col_l, w in ws_widths.items():
                worksheet.column_dimensions[col_l].width = w
                
            for row in worksheet.iter_rows(min_row=2):
                row[0].alignment = center_align # No
                row[1].alignment = left_align # Name
                row[2].alignment = center_align # Qty
                row[3].alignment = center_align # Price
                row[3].number_format = '#,##0.00'
                row[4].alignment = center_align # Cost
                row[4].number_format = '#,##0.00'
                row[5].alignment = left_align # Docs
                row[6].alignment = left_align # Provider
            
            # === Add Grand Total Row ===
            last_data_row = len(result_df) + 1
            total_row = last_data_row + 2
            
            # Total label
            total_label_cell = worksheet.cell(row=total_row, column=4, value="ИТОГО:")
            total_label_cell.font = Font(bold=True, size=12)
            total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Total value (вычисленное значение, не формула - для корректного отображения в PDF)
            total_sum = result_df['Стоимость'].sum() if 'Стоимость' in result_df.columns else 0
            total_value_cell = worksheet.cell(row=total_row, column=5, value=total_sum)
            total_value_cell.font = Font(bold=True, size=12)
            total_value_cell.alignment = Alignment(horizontal='center', vertical='center')
            total_value_cell.number_format = '#,##0.00'
            
            # Styling for total row
            top_border = Border(top=Side(border_style="double", color="000000"))
            total_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            total_label_cell.border = top_border
            total_value_cell.border = top_border
            total_label_cell.fill = total_fill
            total_value_cell.fill = total_fill

        return True, f"РКМ обработан: {len(result_df)} строк (дедуплицировано)"


    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, f"Ошибка РКМ: {str(e)}"


def process_tru_rkm_files(file_paths: List[str], progress_callback=None) -> Dict[str, Dict[str, any]]:
    """
    Обрабатывает список ТРУ/РКМ файлов
    ТРУ файлы объединяются в один.
    """
    results = {}
    
    # Разделяем файлы по типам
    tru_files = []
    rkm_files = []
    
    for i, path in enumerate(file_paths):
        ft = detect_file_type(os.path.basename(path))
        if ft == 'tpy':
            tru_files.append(path)
        elif ft == 'rkm':
            rkm_files.append(path)
        else:
            results[path] = {'success': False, 'message': 'Unknown file type'}
            if progress_callback: progress_callback(i+1, len(file_paths), os.path.basename(path), False)
            
    # Обработка ТРУ файлов (по каждому файлу отдельно)
    if tru_files:
        for idx, path in enumerate(tru_files, start=1):
            output_path = generate_output_filename(path, 'tpy')
            if progress_callback:
                progress_callback(idx, len(file_paths), f"Обработка ТРУ: {os.path.basename(path)}", True)
            success, msg = process_tru_files_batch([path], output_path)
            results[path] = {
                'success': success,
                'output_path': output_path if success else None,
                'message': msg
            }
            
    # Обработка РКМ (пока поштучно, т.к. не реализовано)
    for path in rkm_files:
         output_path = generate_output_filename(path, 'rkm')
         success, msg = process_rkm_file(path, output_path)
         results[path] = {'success': success, 'output_path': output_path, 'message': msg}
         
    return results
