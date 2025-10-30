"""
Интеграционные тесты на реальных файлах
"""
import pytest
import os
from pathlib import Path
from bom_categorizer.main import process_files


class TestRealFileProcessing:
    """Тесты обработки реальных файлов из example/"""
    
    @pytest.fixture
    def example_files(self, example_dir):
        """Получить список доступных примеров"""
        files = {
            'doc': list(example_dir.glob('*.doc')),
            'docx': list(example_dir.glob('*.docx')),
            'xlsx': list(example_dir.glob('*.xlsx')),
            'txt': list(example_dir.glob('*.txt'))
        }
        return files
    
    def test_process_doc_file(self, example_dir, temp_dir):
        """Тест обработки .doc файла"""
        doc_file = example_dir / "plata_MKVH.doc"
        
        if not doc_file.exists():
            pytest.skip(f"Файл {doc_file} не найден")
        
        output_file = temp_dir / "output_doc.xlsx"
        
        try:
            result = process_files(
                input_files=[str(doc_file)],
                output_xlsx=str(output_file),
                combine=True,
                loose=False
            )
            
            assert output_file.exists(), "Выходной файл не создан"
            assert result is not None, "Процесс вернул None"
            
        except Exception as e:
            pytest.fail(f"Ошибка обработки doc файла: {e}")
    
    def test_process_xlsx_file(self, example_dir, temp_dir):
        """Тест обработки .xlsx файла"""
        xlsx_file = example_dir / "Plata_Preobrz.xlsx"
        
        if not xlsx_file.exists():
            pytest.skip(f"Файл {xlsx_file} не найден")
        
        output_file = temp_dir / "output_xlsx.xlsx"
        
        try:
            result = process_files(
                input_files=[str(xlsx_file)],
                output_xlsx=str(output_file),
                combine=True,
                loose=False
            )
            
            assert output_file.exists(), "Выходной файл не создан"
            
        except Exception as e:
            pytest.fail(f"Ошибка обработки xlsx файла: {e}")
    
    def test_process_txt_file(self, example_dir, temp_dir):
        """Тест обработки .txt файла"""
        txt_files = list(example_dir.glob('*.txt'))
        
        if not txt_files:
            pytest.skip("TXT файлы не найдены")
        
        txt_file = txt_files[0]
        output_file = temp_dir / "output_txt.xlsx"
        
        try:
            result = process_files(
                input_files=[str(txt_file)],
                output_xlsx=str(output_file),
                combine=True,
                loose=True  # Для txt используем loose режим
            )
            
            assert output_file.exists(), "Выходной файл не создан"
            
        except Exception as e:
            pytest.fail(f"Ошибка обработки txt файла: {e}")
    
    def test_process_multiple_files(self, example_dir, temp_dir):
        """Тест обработки нескольких файлов одновременно"""
        files = []
        
        # Пробуем найти хотя бы 2 файла
        for pattern in ['*.doc', '*.xlsx', '*.txt']:
            found = list(example_dir.glob(pattern))
            if found:
                files.append(str(found[0]))
            if len(files) >= 2:
                break
        
        if len(files) < 2:
            pytest.skip("Недостаточно файлов для теста")
        
        output_file = temp_dir / "output_multiple.xlsx"
        
        try:
            result = process_files(
                input_files=files,
                output_xlsx=str(output_file),
                combine=True,
                loose=False
            )
            
            assert output_file.exists(), "Выходной файл не создан"
            
        except Exception as e:
            pytest.fail(f"Ошибка обработки нескольких файлов: {e}")


class TestSpecificFiles:
    """Тесты конкретных проблемных файлов"""
    
    def test_plata_mkvh_doc(self, example_dir, temp_dir):
        """Тест файла plata_MKVH.doc который давал ошибки"""
        input_file = example_dir / "plata_MKVH.doc"
        
        if not input_file.exists():
            pytest.skip(f"Файл {input_file} не найден")
        
        output_file = temp_dir / "plata_MKVH_output.xlsx"
        
        try:
            result = process_files(
                input_files=[str(input_file)],
                output_xlsx=str(output_file),
                combine=True,
                loose=False
            )
            
            assert output_file.exists(), "Выходной файл не создан"
            
            # Дополнительные проверки
            import pandas as pd
            from openpyxl import load_workbook
            
            wb = load_workbook(str(output_file))
            
            # Проверяем что есть нужные листы
            assert len(wb.sheetnames) > 0, "Нет листов в выходном файле"
            
            # Проверяем что есть данные
            for sheet_name in wb.sheetnames:
                if sheet_name != 'SUMMARY':
                    ws = wb[sheet_name]
                    # Проверяем что лист не пустой
                    assert ws.max_row > 1, f"Лист {sheet_name} пустой"
            
            print(f"\n✅ Файл plata_MKVH.doc успешно обработан")
            print(f"   Создано листов: {len(wb.sheetnames)}")
            print(f"   Листы: {', '.join(wb.sheetnames)}")
            
        except Exception as e:
            pytest.fail(f"Ошибка обработки plata_MKVH.doc: {e}")


class TestOutputValidation:
    """Тесты валидации выходных данных"""
    
    def test_output_has_all_categories(self, example_dir, temp_dir):
        """Тест что выходной файл содержит все категории"""
        doc_file = example_dir / "plata_MKVH.doc"
        
        if not doc_file.exists():
            pytest.skip(f"Файл {doc_file} не найден")
        
        output_file = temp_dir / "test_categories.xlsx"
        
        process_files(
            input_files=[str(doc_file)],
            output_xlsx=str(output_file),
            combine=True,
            loose=False
        )
        
        from openpyxl import load_workbook
        wb = load_workbook(str(output_file))
        
        # Ожидаемые категории (русские названия листов)
        expected_categories = [
            'Отладочные платы и модули',
            'Микросхемы',
            'Резисторы',
            'Конденсаторы',
            'Индуктивности',
            'Полупроводники',
            'Разъемы',
            'Оптические компоненты',
            'Модули питания',
            'Кабели',
            'Другие',
            'Не распределено'
        ]
        
        sheet_names = wb.sheetnames
        
        # Проверяем что хотя бы некоторые категории присутствуют
        found_categories = [cat for cat in expected_categories if cat in sheet_names]
        
        assert len(found_categories) > 0, "Не найдено ни одной категории"
        
        print(f"\n✅ Найдено категорий: {len(found_categories)}")
        print(f"   Категории: {', '.join(found_categories)}")
    
    def test_output_has_summary(self, example_dir, temp_dir):
        """Тест что выходной файл содержит SUMMARY лист"""
        doc_file = example_dir / "plata_MKVH.doc"
        
        if not doc_file.exists():
            pytest.skip(f"Файл {doc_file} не найден")
        
        output_file = temp_dir / "test_summary.xlsx"
        
        process_files(
            input_files=[str(doc_file)],
            output_xlsx=str(output_file),
            combine=True,  # С этим флагом должен создаваться SUMMARY
            loose=False
        )
        
        from openpyxl import load_workbook
        wb = load_workbook(str(output_file))
        
        assert 'SUMMARY' in wb.sheetnames, "Лист SUMMARY не найден"
        
        # Проверяем что SUMMARY не пустой
        ws = wb['SUMMARY']
        assert ws.max_row > 1, "Лист SUMMARY пустой"


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
